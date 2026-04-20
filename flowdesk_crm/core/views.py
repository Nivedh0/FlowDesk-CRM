import json
import calendar
import re
from types import SimpleNamespace
from urllib import request
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.mail import EmailMultiAlternatives
from django.core.validators import validate_email
from django.shortcuts import get_object_or_404, render,redirect
from django.urls import reverse
from .models import Status, Source, Department, Course, UserProfile, StudentEnquiry, StudentCourse, FeeInstallment, TrainerSpecialization
from .models import StudentEnquiry, Department, Course, Status, Source, UserProfile, LeadActivity, LeadTask, FollowUp, ExamPerformance, Exam, Student, Batch, Fee, Payment, PendingMail, MailLog, DismissedNotification
from django.contrib.auth.decorators import login_required
from django.contrib.auth import views as auth_views
from django.contrib.auth.models import User
from django.contrib.auth.validators import UnicodeUsernameValidator
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout as auth_logout
from django.db.models import ProtectedError, Q, OuterRef, Subquery, F, Prefetch, Sum, ExpressionWrapper, DecimalField, Exists
from django.http import JsonResponse
from django.db import IntegrityError, transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db.models import ProtectedError
from .models import Topic,Module,TopicProgress,QuestionPaper,QuestionPart
from .decorators import role_required
from .forms import UsernameAwarePasswordResetForm
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.views.decorators.cache import never_cache
from datetime import timedelta



# Create your views here.
def base(request):
    return render(request, 'base.html')


class UsernamePasswordResetView(auth_views.PasswordResetView):
    form_class = UsernameAwarePasswordResetForm
    template_name = 'auth/password_reset_form.html'
    email_template_name = 'auth/password_reset_email.txt'
    subject_template_name = 'auth/password_reset_subject.txt'

    def get_initial(self):
        initial = super().get_initial()
        username = (self.request.GET.get('username') or '').strip()
        initial['username'] = username

        if username:
            try:
                user = User.objects.get(username__iexact=username)
            except User.DoesNotExist:
                return initial

            if user.email:
                initial['email'] = user.email

        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        username = (self.request.GET.get('username') or '').strip()
        email = form.initial.get('email')

        form.fields['username'].widget.attrs.update({
            'class': 'form-control',
            'placeholder': 'Enter your username',
            'autofocus': True,
        })
        form.fields['email'].widget.attrs.update({
            'class': 'form-control',
            'placeholder': 'Linked email will appear here',
        })

        if username:
            form.fields['username'].widget.attrs['value'] = form.initial.get('username', username)

        if email:
            form.fields['email'].widget.attrs['readonly'] = True

        return form


def _normalize_title_key(value):
    return re.sub(r"\s+", "", (value or "")).casefold()


def _normalize_source_name(value):
    return re.sub(r"\s+", " ", (value or "").strip())


def _get_source_by_input(value):
    normalized = _normalize_source_name(value)
    if not normalized:
        return None
    return Source.objects.filter(source_name__iexact=normalized).first()


def _get_cre_followup_context(user_profile):
    followups_qs = FollowUp.objects.select_related(
        'lead__status', 'assigned_to__user'
    ).filter(assigned_to=user_profile).order_by('-followup_date', '-followup_time')

    today = timezone.localdate()
    selected_followup_date = today
    display_month = today.replace(day=1)

    calendar_followups = [
        {
            'id': followup.id,
            'title': followup.title,
            'date': followup.followup_date.isoformat(),
            'time': followup.followup_time.strftime('%I:%M %p') if followup.followup_time else '',
            'status': followup.status,
            'status_display': followup.get_status_display(),
            'notes': followup.notes or '',
            'lead_id': followup.lead.id,
            'lead_name': followup.lead.full_name,
            'lead_mobile': followup.lead.mobile or '',
            'lead_email': followup.lead.email or '',
        }
        for followup in followups_qs.order_by('followup_date', 'followup_time', 'created_at')
    ]

    cre_fees_qs = Fee.objects.select_related('student', 'course').all()

    pending_expr = ExpressionWrapper(
        F('total_fee') - F('paid_amount'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    cre_fees_with_pending = cre_fees_qs.annotate(calc_pending=pending_expr)
    pending_fees_qs = cre_fees_with_pending.filter(calc_pending__gt=0)
    paid_fees_qs = cre_fees_with_pending.filter(calc_pending__lte=0)
    overdue_fees_qs = pending_fees_qs.filter(due_date__lt=today)
    due_soon_to = today + timedelta(days=7)
    due_soon_fees_qs = pending_fees_qs.filter(due_date__gte=today, due_date__lte=due_soon_to)

    payment_pending_amount = pending_fees_qs.aggregate(total=Sum('calc_pending')).get('total') or 0
    payment_overdue_amount = overdue_fees_qs.aggregate(total=Sum('calc_pending')).get('total') or 0
    payment_due_soon_amount = due_soon_fees_qs.aggregate(total=Sum('calc_pending')).get('total') or 0
    payment_collected_amount = cre_fees_qs.aggregate(total=Sum('paid_amount')).get('total') or 0

    total_fees_count = cre_fees_qs.count()
    payment_overdue_count = overdue_fees_qs.count()
    payment_due_soon_count = due_soon_fees_qs.count()
    payment_paid_count = paid_fees_qs.count()
    payment_collection_rate = round((payment_paid_count / total_fees_count * 100), 2) if total_fees_count > 0 else 0
    followup_completion_rate = round(
        (followups_qs.filter(status='completed').count() / followups_qs.count() * 100),
        2
    ) if followups_qs.count() > 0 else 0
    if payment_overdue_count > 0:
        payment_important_title = 'Overdue Payments'
        payment_important_count = payment_overdue_count
        payment_important_amount = payment_overdue_amount
        payment_important_note = 'Immediate follow-up required'
    elif payment_due_soon_count > 0:
        payment_important_title = 'Due This Week'
        payment_important_count = payment_due_soon_count
        payment_important_amount = payment_due_soon_amount
        payment_important_note = 'Upcoming payment due soon'
    else:
        payment_important_title = 'No Urgent Dues'
        payment_important_count = 0
        payment_important_amount = 0
        payment_important_note = 'All student payments are stable'

    return {
        'total_followups_today': followups_qs.filter(followup_date=today).count(),
        'completed_followups': followups_qs.filter(followup_date=today, status='completed').count(),
        'total_followups_count': followups_qs.count(),
        'completed_followups_count': followups_qs.filter(status='completed').count(),
        'pending_followups_count': followups_qs.filter(status='pending').count(),
        'cancelled_followups_count': followups_qs.filter(status='cancelled').count(),
        'all_pending_followups': followups_qs.filter(status='pending').order_by('followup_date', 'followup_time', 'created_at'),
        'all_completed_followups': followups_qs.filter(status='completed').order_by('-followup_date', '-followup_time', '-created_at'),
        'today': today,
        'selected_followup_date': selected_followup_date,
        'display_month': display_month,
        'calendar_followups_json': calendar_followups,
        'payment_total_fees_count': total_fees_count,
        'payment_pending_count': pending_fees_qs.count(),
        'payment_overdue_count': payment_overdue_count,
        'payment_due_soon_count': payment_due_soon_count,
        'payment_paid_count': payment_paid_count,
        'payment_collection_rate': payment_collection_rate,
        'followup_completion_rate': followup_completion_rate,
        'payment_pending_amount': payment_pending_amount,
        'payment_overdue_amount': payment_overdue_amount,
        'payment_due_soon_amount': payment_due_soon_amount,
        'payment_collected_amount': payment_collected_amount,
        'payment_important_title': payment_important_title,
        'payment_important_count': payment_important_count,
        'payment_important_amount': payment_important_amount,
        'payment_important_note': payment_important_note,
    }


@login_required
def dashboard(request):
    from django.db.models import Sum, Count, Q
    
    user_profile = request.user.userprofile
    role = user_profile.role
    
    context = {'role': role}
    
    if role == 'admin':
        lead_range = (request.GET.get('lead_range') or '').strip().lower()
        allowed_lead_ranges = {'today', 'weekly', 'monthly', 'yearly', 'custom'}
        if lead_range not in allowed_lead_ranges:
            lead_range = ''

        today = timezone.localdate()
        lead_custom_from = parse_date(request.GET.get('lead_from') or '')
        lead_custom_to = parse_date(request.GET.get('lead_to') or '')

        if lead_range == 'weekly':
            lead_date_from = today - timedelta(days=today.weekday())
            lead_date_to = lead_date_from + timedelta(days=6)
            lead_range_label = "This Week"
        elif lead_range == 'monthly':
            lead_date_from = today.replace(day=1)
            if lead_date_from.month == 12:
                next_month_start = lead_date_from.replace(year=lead_date_from.year + 1, month=1, day=1)
            else:
                next_month_start = lead_date_from.replace(month=lead_date_from.month + 1, day=1)
            lead_date_to = next_month_start - timedelta(days=1)
            lead_range_label = "This Month"
        elif lead_range == 'yearly':
            lead_date_from = today.replace(month=1, day=1)
            lead_date_to = today.replace(month=12, day=31)
            lead_range_label = "This Year"
        elif lead_range == 'custom':
            lead_date_from = lead_custom_from or today
            lead_date_to = lead_custom_to or lead_date_from
            if lead_date_from > lead_date_to:
                lead_date_from, lead_date_to = lead_date_to, lead_date_from
            lead_range_label = "Custom Range"
        elif lead_range == 'today':
            lead_date_from = today
            lead_date_to = today
            lead_range_label = "Today"
        else:
            lead_date_from = None
            lead_date_to = None
            lead_range_label = "All Leads"

        if lead_date_from and lead_date_to:
            lead_period_display = (
                lead_date_from.strftime('%d %b %Y')
                if lead_date_from == lead_date_to
                else f"{lead_date_from.strftime('%d %b %Y')} - {lead_date_to.strftime('%d %b %Y')}"
            )
            filtered_leads = StudentEnquiry.objects.filter(
                enquiry_date__gte=lead_date_from,
                enquiry_date__lte=lead_date_to,
            )
        else:
            lead_period_display = "Full Dataset"
            filtered_leads = StudentEnquiry.objects.all()

        # Lead-list based totals
        total_leads = filtered_leads.count()
        converted_leads = filtered_leads.filter(
            status__status_name__iexact='enrolled'
        ).count()
        
        # Total Students
        total_students = Student.objects.count()
        
        lead_statuses = filtered_leads.values('status__status_name').annotate(count=Count('id'))
        
        # Lead status data for chart
        lead_status_labels = []
        lead_status_counts = []
        for status in lead_statuses:
            lead_status_labels.append(status['status__status_name'] or 'Unknown')
            lead_status_counts.append(status['count'])
        
        active_students = Student.objects.filter(status='active').count()
        enrolled_students = Student.objects.filter(status='enrolled').count()
        completed_students = Student.objects.filter(status='completed').count()
        
        # Revenue
        total_revenue = Fee.objects.aggregate(total=Sum('total_fee'))['total'] or 0
        collected_revenue = Fee.objects.aggregate(collected=Sum('paid_amount'))['collected'] or 0
        pending_revenue = total_revenue - collected_revenue
        
        # Active Batches
        active_batches = Batch.objects.filter(status='Active').count()
        batches = Batch.objects.count()
        
        # Latest Leads
        latest_leads = filtered_leads.select_related('status').order_by('-enquiry_date', '-created_at')[:5]
        
        # Latest Students
        latest_students = Student.objects.prefetch_related('student_courses__course').order_by('-id')[:5]
        
        # Mini Cards Data
        total_count = total_leads
        today_count = filtered_leads.filter(enquiry_date=today).count()
        followup_count = filtered_leads.filter(followup_date__isnull=False).count()
        enrolled_count = converted_leads

        # Hot leads by lead type within the selected range
        seven_days_ago = today - timedelta(days=7)
        hot_leads_count = filtered_leads.filter(lead_type='hot').count()

        # Conversion rate
        converted_count = round((enrolled_count / total_count * 100), 2) if total_count > 0 else 0

        # Calculate percentage changes for mini cards
        # Total Inquiry: Compare the selected range with the immediately previous range of equal length.
        if lead_date_from and lead_date_to:
            selected_range_days = (lead_date_to - lead_date_from).days + 1
            previous_range_end = lead_date_from - timedelta(days=1)
            previous_range_start = previous_range_end - timedelta(days=selected_range_days - 1)

            previous_period_count = StudentEnquiry.objects.filter(
                enquiry_date__gte=previous_range_start,
                enquiry_date__lte=previous_range_end,
            ).count()
            total_inquiry_change = round(
                ((total_count - previous_period_count) / previous_period_count * 100), 2
            ) if previous_period_count > 0 else 0
        else:
            total_inquiry_change = 0

        # Today Leads: Compare today vs average of last 7 days
        last_7_days_count = StudentEnquiry.objects.filter(
            enquiry_date__gte=seven_days_ago,
            enquiry_date__lt=today
        ).count()
        avg_daily_leads = round(last_7_days_count / 7, 2) if last_7_days_count > 0 else 1
        if lead_date_from and lead_date_to and lead_date_from <= today <= lead_date_to and avg_daily_leads > 0:
            today_leads_change = round(((today_count - avg_daily_leads) / avg_daily_leads * 100), 2)
        else:
            today_leads_change = 0
        
        # With Follow-up: Percentage of leads with follow-ups
        followup_percentage = round((followup_count / total_count * 100), 2) if total_count > 0 else 0
        
        # Hot Leads: Percentage of hot leads vs total
        hot_leads_percentage = round((hot_leads_count / total_count * 100), 2) if total_count > 0 else 0

        cre_range = (request.GET.get('cre_range') or 'today').strip().lower()
        allowed_cre_ranges = {'today', 'weekly', 'monthly', 'yearly', 'custom'}
        if cre_range not in allowed_cre_ranges:
            cre_range = 'today'

        cre_followup_date = parse_date(request.GET.get('cre_followup_date') or '') or today
        custom_from = parse_date(request.GET.get('cre_from') or '')
        custom_to = parse_date(request.GET.get('cre_to') or '')

        if cre_range == 'weekly':
            cre_date_from = today - timedelta(days=today.weekday())
            cre_date_to = cre_date_from + timedelta(days=6)
            cre_range_label = "This Week"
        elif cre_range == 'monthly':
            cre_date_from = today.replace(day=1)
            if cre_date_from.month == 12:
                next_month_start = cre_date_from.replace(year=cre_date_from.year + 1, month=1, day=1)
            else:
                next_month_start = cre_date_from.replace(month=cre_date_from.month + 1, day=1)
            cre_date_to = next_month_start - timedelta(days=1)
            cre_range_label = "This Month"
        elif cre_range == 'yearly':
            cre_date_from = today.replace(month=1, day=1)
            cre_date_to = today.replace(month=12, day=31)
            cre_range_label = "This Year"
        elif cre_range == 'custom':
            cre_date_from = custom_from or today
            cre_date_to = custom_to or cre_date_from
            if cre_date_from > cre_date_to:
                cre_date_from, cre_date_to = cre_date_to, cre_date_from
            cre_range_label = "Custom Range"
        else:
            cre_date_from = cre_followup_date
            cre_date_to = cre_followup_date
            cre_range_label = "Today"

        cre_period_display = (
            cre_date_from.strftime('%d %b %Y')
            if cre_date_from == cre_date_to
            else f"{cre_date_from.strftime('%d %b %Y')} - {cre_date_to.strftime('%d %b %Y')}"
        )
        cre_profiles = UserProfile.objects.select_related('user').filter(role='cre').order_by(
            'user__first_name', 'user__username'
        )
        cre_followup_stats = FollowUp.objects.filter(
            assigned_to__in=cre_profiles,
            followup_date__gte=cre_date_from,
            followup_date__lte=cre_date_to,
        ).values('assigned_to_id').annotate(
            total=Count('id'),
            pending=Count('id', filter=Q(status='pending')),
            completed=Count('id', filter=Q(status='completed')),
            cancelled=Count('id', filter=Q(status='cancelled')),
        )
        cre_followup_stats_map = {row['assigned_to_id']: row for row in cre_followup_stats}
        cre_followup_summaries = []
        for cre in cre_profiles:
            stats = cre_followup_stats_map.get(cre.id, {})
            cre_followup_summaries.append({
                'id': cre.id,
                'name': cre.user.get_full_name() or cre.user.username,
                'mobile': cre.mobile,
                'total': stats.get('total', 0),
                'pending': stats.get('pending', 0),
                'completed': stats.get('completed', 0),
                'cancelled': stats.get('cancelled', 0),
            })
        cre_followups_total = sum(item['total'] for item in cre_followup_summaries)
        active_cre_count = sum(1 for item in cre_followup_summaries if item['total'] > 0)
        
        context.update({
            'lead_range': lead_range,
            'lead_range_label': lead_range_label,
            'lead_from': lead_date_from,
            'lead_to': lead_date_to,
            'lead_period_display': lead_period_display,
            'total_leads': total_leads,
            'converted_leads': converted_leads,
            'lead_statuses': lead_statuses,
            'lead_status_labels': lead_status_labels,
            'lead_status_counts': lead_status_counts,
            'total_students': total_students,
            'active_students': active_students,
            'enrolled_students': enrolled_students,
            'completed_students': completed_students,
            'total_revenue': total_revenue,
            'collected_revenue': collected_revenue,
            'pending_revenue': pending_revenue,
            'active_batches': active_batches,
            'batches': batches,
            'latest_leads': latest_leads,
            'latest_students': latest_students,
            # Mini Cards
            'total_count': total_count,
            'today_count': today_count,
            'followup_count': followup_count,
            'enrolled_count': enrolled_count,
            'hot_leads_count': hot_leads_count,
            'converted_count': converted_count,
            # Mini Cards Percentages
            'total_inquiry_change': total_inquiry_change,
            'today_leads_change': today_leads_change,
            'followup_percentage': followup_percentage,
            'hot_leads_percentage': hot_leads_percentage,
            'cre_range': cre_range,
            'cre_range_label': cre_range_label,
            'cre_followup_date': cre_followup_date,
            'cre_from': cre_date_from,
            'cre_to': cre_date_to,
            'cre_period_display': cre_period_display,
            'cre_followup_summaries': cre_followup_summaries,
            'cre_count': len(cre_followup_summaries),
            'active_cre_count': active_cre_count,
            'cre_followups_total': cre_followups_total,
        })
        
    
    elif role == 'trainer':
        from .models import SessionUpdate, Assignment, AssignmentSubmission

        # My Active Batches
        my_batches = Batch.objects.filter(
            trainer=user_profile,
            status='Active'
        ).select_related('course', 'department').order_by('batch_name')
        active_batches_count = my_batches.count()

        today = timezone.localdate()
        week_ahead = today + timedelta(days=7)

        # Total students in trainer's active batches
        student_counts = StudentCourse.objects.filter(
            batch__in=my_batches,
            course=F('batch__course')
        ).values('batch_id').annotate(total=Count('student', distinct=True))
        student_counts_map = {row['batch_id']: row['total'] for row in student_counts}
        total_students = sum(student_counts_map.values())

        # Topic totals and progress are tracked per batch-course
        course_ids = list(my_batches.values_list('course_id', flat=True))
        topic_counts = Topic.objects.filter(
            module__course_id__in=course_ids
        ).values('module__course_id').annotate(total=Count('id'))
        topic_counts_map = {row['module__course_id']: row['total'] for row in topic_counts}

        topic_progress = TopicProgress.objects.filter(
            batch__in=my_batches
        ).values('batch_id').annotate(
            completed=Count('id', filter=Q(status='completed')),
            partial=Count('id', filter=Q(status='partial')),
        )
        topic_progress_map = {row['batch_id']: row for row in topic_progress}

        total_topic_slots = 0
        completed_topic_slots = 0
        partial_topic_slots = 0
        trainer_batches = []

        for batch in my_batches:
            batch_total_topics = topic_counts_map.get(batch.course_id, 0)
            progress_stats = topic_progress_map.get(batch.id, {})
            batch_completed_topics = progress_stats.get('completed', 0)
            batch_partial_topics = progress_stats.get('partial', 0)
            total_topic_slots += batch_total_topics
            completed_topic_slots += batch_completed_topics
            partial_topic_slots += batch_partial_topics

            trainer_batches.append({
                'id': batch.id,
                'batch_name': batch.batch_name,
                'course_name': batch.course.course_name if batch.course else '-',
                'department_name': batch.department.department_name if batch.department else '-',
                'students_count': student_counts_map.get(batch.id, 0),
                'start_date': batch.start_date,
                'end_date': batch.end_date,
                'mode': batch.mode,
                'completed_topics': batch_completed_topics,
                'partial_topics': batch_partial_topics,
                'total_topics': batch_total_topics,
                'progress_percent': round((batch_completed_topics / batch_total_topics) * 100, 1) if batch_total_topics else 0,
            })

        overall_syllabus_completion = round(
            (completed_topic_slots / total_topic_slots) * 100, 1
        ) if total_topic_slots else 0

        today_sessions = SessionUpdate.objects.filter(
            batch__in=my_batches,
            session_date=today
        ).select_related('batch', 'batch__course').order_by('batch__batch_name')
        today_sessions_count = today_sessions.count()
        attendance_marked_today = today_sessions.filter(attendance_marked=True).count()

        recent_sessions = SessionUpdate.objects.filter(
            batch__in=my_batches
        ).select_related('batch', 'batch__course').order_by('-session_date', 'batch__batch_name')[:5]

        assignments_due_this_week = Assignment.objects.filter(
            batch__in=my_batches,
            due_date__gte=today,
            due_date__lte=week_ahead
        ).select_related('batch', 'batch__course').order_by('due_date')

        upcoming_exams = Exam.objects.filter(
            batch__in=my_batches,
            exam_date__gte=today,
            exam_date__lte=week_ahead
        ).select_related('batch', 'batch__course').order_by('exam_date')

        pending_submissions_count = AssignmentSubmission.objects.filter(
            assignment__batch__in=my_batches,
            status__in=['pending', 'not_submitted']
        ).count()

        context.update({
            'active_batches_count': active_batches_count,
            'total_students': total_students,
            'my_batches': my_batches,
            'trainer_batches': trainer_batches,
            'overall_syllabus_completion': overall_syllabus_completion,
            'total_topic_slots': total_topic_slots,
            'completed_topic_slots': completed_topic_slots,
            'partial_topic_slots': partial_topic_slots,
            'today_sessions_count': today_sessions_count,
            'attendance_marked_today': attendance_marked_today,
            'assignments_due_this_week_count': assignments_due_this_week.count(),
            'upcoming_exams_count': upcoming_exams.count(),
            'pending_submissions_count': pending_submissions_count,
            'recent_sessions': recent_sessions,
            'upcoming_exams': upcoming_exams[:5],
            'assignments_due_this_week': assignments_due_this_week[:5],
            'today_for_dashboard': today,
        })
    
    elif role == 'cre':
        # My Leads
        my_leads = StudentEnquiry.objects.filter(assigned=user_profile)
        total_leads = my_leads.count()
        cre_followup_context = _get_cre_followup_context(user_profile)

        context.update({
            'total_leads': total_leads,
            'show_payment_cards': True,
            **cre_followup_context,
        })
    return render(request, 'dashboard.html', context)
#=======================================LOGIN=========================================================================

# LOGIN REDIRECT BY ROLE

@never_cache
def login_view(request):

    if request.method == "POST":

        login_input = request.POST.get("username")
        password = request.POST.get("password")

        # Support login with email
        if "@" in login_input:
            try:
                user_obj = User.objects.get(email__iexact=login_input)
                username = user_obj.username
            except User.DoesNotExist:
                messages.error(request, "Invalid username/email or password")
                return render(request, "login.html")
        else:
            username = login_input

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            request.session.set_expiry(settings.SESSION_COOKIE_AGE)

            # Superuser → Django Admin Panel
            if user.is_superuser:
                return redirect("/admin/")

            # Everyone else → Same dashboard
            return redirect("dashboard")

        else:
            messages.error(request, "Invalid username/email or password")

    return render(request, "login.html")




@never_cache
def logout_view(request):
    from django.contrib import messages as django_messages
    storage = django_messages.get_messages(request)
    storage.used = True
    auth_logout(request)
    return redirect("login")


#===============================ENQUIRY==================================================================

@login_required

def create_lead(request):
    import re

    default_status = Status.objects.filter(status_name__iexact='New Enquiry').first()
    current_user_profile = UserProfile.objects.filter(user=request.user).first()
    sources = Source.objects.all()

    if request.method == "POST":

        REQUIRED_FIELDS = ['name', 'email', 'phone']
        for field in REQUIRED_FIELDS:
            if not request.POST.get(field):
                messages.error(request, "Please fill all mandatory fields.")
                return redirect('create_lead')

        if not default_status:
            messages.error(request, "Default status 'New Enquiry' was not found. Please create it first.")
            return redirect('create_lead')

        enquiry_date_raw = (request.POST.get('date') or '').strip()
        if enquiry_date_raw:
            enquiry_date = parse_date(enquiry_date_raw)
            if enquiry_date is None:
                messages.error(request, "Please enter a valid date in YYYY-MM-DD format.")
                return redirect('create_lead')
        else:
            enquiry_date = timezone.localdate()

        field_validations = {
            'name': ('Full name', 50),
            'email': ('Email', 50),
            'phone': ('Mobile', 25),
            'qualification': ('Qualification', 50),
            'guardian_number': ('Guardian number', 25),
            'house_name': ('House name', 150),
            'place': ('Place', 150),
            'district': ('District', 150),
            'state': ('State', 150),
            'location': ('Location', 100),
            'feedback': ('Feedback', 150),
            'followup_title': ('Follow-up title', 150),
            'college_name': ('College name', 150),
            'campaign_name': ('Campaign name', 150),
            'campaign_adset': ('Campaign adset', 150),
            'campaign_content': ('Campaign content', 150),
            'whatsapp': ('WhatsApp', 20),
        }
        
        for field, (label, max_len) in field_validations.items():
            value = request.POST.get(field, '')
            if value and len(value) > max_len:
                messages.error(request, f"{label} must be {max_len} characters or fewer (currently {len(value)} characters).")
                return redirect('create_lead')

        def normalize_phone_number(value):
            cleaned = re.sub(r"[^\d+]", "", (value or "").strip())
            if cleaned.startswith("+"):
                cleaned = "+" + cleaned[1:].replace("+", "")
            else:
                cleaned = cleaned.replace("+", "")
            return cleaned

        def duplicate_phone_key(value):
            digits = re.sub(r"\D", "", value or "")
            if len(digits) == 12 and digits.startswith("91"):
                return digits[2:]
            return digits

        email = (request.POST.get('email') or '').strip()
        phone = normalize_phone_number(request.POST.get('phone'))
        whatsapp = normalize_phone_number(request.POST.get('whatsapp'))
        guardian_number = normalize_phone_number(request.POST.get('guardian_number'))
        phone_pattern = re.compile(r"^\+?\d{10,15}$")

        if phone and not phone_pattern.match(phone):
            messages.error(request, "Phone number must be 10 to 15 digits and can start with +91.")
            return redirect('create_lead')

        if whatsapp and not phone_pattern.match(whatsapp):
            messages.error(request, "WhatsApp number must be 10 to 15 digits and can start with +91.")
            return redirect('create_lead')

        if guardian_number and not phone_pattern.match(guardian_number):
            messages.error(request, "Guardian number must be 10 to 15 digits and can start with +91.")
            return redirect('create_lead')

        dob_raw = (request.POST.get('dob') or '').strip()
        dob = parse_date(dob_raw) if dob_raw else None
        if dob_raw and dob is None:
            messages.error(request, "Please enter a valid DOB in YYYY-MM-DD format.")
            return redirect('create_lead')

        year_of_passing_raw = (request.POST.get('year_of_passing') or '').strip()
        year_of_passing = None
        if year_of_passing_raw:
            if not year_of_passing_raw.isdigit():
                messages.error(request, "Year of passing must be a valid year.")
                return redirect('create_lead')
            year_of_passing = int(year_of_passing_raw)

        source_obj = None
        source_id = request.POST.get('source')
        if source_id:
            source_obj = Source.objects.filter(id=source_id).first()
            if not source_obj:
                messages.error(request, "Please select a valid source.")
                return redirect('create_lead')

        if StudentEnquiry.objects.filter(email__iexact=email).exists():
            messages.error(request, "This email address already exists.")
            return redirect('create_lead')

        submitted_phone_key = duplicate_phone_key(phone)
        existing_numbers = list(StudentEnquiry.objects.values_list('mobile', flat=True))
        existing_numbers.extend(Student.objects.values_list('mobile', flat=True))

        if submitted_phone_key and any(
            duplicate_phone_key(existing_number) == submitted_phone_key
            for existing_number in existing_numbers
        ):
            messages.error(request, "This phone number already exists.")
            return redirect('create_lead')

        #  MULTIPLE COURSES
        department_ids = request.POST.getlist("departments[]")
        course_ids = request.POST.getlist("courses[]")

        if current_user_profile and current_user_profile.role == 'cre':
            assigned = current_user_profile
        else:
            assigned_to_id = request.POST.get('assigned_to')
            assigned = (
                UserProfile.objects.filter(id=assigned_to_id).first()
                if assigned_to_id else None
            )

        # CREATE LEAD (ADDRESS NORMALIZED)
        lead = StudentEnquiry.objects.create(

            full_name=request.POST.get('name'),
            email=email,
            mobile=phone,
            enquiry_date=enquiry_date,
            qualification=request.POST.get('qualification'),
            dob=dob,
            guardian_number=guardian_number or None,
            year_of_passing=year_of_passing,

            house_name=request.POST.get('house_name'),
            place=request.POST.get('place'),
            district=request.POST.get('district'),
            state=request.POST.get('state'),

            location=request.POST.get('location'),
            feedback=request.POST.get('feedback'),
            source=source_obj,

            status=default_status,
            lead_type='new',
            assigned=assigned,
            created_by=request.user,

            followup_title=request.POST.get('followup_title'),
            followup_date=request.POST.get('followup_date') or None,
            followup_time=request.POST.get('followup_time') or None,

            college_name=request.POST.get('college_name'),
            mode=request.POST.get('mode'),
            campaign_name=request.POST.get('campaign_name'),
            campaign_adset=request.POST.get('campaign_adset'),
            campaign_content=request.POST.get('campaign_content'),
            whatsapp=whatsapp,
        )

        LeadActivity.objects.create(
            lead=lead,
            user=request.user,
            action="Lead Created",
            new_value=lead.full_name
        )

        #  SAVE MULTIPLE COURSES
        for dept_id, course_id in zip(department_ids, course_ids):
            if dept_id and course_id:
                LeadCourse.objects.create(
                    lead=lead,
                    department_id=dept_id,
                    course_id=course_id
        )

        # Create FollowUp entry if follow-up details provided
        if request.POST.get('followup_title') and request.POST.get('followup_date'):
            FollowUp.objects.create(
                lead=lead,
                title=request.POST.get('followup_title'),
                followup_date=request.POST.get('followup_date'),
                followup_time=request.POST.get('followup_time') or None,
                assigned_to=assigned,
                created_by=request.user
            )
            
            # Create activity for follow-up
            LeadActivity.objects.create(
                lead=lead,
                action="Follow-up Added",
                new_value=f"{request.POST.get('followup_title')} on {request.POST.get('followup_date')}"
            )

        messages.success(request, "Lead created successfully")
        return redirect('create_lead')

    context = {
        'departments': Department.objects.all(),
        'courses': Course.objects.all(),
        'statuses': Status.objects.all(),
        'sources': sources,
        'default_status': default_status,
        'cre_users': UserProfile.objects.filter(role='cre').select_related('user'),
        'current_user_profile': current_user_profile,
    }
    return render(request, 'create_lead.html', context)

def import_leads_excel(request):
    import openpyxl
    import re
    from datetime import datetime
    if request.method != "POST":
        return redirect('create_lead')

    excel_file = request.FILES.get('excel_file')

    if not excel_file or not excel_file.name.endswith('.xlsx'):
        messages.error(request, "Please upload a valid .xlsx file")
        return redirect('create_lead')

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # Read header row → map column names to index
    headers = {}
    for idx, cell in enumerate(sheet[1]):
        if cell.value:
            headers[cell.value.strip().lower()] = idx

    created = 0
    updated = 0
    skipped = 0
    skipped_reasons = []
    default_status = Status.objects.filter(status_name__iexact='New Enquiry').first()

    if not default_status:
        messages.error(request, "Default status 'New Enquiry' was not found. Please create it first.")
        return redirect('create_lead')

    def split_excel_values(value):
        if value in (None, ""):
            return []
        return [item.strip() for item in re.split(r"[,\n;|]+", str(value)) if item and item.strip()]

    def normalize_import_label(value):
        if value in (None, ""):
            return ""
        return re.sub(r"\s+", "", str(value)).lower()

    department_lookup = {
        normalize_import_label(department.department_name): department
        for department in Department.objects.all()
    }
    source_lookup = {
        normalize_import_label(source.source_name): source
        for source in Source.objects.all()
    }
    course_lookup = {}
    for course in Course.objects.select_related('department'):
        normalized_course = normalize_import_label(course.course_name)
        course_lookup.setdefault(normalized_course, []).append(course)

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            def get(col):
                return row[headers[col]] if col in headers else None

            name = get('name')
            email = get('email')
            phone = get('phone')

            if not email and not phone:
                skipped += 1
                skipped_reasons.append(f"Row {row_number}: missing both email and phone")
                continue

            # DUPLICATE CHECK
            lead = StudentEnquiry.objects.filter(
                Q(email__iexact=email) | Q(mobile=phone)
            ).first()

            is_new = False
            if not lead:
                lead = StudentEnquiry(
                    created_by=request.user,
                    enquiry_date=get('date') or datetime.today().date()
                )
                is_new = True

            #  TEXT → FK RESOLUTION
            dept_name = get('department')
            dept_names = split_excel_values(dept_name)
            resolved_departments = [
                department_lookup.get(normalize_import_label(name))
                for name in dept_names
                if department_lookup.get(normalize_import_label(name))
            ]

            course_name = get('course')
            course_names = split_excel_values(course_name)

            lead.status = default_status
            if not lead.lead_type:
                lead.lead_type = 'new'

            imported_source = _normalize_source_name(get('source'))
            resolved_source = source_lookup.get(normalize_import_label(imported_source)) if imported_source else None

            #  FIELD AUTO-MERGE (only if value exists)
            field_map = {
                'full_name': name,
                'email': email,
                'mobile': phone,
                'qualification': get('qualification'),
                'address': get('address'),
                'location': get('location'),
                'followup_title': get('followup_title'),
                'followup_date': get('followup_date'),
                'followup_time': get('followup_time'),
                'college_name': get('college_name'),
                'mode': get('mode'),
                'campaign_name': get('campaign_name'),
                'campaign_adset': get('campaign_adset'),
                'campaign_content': get('campaign_content'),
                'whatsapp': get('whatsapp'),
            }

            for field, value in field_map.items():
                if value not in [None, ""]:
                    setattr(lead, field, value)

            if resolved_source:
                lead.source = resolved_source

            # ASSIGN DEFAULT OWNER IF EMPTY
            if not lead.assigned:
                lead.assigned = UserProfile.objects.filter(user=request.user).first()

            lead.save()

            # Store imported department/course pairs in the relation table used by lead listing.
            lead_course_pairs = []

            if course_names:
                for index, course_label in enumerate(course_names):
                    department = resolved_departments[index] if index < len(resolved_departments) else None
                    normalized_course_label = normalize_import_label(course_label)
                    matching_courses = course_lookup.get(normalized_course_label, [])

                    course = next(
                        (item for item in matching_courses if not department or item.department_id == department.id),
                        None
                    )
                    if not course and department:
                        # Fall back to the course's own department if the sheet's department text
                        # doesn't line up perfectly with the stored mapping.
                        course = matching_courses[0] if matching_courses else None

                    if not course:
                        continue

                    lead_course_pairs.append((department or course.department, course))
            elif resolved_departments:
                for department in resolved_departments:
                    lead_course_pairs.append((department, None))

            for department, course in lead_course_pairs:
                if department and course:
                    LeadCourse.objects.get_or_create(
                        lead=lead,
                        department=department,
                        course=course
                    )

            if is_new:
                created += 1
            else:
                updated += 1

        except Exception as exc:
            skipped += 1
            skipped_reasons.append(f"Row {row_number}: {exc}")
            continue

    messages.success(
        request,
        f"Import completed: {created} new, {updated} updated, {skipped} skipped"
    )

    if skipped_reasons:
        preview_count = 10
        preview = "; ".join(skipped_reasons[:preview_count])
        if len(skipped_reasons) > preview_count:
            preview += f"; and {len(skipped_reasons) - preview_count} more"
        messages.warning(request, f"Skip reasons: {preview}")

    return redirect('create_lead')

#===============================LEAD==============================================================================

from django.core.paginator import Paginator
from django.db.models import Q

from django.db.models import Q, Prefetch

def leads__filter(request):

    # 🚀 Prefetch multiple courses
    lead_courses_prefetch = Prefetch(
        'lead_courses',
        queryset=LeadCourse.objects.select_related('course','department')
    )

    leads_qs = StudentEnquiry.objects.select_related(
        'status',
        'source',
        'assigned__user',
        'created_by'
    ).prefetch_related(
        lead_courses_prefetch
    ).order_by('-created_at')

    # ================= PERMISSION FILTER =================
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role == 'cre':
        leads_qs = leads_qs.filter(assigned=user_profile)

    # ================= SEARCH =================
    q = request.GET.get('q')
    if q:
        leads_qs = leads_qs.filter(
            Q(full_name__icontains=q) |
            Q(email__icontains=q) |
            Q(mobile__icontains=q)
        )

    # ================= DATE FILTER =================
    date_from = request.GET.get('date_from')
    if date_from:
        leads_qs = leads_qs.filter(enquiry_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        leads_qs = leads_qs.filter(enquiry_date__lte=date_to)

    # ================= STATUS =================
    status = request.GET.get('status')
    if status:
        leads_qs = leads_qs.filter(status_id=status)

    # ================= SOURCE =================
    source = request.GET.get('source')
    if source:
        leads_qs = leads_qs.filter(source_id=source)

    # ================= DEPARTMENT (NEW LOGIC) =================
    department = request.GET.get('department')
    if department:
        leads_qs = leads_qs.filter(
            lead_courses__department_id=department
        ).distinct()

    # ================= COURSE (NEW LOGIC) =================
    course = request.GET.get('course')
    if course:
        leads_qs = leads_qs.filter(
            lead_courses__course_id=course
        ).distinct()

    # ================= ASSIGNED =================
    assigned_to = request.GET.get('assigned_to')
    if assigned_to:
        leads_qs = leads_qs.filter(assigned_id=assigned_to)

    # ================= PAGINATION =================
    per_page = int(request.GET.get('per_page', 10))
    paginator = Paginator(leads_qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'leads': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,

        'statuses': Status.objects.all(),
        'sources': Source.objects.all(),
        'departments': Department.objects.all(),
        'courses': Course.objects.all(),
        'users': UserProfile.objects.filter(role='cre').select_related('user'),
    }

    return render(request, 'leads.html', context)




from django.db.models import Prefetch

def lead_list(request):
    query = request.GET.get('q')

    # 🚀 Prefetch multiple courses
    lead_courses_prefetch = Prefetch(
        'lead_courses',
        queryset=LeadCourse.objects.select_related('course', 'department')
    )

    leads_qs = StudentEnquiry.objects.select_related(
        'status',
        'source',
        'assigned',
        'created_by'
    ).prefetch_related(
        lead_courses_prefetch
    )

    # Filter by assigned CRE
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role == 'cre':
        leads_qs = leads_qs.filter(assigned=user_profile)

    # Search
    if query:
        leads_qs = leads_qs.filter(
            Q(full_name__icontains=query) |
            Q(email__icontains=query) |
            Q(mobile__icontains=query)
        )

    leads_qs = leads_qs.order_by('-created_at')

    # Pagination
    per_page = int(request.GET.get('per_page', 10))
    paginator = Paginator(leads_qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'leads.html', {
        'leads': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,

        'statuses': Status.objects.all(),
        'sources': Source.objects.all(),
        'departments': Department.objects.all(),
        'courses': Course.objects.all(),
        'users': UserProfile.objects.filter(role='cre').select_related('user'),
    })

def lead_detail(request, lead_id):
    lead = get_object_or_404(
        StudentEnquiry.objects.prefetch_related(
            Prefetch(
                'lead_courses',
                queryset=LeadCourse.objects.select_related('course','department')
            )
        ),
        id=lead_id
    )

    return render(request, 'lead_detail.html', {
        'lead': lead,
        'lead_courses': lead.lead_courses.all()
    })


from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages

from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Prefetch
from .models import LeadCourse
def edit_lead(request, lead_id):
    lead = get_object_or_404(StudentEnquiry, id=lead_id)

    departments = Department.objects.all()
    courses = Course.objects.all()
    statuses = Status.objects.all()
    sources = Source.objects.all()
    cre_users = UserProfile.objects.filter(role='cre').select_related('user')
    existing_courses = LeadCourse.objects.filter(lead=lead)

    if request.method == "POST":

        # ===== STORE OLD VALUES =====
        old_status_id = str(lead.status_id) if lead.status_id else None

        old_data = {
            "Name": lead.full_name or None,
            "Phone": lead.mobile or None,
            "Email": lead.email or None,
            "Qualification": lead.qualification or None,
            "DOB": str(lead.dob) if lead.dob else None,
            "Guardian Number": lead.guardian_number or None,
            "Year of Passing": str(lead.year_of_passing) if lead.year_of_passing else None,
            "Address": f"{lead.house_name or ''}, {lead.place or ''}, {lead.district or ''}, {lead.state or ''}".strip(", ") or None,
            "Location": lead.location or None,
            "Source": lead.source.source_name if lead.source else None,
           "Assigned To": (
                lead.assigned.user.get_full_name() 
                or lead.assigned.user.username
            ) if lead.assigned else None,
            "Follow-up Title": lead.followup_title ,
            "Follow-up Date": str(lead.followup_date) if lead.followup_date else None,
            "Follow-up Time": str(lead.followup_time)[:5] if lead.followup_time else None,
        }

        # ===== UPDATE FIELDS =====
        lead.full_name = request.POST.get('name')
        lead.email = request.POST.get('email')
        lead.mobile = request.POST.get('phone')
        lead.enquiry_date = request.POST.get('date')
        lead.qualification = request.POST.get('qualification')
        lead.dob = request.POST.get('dob') or None
        lead.guardian_number = request.POST.get('guardian_number') or None
        lead.year_of_passing = request.POST.get('year_of_passing') or None
        lead.location = request.POST.get('location')
        source_id = request.POST.get('source')
        lead.source = Source.objects.filter(id=source_id).first() if source_id else None
        lead.feedback = request.POST.get('feedback')

        lead.house_name = request.POST.get('house_name')
        lead.place = request.POST.get('place')
        lead.district = request.POST.get('district')
        lead.state = request.POST.get('state')

        lead.followup_title = request.POST.get('followup_title') 
        lead.followup_date = request.POST.get('followup_date') or None
        lead.followup_time = request.POST.get('followup_time') or None

        lead.college_name = request.POST.get('college_name')
        lead.mode = request.POST.get('mode')
        lead.campaign_name = request.POST.get('campaign_name')
        lead.campaign_adset = request.POST.get('campaign_adset')
        lead.campaign_content = request.POST.get('campaign_content')
        lead.whatsapp = request.POST.get('whatsapp')

        if request.POST.get('status'):
            lead.status_id = request.POST.get('status')
            if lead.status and lead.status.status_name.strip().lower() == "enrolled":
                lead.lead_type = 'done'

        assigned_to_id = request.POST.get('assigned_to')
        lead.assigned = (
            UserProfile.objects.filter(id=assigned_to_id).first()
            if assigned_to_id else None
        )

        lead.save()

        # ===== STATUS CHANGE =====
        new_status_id = str(lead.status_id) if lead.status_id else None

        if old_status_id != new_status_id:
            old_status_name = Status.objects.get(id=old_status_id).status_name if old_status_id else "—"
            new_status_name = lead.status.status_name if lead.status else "—"

            if old_status_name != new_status_name:
                LeadActivity.objects.create(
                    lead=lead,
                    user=request.user,
                    action="Updated Status",
                    old_value=old_status_name,
                    new_value=new_status_name
                )

        # ===== NEW VALUES =====
        new_data = {
            "Name": lead.full_name or None,
            "Phone": lead.mobile or None,
            "Email": lead.email or None,
            "Qualification": lead.qualification or None,
            "DOB": str(lead.dob) if lead.dob else None,
            "Guardian Number": lead.guardian_number or None,
            "Year of Passing": str(lead.year_of_passing) if lead.year_of_passing else None,
            "Address": f"{lead.house_name or ''}, {lead.place or ''}, {lead.district or ''}, {lead.state or ''}".strip(", ") or None,
            "Location": lead.location or None,
            "Source": lead.source.source_name if lead.source else None,
            "Assigned To": (
                lead.assigned.user.get_full_name() 
                or lead.assigned.user.username
            ) if lead.assigned else None,
            "Follow-up Title": lead.followup_title ,
            "Follow-up Date": str(lead.followup_date) if lead.followup_date else None,
            "Follow-up Time": str(lead.followup_time)[:5] if lead.followup_time else None,
        }

        for field in old_data:
            if old_data[field] != new_data[field]:
                LeadActivity.objects.create(
                    lead=lead,
                    user=request.user,
                    action=f"Updated {field}",
                    old_value=old_data[field] or "—",
                    new_value=new_data[field] or "—"
                )

        # ===== COURSE CHANGE =====
        department_ids = request.POST.getlist("departments[]")
        course_ids = request.POST.getlist("courses[]")

        old_courses = list(
            LeadCourse.objects.filter(lead=lead)
            .values_list("course__course_name", flat=True)
        )

        LeadCourse.objects.filter(lead=lead).delete()

        for dept_id, course_id in zip(department_ids, course_ids):
            if dept_id and course_id:
                LeadCourse.objects.create(
                    lead=lead,
                    department_id=dept_id,
                    course_id=course_id
                )

        new_courses = list(
            LeadCourse.objects.filter(lead=lead)
            .values_list("course__course_name", flat=True)
        )

        if set(old_courses) != set(new_courses):
            LeadActivity.objects.create(
                lead=lead,
                user=request.user,
                action="Updated Courses",
                old_value=", ".join(old_courses) if old_courses else "—",
                new_value=", ".join(new_courses) if new_courses else "—"
            )

        messages.success(request, "Lead updated successfully")
        return redirect('lead_profile', lead.id)

    # GET request
    return render(request, 'edit_lead.html', {
        'lead': lead,
        'departments': departments,
        'courses': courses,
        'statuses': statuses,
        'sources': sources,
        'cre_users': cre_users,
        'existing_courses': existing_courses,
    })

from django.utils.timezone import now



def update_lead_status(request, lead_id):
    if request.method != "POST":
        return redirect('leads')

    lead = get_object_or_404(StudentEnquiry, id=lead_id)
    if lead.status and lead.status.status_name.strip().lower() == "enrolled":
        messages.error(request, "Enrolled lead is locked and status cannot be changed.")
        return redirect('leads')

    new_status_id = request.POST.get('status')
    new_status = get_object_or_404(Status, id=new_status_id)
    normalized_status_name = new_status.status_name.strip().lower()
    is_enrolled = normalized_status_name == "enrolled"
    is_followup_status = 'follow' in normalized_status_name

    followup_title = (request.POST.get('followup_title') or '').strip()
    followup_date = (request.POST.get('followup_date') or '').strip()
    followup_time = (request.POST.get('followup_time') or '').strip()
    followup_notes = (request.POST.get('followup_notes') or '').strip()

    if is_followup_status and (not followup_title or not followup_date):
        messages.error(request, "Title and date are required for follow-up status.")
        return redirect('leads')

    old_status = str(lead.status) if lead.status else "-"

    lead.status = new_status
    update_fields = ['status']

    if is_enrolled and lead.lead_type != 'done':
        lead.lead_type = 'done'
        update_fields.append('lead_type')

    if is_followup_status:
        lead.followup_title = followup_title
        lead.followup_date = followup_date
        lead.followup_time = followup_time or None
        update_fields.extend(['followup_title', 'followup_date', 'followup_time'])

    lead.save(update_fields=update_fields)

    LeadActivity.objects.create(
        lead=lead,
        user=request.user,
        action="Updated Status",
        old_value=old_status,
        new_value=str(new_status)
    )

    if is_followup_status:
        FollowUp.objects.create(
            lead=lead,
            title=followup_title,
            followup_date=followup_date,
            followup_time=followup_time or None,
            notes=followup_notes,
            assigned_to=lead.assigned,
            created_by=request.user,
            status='pending'
        )

        followup_summary = f"{followup_title} on {followup_date}"
        if followup_time:
            followup_summary = f"{followup_summary} at {followup_time}"

        LeadActivity.objects.create(
            lead=lead,
            user=request.user,
            action="Follow-up Added",
            new_value=followup_summary
        )

    if is_enrolled:
        if not Student.objects.filter(mobile=lead.mobile).exists():
            student = Student.objects.create(
                student_name=lead.full_name,
                email=lead.email,
                mobile=lead.mobile,
                house_name=lead.house_name,
                place=lead.place,
                district=lead.district,
                state=lead.state,
                location=lead.location,
                enrolled_date=now().date()
            )

            for lead_course in LeadCourse.objects.filter(lead=lead).select_related('department', 'course'):
                StudentCourse.objects.create(
                    student=student,
                    department=lead_course.department,
                    course=lead_course.course
                )

            messages.success(request, "Lead converted to student successfully")
        else:
            messages.warning(request, "Student already exists for this lead")
        return redirect('leads')

    if is_followup_status:
        messages.success(request, "Follow-up created successfully")
        return redirect('followups')

    messages.success(request, "Lead status updated successfully")
    return redirect('leads')

def update_lead_type(request, lead_id):
    if request.method != "POST":
        return redirect('leads')

    lead = get_object_or_404(StudentEnquiry, id=lead_id)
    if lead.status and lead.status.status_name.strip().lower() == "enrolled":
        messages.error(request, "Type cannot be changed for enrolled leads.")
        return redirect('leads')

    new_type = (request.POST.get('lead_type') or '').strip().lower()
    valid_types = {choice[0] for choice in StudentEnquiry.TYPE_CHOICES}
    if new_type not in valid_types:
        messages.error(request, "Invalid lead type selected.")
        return redirect('leads')

    note = (request.POST.get('interest_notes') or '').strip()
    if not note:
        messages.error(request, "Interest note is required when changing lead type.")
        return redirect('leads')

    old_type = lead.get_lead_type_display() if lead.lead_type else "—"
    lead.lead_type = new_type
    lead.interest_notes = note
    lead.save(update_fields=['lead_type', 'interest_notes'])

    LeadActivity.objects.create(
        lead=lead,
        user=request.user,
        action="Updated Lead Type",
        old_value=old_type,
        new_value=f"{lead.get_lead_type_display()} | Note: {note}"
    )

    messages.success(request, "Lead type updated successfully")
    return redirect('leads')

def delete_lead(request, lead_id):
    lead = get_object_or_404(StudentEnquiry, id=lead_id)
    lead.delete()
    return redirect('leads')


def lead_profile(request, lead_id):
    lead = get_object_or_404(StudentEnquiry, id=lead_id)

    activities = list(lead.activities.select_related('user').order_by('-created_at'))


    tasks = LeadTask.objects.filter(
        lead=lead
    ).order_by('-created_at')
    
    followups = FollowUp.objects.filter(
        lead=lead
    ).order_by('-followup_date', '-followup_time')

    followup_status_by_activity_value = {}
    for followup in followups:
        activity_value = f"{followup.title} on {followup.followup_date}"
        existing_status = followup_status_by_activity_value.get(activity_value)

        # Show a single effective status in activity: completed wins, otherwise pending.
        if followup.status == 'completed' or existing_status is None:
            followup_status_by_activity_value[activity_value] = (
                'completed' if followup.status == 'completed' else 'pending'
            )

    for activity in activities:
        activity.followup_display_status = followup_status_by_activity_value.get(activity.new_value)

    users = UserProfile.objects.all()

    return render(request, "lead_profile.html", {
        "lead": lead,
        "activities": activities,
        "tasks": tasks,
        "followups": followups,
        "users": users,
    })

from django.views.decorators.http import require_POST
from django.http import JsonResponse


@require_POST
def bulk_update_status(request):
    lead_ids = request.POST.getlist('lead_ids')
    status_id = request.POST.get('status_id')

    if not lead_ids or not status_id:
        return JsonResponse(
            {'success': False, 'error': 'Lead IDs and status are required'},
            status=400
        )

    new_status = Status.objects.filter(id=status_id).first()
    if not new_status:
        return JsonResponse(
            {'success': False, 'error': 'Invalid status'},
            status=400
        )

    leads = StudentEnquiry.objects.filter(id__in=lead_ids)

    converted_count = 0
    updated_count = 0
    locked_count = 0
    is_enrolled = new_status.status_name.strip().lower() == "enrolled"

    for lead in leads:
        if lead.status and lead.status.status_name.strip().lower() == "enrolled":
            locked_count += 1
            continue

        old_status = str(lead.status) if lead.status else "—"

        lead.status = new_status
        update_fields = ['status']
        if is_enrolled and lead.lead_type != 'done':
            lead.lead_type = 'done'
            update_fields.append('lead_type')
        lead.save(update_fields=update_fields)
        updated_count += 1

        LeadActivity.objects.create(
            lead=lead,
            action="Updated Status (Bulk)",
            old_value=old_status,
            new_value=str(new_status)
        )

        if is_enrolled:
            mobile = (lead.mobile or "").strip()
            if not mobile:
                continue

            if not Student.objects.filter(mobile=mobile).exists():
                student = Student.objects.create(
                    student_name=lead.full_name,
                    email=lead.email,
                    mobile=mobile,
                    house_name=lead.house_name,
                    place=lead.place,
                    district=lead.district,
                    state=lead.state,
                    location=lead.location,
                    enrolled_date=now().date()
                )

                for lead_course in LeadCourse.objects.filter(lead=lead).select_related('department','course'):
                    StudentCourse.objects.create(
                        student=student,
                        department=lead_course.department,
                        course=lead_course.course
                    )

                converted_count += 1

    return JsonResponse({
        'success': True,
        'updated': updated_count,
        'converted': converted_count,
        'locked': locked_count
    })


@require_POST
def bulk_delete_leads(request):
    lead_ids = request.POST.getlist('lead_ids')

    StudentEnquiry.objects.filter(id__in=lead_ids).delete()

    return JsonResponse({'success': True})

@require_POST
def bulk_assign_leads(request):
    lead_ids = request.POST.getlist('lead_ids')
    user_id = request.POST.get('user_id')
    
    user_profile = UserProfile.objects.filter(id=user_id, role='cre').first()
    if not user_profile:
        return JsonResponse({'success': False, 'error': 'CRE user not found'})
    
    StudentEnquiry.objects.filter(id__in=lead_ids).update(assigned=user_profile)
    
    return JsonResponse({'success': True})

def lead_live_search(request):
    query = request.GET.get('q', '')
    
    lead_courses_prefetch = Prefetch(
        'lead_courses',
        queryset=LeadCourse.objects.select_related('course', 'department')
    )
    
    leads = StudentEnquiry.objects.select_related(
        'status', 'assigned__user', 'created_by'
    ).prefetch_related(lead_courses_prefetch)
    
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role == 'cre':
        leads = leads.filter(assigned=user_profile)
    
    if query:
        leads = leads.filter(
            Q(full_name__icontains=query) |
            Q(email__icontains=query) |
            Q(mobile__icontains=query)
        )
    
    leads = leads.order_by('-created_at')[:50]
    
    data = []
    for lead in leads:
        courses = [lc.course.course_name for lc in lead.lead_courses.all()]
        departments = [lc.department.department_name for lc in lead.lead_courses.all()]
        
        data.append({
            "id": lead.id,
            "full_name": lead.full_name,
            "mobile": lead.mobile,
            "assigned": lead.assigned.user.username if lead.assigned else "",
            "courses": courses,
            "departments": departments,
            "mode": lead.mode or "",
            "status": lead.status.status_name if lead.status else "",
            "status_id": lead.status.id if lead.status else None,
            "lead_type": lead.lead_type or "hot",
            "source": lead.source.source_name if lead.source else "",
            "created_by": lead.created_by.get_full_name() or lead.created_by.username if lead.created_by else "",
            "created_at": lead.created_at.strftime("%d %b %Y"),
            "followup_date": lead.followup_date.strftime("%d %b %Y") if lead.followup_date else "",
            "followup_time": str(lead.followup_time)[:5] if lead.followup_time else "",
            "enquiry_date": lead.enquiry_date.strftime("%Y-%m-%d") if lead.enquiry_date else "",
            "is_enrolled": lead.status.status_name.lower() == 'enrolled' if lead.status else False,
        })
    
    return JsonResponse({"leads": data})
#=======================================STUDENT=========================================================================

from .models import Student
from core.models import Department, Course, Batch

def sync_student_fee(student):
    from datetime import timedelta
    from decimal import Decimal
    default_due_date = student.enrolled_date or now().date()
    student_courses = list(
        StudentCourse.objects.filter(student=student).select_related('course', 'batch')
    )
    active_course_ids = {sc.course_id for sc in student_courses}

    def get_final_due_date(student_course):
        if student_course.batch and student_course.batch.end_date:
            return student_course.batch.end_date - timedelta(days=5)
        return default_due_date

    def sync_fee_installments(fee, course, final_due_date):
        if fee.custom_installment_plan:
            return

        if not course.installment_enabled or course.installment_count <= 1:
            return

        total_fee = Decimal(str(fee.total_fee)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        count = course.installment_count
        base_amount = Decimal(int(total_fee) // count)
        existing_installments = {
            inst.installment_number: inst
            for inst in fee.installments.all()
        }
        expected_numbers = set()

        for i in range(1, count + 1):
            installment_due_date = final_due_date - timedelta(
                days=(count - i) * course.installment_interval_days
            )
            if i == count:
                amount = total_fee - (base_amount * (count - 1))
            else:
                amount = base_amount

            expected_numbers.add(i)
            installment = existing_installments.get(i)
            if installment:
                update_fields = []
                if installment.amount != amount:
                    installment.amount = amount
                    update_fields.append('amount')
                if installment.due_date != installment_due_date:
                    installment.due_date = installment_due_date
                    update_fields.append('due_date')
                if update_fields:
                    installment.save(update_fields=update_fields)
            else:
                FeeInstallment.objects.create(
                    fee=fee,
                    installment_number=i,
                    amount=amount,
                    due_date=installment_due_date
                )

        for number, installment in existing_installments.items():
            if number in expected_numbers:
                continue
            if installment.payments.exists():
                continue
            installment.delete()

    existing_fees = list(
        Fee.objects.filter(student=student, course__isnull=False)
        .select_related('course')
        .prefetch_related('payments', 'installments__payments')
    )

    for fee in existing_fees:
        if fee.course_id in active_course_ids:
            continue

        if fee.payments.exists():
            continue

        fee.delete()

    for sc in student_courses:
        if not sc.course.fees:
            continue

        fee_total = Decimal(str(sc.course.fees)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        final_due_date = get_final_due_date(sc)
        fee, created = Fee.objects.get_or_create(
            student=student,
            course=sc.course,
            defaults={
                "total_fee": fee_total,
                "paid_amount": Decimal('0.00'),
                "advance_paid": Decimal('0.00'),
                "course_fee_paid": Decimal('0.00'),
                "remaining_balance": fee_total,
                "advance_amount": min(fee_total, sc.course.advance_payment_amount or Decimal('0.00')),
                "due_date": final_due_date,
            }
        )

        update_fields = []
        if fee.total_fee != fee_total:
            fee.total_fee = fee_total
            update_fields.append("total_fee")

        if fee.due_date != final_due_date:
            fee.due_date = final_due_date
            update_fields.append("due_date")

        expected_advance_amount = min(fee_total, sc.course.advance_payment_amount or Decimal('0.00'))
        if fee.advance_amount != expected_advance_amount:
            fee.advance_amount = expected_advance_amount
            update_fields.append("advance_amount")

        if update_fields:
            fee.save(update_fields=update_fields)

        if sc.course.installment_enabled and sc.course.installment_count > 1:
            sync_fee_installments(fee, sc.course, final_due_date)

        fee.recalculate_payment_status()

def create_student(request):
    departments = Department.objects.all()
    courses = Course.objects.select_related("department").all()

    def blank_course_row():
        return {
            "department_id": "",
            "course_id": "",
            "department_error": "",
            "course_error": "",
            "general_error": "",
        }

    def build_context(form_data=None, field_errors=None, course_rows=None):
        return {
            "departments": departments,
            "courses": courses,
            "form_data": form_data or {},
            "field_errors": field_errors or {},
            "course_rows": course_rows or [blank_course_row()],
        }

    if request.method == "POST":
        student_name = (request.POST.get("student_name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        status = "enrolled"
        house_name = (request.POST.get("house_name") or "").strip()
        place = (request.POST.get("place") or "").strip()
        district = (request.POST.get("district") or "").strip()
        state = (request.POST.get("state") or "").strip()
        location = (request.POST.get("location") or "").strip()

        def normalize_phone_number(value):
            cleaned = re.sub(r"[^\d+]", "", (value or "").strip())
            if cleaned.startswith("+"):
                cleaned = "+" + cleaned[1:].replace("+", "")
            else:
                cleaned = cleaned.replace("+", "")
            return cleaned

        mobile = normalize_phone_number(request.POST.get("mobile"))

        department_ids = request.POST.getlist("departments[]")
        course_ids = request.POST.getlist("courses[]")

        form_data = {
            "student_name": student_name,
            "email": email,
            "mobile": mobile,
            "status": status,
            "house_name": house_name,
            "place": place,
            "district": district,
            "state": state,
            "location": location,
        }
        field_errors = {}
        course_rows = []
        valid_course_pairs = []

        name_pattern = re.compile(r"^[A-Za-z][A-Za-z .'-]*$")
        mobile_pattern = re.compile(r"^\+?\d{10,15}$")
        department_lookup = {str(department.id): department for department in departments}
        course_lookup = {str(course.id): course for course in courses}
        selected_course_ids = set()

        if not student_name:
            field_errors["student_name"] = "Student name is required."
        elif len(student_name) < 2:
            field_errors["student_name"] = "Student name must be at least 2 characters."
        elif len(student_name) > 50:
            field_errors["student_name"] = "Student name must be 50 characters or fewer."
        elif not name_pattern.match(student_name):
            field_errors["student_name"] = "Student name contains invalid characters."

        if email:
            if len(email) > 254:
                field_errors["email"] = "Email must be 254 characters or fewer."
            else:
                try:
                    validate_email(email)
                except ValidationError:
                    field_errors["email"] = "Enter a valid email address."
                else:
                    if Student.objects.filter(email__iexact=email).exists():
                        field_errors["email"] = "This email address already exists."

        if not mobile:
            field_errors["mobile"] = "Mobile number is required."
        elif not mobile_pattern.match(mobile):
            field_errors["mobile"] = "Enter a valid mobile number (10 to 15 digits, optional +)."
        elif Student.objects.filter(mobile=mobile).exists():
            field_errors["mobile"] = "This mobile number already exists."

        length_limits = {
            "house_name": 150,
            "place": 150,
            "district": 150,
            "state": 150,
            "location": 100,
        }
        for field_name, max_length in length_limits.items():
            value = form_data.get(field_name, "")
            if value and len(value) > max_length:
                field_errors[field_name] = f"{field_name.replace('_', ' ').title()} must be {max_length} characters or fewer."

        max_rows = max(len(department_ids), len(course_ids), 1)
        for index in range(max_rows):
            dept_id = (department_ids[index] if index < len(department_ids) else "").strip()
            course_id = (course_ids[index] if index < len(course_ids) else "").strip()
            row = blank_course_row()
            row["department_id"] = dept_id
            row["course_id"] = course_id

            if not dept_id and not course_id:
                course_rows.append(row)
                continue

            if not dept_id:
                row["department_error"] = "Department is required."
            elif dept_id not in department_lookup:
                row["department_error"] = "Select a valid department."

            course = None
            if not course_id:
                row["course_error"] = "Course is required."
            else:
                course = course_lookup.get(course_id)
                if not course:
                    row["course_error"] = "Select a valid course."

            if not row["department_error"] and not row["course_error"] and course:
                if str(course.department_id) != dept_id:
                    row["course_error"] = "Selected course does not belong to the chosen department."
                elif course_id in selected_course_ids:
                    row["course_error"] = "This course is already selected."
                else:
                    selected_course_ids.add(course_id)
                    valid_course_pairs.append((dept_id, course_id))

            course_rows.append(row)

        if not valid_course_pairs:
            field_errors["courses"] = "Select at least one department and course."

        if field_errors or any(
            row["department_error"] or row["course_error"] or row["general_error"]
            for row in course_rows
        ):
            messages.error(request, "Please correct the highlighted errors.")
            return render(
                request,
                "create_student.html",
                build_context(form_data=form_data, field_errors=field_errors, course_rows=course_rows),
            )

        student = Student.objects.create(
            student_name=student_name,
            email=email or None,
            mobile=mobile,
            status=status,
            house_name=house_name or None,
            place=place or None,
            district=district or None,
            state=state or None,
            location=location or None,
            enrolled_date=request.POST.get("enrolled_date") or now().date(),
        )

        for dept_id, course_id in valid_course_pairs:
            StudentCourse.objects.create(
                student=student,
                department_id=dept_id,
                course_id=course_id
            )

        sync_student_fee(student)
        messages.success(request, "Student created successfully.")
        return redirect("students")

    return render(request, "create_student.html", build_context())


def _clear_student_batch_assignments(student_ids):
    if not student_ids:
        return

    StudentCourse.objects.filter(
        student_id__in=student_ids,
        batch__isnull=False
    ).update(batch=None)


def _format_inr_amount(amount):
    if amount is None:
        return "0"

    normalized = amount.quantize(Decimal('0.01'))
    if normalized == normalized.to_integral():
        return str(int(normalized))
    return f"{normalized:.2f}"


def _get_student_course_fee(student, course, sync_if_missing=False):
    if not student or not course:
        return None

    if sync_if_missing:
        sync_student_fee(student)

    return Fee.objects.filter(student=student, course=course).first()


def _get_batch_assignment_block_reason(student, course=None):
    if student.status == "dropped":
        return "Dropped students cannot be assigned to a batch. Change the student status first."

    if not course or not course.fees:
        return None

    fee = _get_student_course_fee(student, course, sync_if_missing=True)
    if not fee:
        return (
            f"No fee record found for {course.course_name}. "
            "Add the fee and collect the advance payment first."
        )

    if fee.has_minimum_advance_payment:
        return None

    required = _format_inr_amount(fee.advance_payment_required)
    remaining = _format_inr_amount(fee.advance_payment_remaining)
    return (
        f"Please pay advance first for {course.course_name}. "
        f"Minimum advance is ₹{required}; ₹{remaining} is still pending."
    )
    return None


def _get_batch_assignment_candidates(batch, student_ids):
    cleaned_ids = []
    seen_ids = set()
    for student_id in student_ids:
        student_id = (student_id or "").strip()
        if not student_id:
            continue
        if student_id in seen_ids:
            continue
        seen_ids.add(student_id)
        cleaned_ids.append(student_id)

    if not cleaned_ids:
        return None, None, "Select at least one student."

    students = Student.objects.filter(id__in=cleaned_ids)
    students_by_id = {str(student.id): student for student in students}

    ordered_students = []
    student_courses = {}
    for student_id in cleaned_ids:
        student = students_by_id.get(str(student_id))
        if not student:
            return None, None, "One or more selected students were not found."

        block_reason = _get_batch_assignment_block_reason(student, batch.course)
        if block_reason:
            return None, None, f"{student.student_name}: {block_reason}"

        student_course = StudentCourse.objects.filter(student=student, course=batch.course).first()
        if not student_course:
            return None, None, f"{student.student_name} doesn't have this course"

        ordered_students.append(student)
        student_courses[student.id] = student_course

    return ordered_students, student_courses, None


def _assign_students_to_batch(batch, students, student_courses):
    with transaction.atomic():
        for student in students:
            student_course = student_courses[student.id]
            student_course.batch = batch
            student_course.save(update_fields=['batch'])
            sync_student_fee(student)

            student.status = 'active'
            student.save(update_fields=['status'])
            _queue_batch_assignment_mail(student, batch)
            fee = Fee.objects.filter(student=student, course=batch.course).order_by('-id').first()
            if fee:
                _sync_overdue_pending_mails([fee.id])


def _get_student_delete_block_reason(student):
    if not student:
        return "Student not found."

    sync_student_fee(student)
    active_course_ids = list(
        StudentCourse.objects.filter(student=student).values_list('course_id', flat=True)
    )
    unpaid_fees = list(
        Fee.objects.select_related('course').filter(
            student=student,
            course_id__in=active_course_ids,
        )
    )

    blocked_fees = [fee for fee in unpaid_fees if fee.pending_amount > 0]
    if not blocked_fees:
        return None

    course_labels = []
    for fee in blocked_fees[:3]:
        course_name = fee.course.course_name if fee.course else 'Course'
        course_labels.append(f"{course_name} (₹{_format_inr_amount(fee.pending_amount)} pending)")

    more_count = len(blocked_fees) - len(course_labels)
    course_summary = ", ".join(course_labels)
    if more_count > 0:
        course_summary = f"{course_summary} and {more_count} more"

    return (
        "Student cannot be deleted until payment is complete. "
        f"Pending fee: {course_summary}."
    )



def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    block_reason = _get_student_delete_block_reason(student)
    if block_reason:
        messages.error(request, f"{student.student_name}: {block_reason}")
        return redirect("students")
    student.delete()
    messages.success(request, "Student deleted successfully")
    return redirect("students")



def student_edit(request, id):
    student = get_object_or_404(Student, id=id)

    if request.method == "POST":
        department_ids = request.POST.getlist("departments[]")
        course_ids = request.POST.getlist("courses[]")
        new_status = request.POST.get("status")
        enrolled_date_value = request.POST.get("enrolled_date")
        parsed_enrolled_date = parse_date(enrolled_date_value) if enrolled_date_value else student.enrolled_date

        student.student_name = request.POST.get("student_name")
        student.email = request.POST.get("email")
        student.mobile = request.POST.get("mobile")
        student.status = new_status
        student.house_name = request.POST.get("house_name")
        student.place = request.POST.get("place")
        student.district = request.POST.get("district")
        student.state = request.POST.get("state")
        student.location = request.POST.get("location")
        student.enrolled_date = parsed_enrolled_date

        student.save()

        # Store existing batch assignments
        existing_batches = {}
        for sc in StudentCourse.objects.filter(student=student):
            existing_batches[sc.course_id] = sc.batch

        if new_status == "dropped":
            existing_batches = {}

        StudentCourse.objects.filter(student=student).delete()
        for dept_id, course_id in zip(department_ids, course_ids):
            if dept_id and course_id:
                StudentCourse.objects.create(
                    student=student,
                    department_id=dept_id,
                    course_id=int(course_id),
                    batch=existing_batches.get(int(course_id))
                )

        sync_student_fee(student)
        return redirect("students")

    context = {
        "student": student,
        "departments": Department.objects.all(),
        "courses": Course.objects.all(),
        "existing_courses": StudentCourse.objects.filter(student=student),
    }
    return render(request, "student_edit.html", context)



from datetime import timedelta
def students_list(request):
    q = request.GET.get('q', '')
    department = request.GET.get('department')
    course = request.GET.get('course')
    status = request.GET.get('status')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    batch_status = request.GET.get('batch_status')
    course_count = request.GET.get('course_count')

    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page', 1)

    students_qs = Student.objects.prefetch_related(
        'student_courses__course', 'student_courses__department', 'student_courses__batch'
    ).order_by('-id')
    
    # Trainer sees only students from their batches
    user_profile = request.user.userprofile
    if user_profile.role == 'trainer':
        my_batches = Batch.objects.filter(trainer=user_profile)
        students_qs = students_qs.filter(student_courses__batch__in=my_batches).distinct()

    #  SEARCH
    if q:
        students_qs = students_qs.filter(
            Q(student_name__icontains=q) |
            Q(email__icontains=q) |
            Q(mobile__icontains=q)
        )

    #  FILTERS
    if department:
        students_qs = students_qs.filter(student_courses__department_id=department).distinct()

    if course:
        students_qs = students_qs.filter(student_courses__course_id=course).distinct()

    if status:
        students_qs = students_qs.filter(status=status)

    # Batch Status Filter
    if batch_status == 'without_batch':
        students_qs = students_qs.filter(student_courses__batch__isnull=True).distinct()
    elif batch_status == 'with_batch':
        students_qs = students_qs.filter(student_courses__batch__isnull=False).distinct()
    elif batch_status == 'partial_batch':
        from django.db.models import Count, Q
        students_qs = students_qs.annotate(
            total_courses=Count('student_courses', distinct=True),
            assigned_batches=Count('student_courses', filter=Q(student_courses__batch__isnull=False), distinct=True)
        ).filter(total_courses__gt=1, assigned_batches__gt=0, assigned_batches__lt=F('total_courses'))

    # Course Count Filter
    if course_count == 'multiple':
        from django.db.models import Count
        students_qs = students_qs.annotate(course_count=Count('student_courses')).filter(course_count__gt=1)
    elif course_count == 'single':
        from django.db.models import Count
        students_qs = students_qs.annotate(course_count=Count('student_courses')).filter(course_count=1)

    if from_date and to_date:
        students_qs = students_qs.filter(
            enrolled_date__range=[from_date, to_date]
        )

    elif from_date and not to_date:
        date = parse_date(from_date)
        students_qs = students_qs.filter(
            enrolled_date__range=[
                date,
                date + timedelta(days=1)
            ]
        )

    elif to_date and not from_date:
        date = parse_date(to_date)
        students_qs = students_qs.filter(
            enrolled_date__range=[
                date,
                date + timedelta(days=1)
            ]
        )


    #  PAGINATION (THIS IS THE KEY)
    paginator = Paginator(students_qs, per_page)
    page_obj = paginator.get_page(page_number)

    return render(request, 'student.html', {
    'students': page_obj,
    'page_obj': page_obj,
    'paginator': paginator,
    'per_page': per_page,

    'courses': Course.objects.all(),
    'departments': Department.objects.all(),
    'statuses': [choice[0] for choice in Student.STATUS_CHOICES],

    'all_batches': Batch.objects.select_related('course')
})



def student_view(request, student_id):
    from .models import Assignment, AssignmentSubmission, Attendance
    import calendar
    from datetime import date
    from decimal import Decimal
    
    student = get_object_or_404(
        Student.objects.prefetch_related(
            'student_courses__course',
            'student_courses__batch',
            'fees__payments',
            'fees__course',
            'fees__installments__payments',
        ),
        id=student_id
    )
    
    sync_student_fee(student)
    
    active_course_ids = list(
        student.student_courses.values_list('course_id', flat=True)
    )
    all_fees = student.fees.select_related('course').prefetch_related('payments', 'installments__payments').filter(
        course_id__in=active_course_ids
    ).order_by('-id')
    fee = all_fees.first() if all_fees.exists() else None
    payment_history = fee.payments.all() if fee else []
    billing_overview = {
        'total_fee': Decimal('0.00'),
        'total_paid': Decimal('0.00'),
        'total_pending': Decimal('0.00'),
        'course_count': 0,
    }

    for student_fee in all_fees:
        billing_overview['total_fee'] += student_fee.total_fee or Decimal('0.00')
        billing_overview['total_paid'] += student_fee.paid_amount or Decimal('0.00')
        billing_overview['total_pending'] += student_fee.pending_amount or Decimal('0.00')
        if student_fee.course_id:
            billing_overview['course_count'] += 1

        installment_rows = list(student_fee.installments.all().order_by('installment_number'))
        locked_installments = []
        editable_installments = []
        locked_pending_total = Decimal('0.00')

        for installment in installment_rows:
            installment.has_payments = installment.payments.exists()
            if installment.has_payments:
                locked_installments.append(installment)
                locked_pending_total += installment.pending_amount or Decimal('0.00')
            else:
                editable_installments.append(installment)

        custom_target_amount = (student_fee.pending_amount or Decimal('0.00')) - locked_pending_total
        if custom_target_amount < Decimal('0.00'):
            custom_target_amount = Decimal('0.00')

        student_fee.installment_rows = installment_rows
        student_fee.locked_installments = locked_installments
        student_fee.editable_installments = editable_installments
        student_fee.custom_installment_target = custom_target_amount

    # Get student courses with batches
    student_courses_with_batch = student.student_courses.filter(batch__isnull=False).select_related('course', 'batch')
    no_batch = not student_courses_with_batch.exists()

    month_param = (request.GET.get('attendance_month') or '').strip()
    today = date.today()
    try:
        selected_month_date = date.fromisoformat(f"{month_param}-01") if month_param else today.replace(day=1)
    except ValueError:
        selected_month_date = today.replace(day=1)

    attendance_records = (
        Attendance.objects.filter(
            student=student,
            date__year=selected_month_date.year,
            date__month=selected_month_date.month,
        )
        .select_related('batch')
        .order_by('date')
    )
    attendance_by_date = {record.date: record for record in attendance_records}

    month_calendar = []
    month_matrix = calendar.Calendar(firstweekday=6).monthdatescalendar(
        selected_month_date.year,
        selected_month_date.month,
    )
    for week in month_matrix:
        week_days = []
        for day_date in week:
            record = attendance_by_date.get(day_date)
            week_days.append({
                'date': day_date,
                'day': day_date.day,
                'in_month': day_date.month == selected_month_date.month,
                'is_today': day_date == today,
                'status': record.status if record else '',
                'duration': record.duration if record else '',
                'batch_name': record.batch.batch_name if record else '',
                'remarks': record.remarks if record else '',
            })
        month_calendar.append(week_days)

    previous_month = (
        selected_month_date.replace(day=1) - timedelta(days=1)
    ).replace(day=1)
    next_month = (
        selected_month_date.replace(day=28) + timedelta(days=4)
    ).replace(day=1)

    attendance_summary = {
        'total': attendance_records.count(),
        'present': attendance_records.filter(status='present').count(),
        'late': attendance_records.filter(status='late').count(),
        'leave': attendance_records.filter(status='leave').count(),
        'absent': attendance_records.filter(status='absent').count(),
    }
    
    # Calculate progress for each course/batch
    course_progress_list = []
    
    if not no_batch:
        for sc in student_courses_with_batch:
            batch = sc.batch
            course = sc.course
            
            # Attendance
            attendance_qs = Attendance.objects.filter(student=student, batch=batch)
            total_attendance = attendance_qs.count()
            present_count = attendance_qs.filter(status__in=["present", "late"]).count()
            attendance_percent = round((present_count / total_attendance) * 100) if total_attendance > 0 else 0
            
            # Assignments
            total_assignments = Assignment.objects.filter(batch=batch).count()
            submitted_count = AssignmentSubmission.objects.filter(
                student=student,
                assignment__batch=batch,
                status='submitted'
            ).count()
            assignment_percent = round((submitted_count / total_assignments) * 100) if total_assignments > 0 else 0
            
            # Exams
            exams_qs = Exam.objects.filter(batch=batch)
            total_exams = exams_qs.count()
            exam_performances = ExamPerformance.objects.filter(
                student=student,
                exam__batch=batch,
                score__isnull=False
            ).select_related('exam')
            
            scored_exam_count = exam_performances.count()
            exam_scored_marks = sum((performance.score or 0) for performance in exam_performances)
            exam_out_of_marks = sum(exam.max_marks for exam in exams_qs)
            exam_percent = round((float(exam_scored_marks) / float(exam_out_of_marks)) * 100) if exam_out_of_marks > 0 else 0
            
            # Overall
            metrics = []
            if total_attendance > 0:
                metrics.append(attendance_percent)
            if total_assignments > 0:
                metrics.append(assignment_percent)
            if total_exams > 0:
                metrics.append(exam_percent)
            overall_percent = round(sum(metrics) / len(metrics)) if metrics else 0
            
            course_progress_list.append({
                'course_name': course.course_name,
                'batch_name': batch.batch_name,
                'attendance_percent': max(0, min(attendance_percent, 100)),
                'attendance_present_count': present_count,
                'attendance_total_count': total_attendance,
                'assignment_percent': max(0, min(assignment_percent, 100)),
                'assignment_submitted_count': submitted_count,
                'assignment_total_count': total_assignments,
                'exam_percent': max(0, min(exam_percent, 100)),
                'exam_total_count': total_exams,
                'exam_scored_count': scored_exam_count,
                'exam_scored_marks': exam_scored_marks,
                'exam_out_of_marks': exam_out_of_marks,
                'overall_percent': max(0, min(overall_percent, 100)),
            })

    context = {
        'student': student,
        'all_fees': all_fees,
        'fee': fee,
        'payment_history': payment_history,
        'billing_overview': billing_overview,
        'no_batch': no_batch,
        'course_progress_list': course_progress_list,
        'attendance_calendar': month_calendar,
        'attendance_selected_month': selected_month_date,
        'attendance_previous_month': previous_month.strftime('%Y-%m'),
        'attendance_next_month': next_month.strftime('%Y-%m'),
        'attendance_summary': attendance_summary,
    }

    return render(request, 'student_view.html', context)

from django.http import JsonResponse
from django.views.decorators.http import require_POST


@login_required
@require_POST
def customize_fee_installments(request, fee_id):
    fee = get_object_or_404(
        Fee.objects.select_related('student', 'course').prefetch_related('installments__payments'),
        id=fee_id
    )

    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role == 'trainer':
        messages.error(request, "Trainers cannot change installment plans.")
        return redirect('student_view', student_id=fee.student_id)

    if request.POST.get('reset_to_course') == '1':
        for installment in fee.installments.all():
            if installment.payments.exists():
                continue
            installment.delete()

        fee.custom_installment_plan = False
        fee.save(update_fields=['custom_installment_plan'])
        sync_student_fee(fee.student)
        messages.success(request, f"Installments reset to the default course plan for {fee.course.course_name}.")
        return redirect('student_view', student_id=fee.student_id)

    amount_inputs = request.POST.getlist('installment_amount[]')
    due_date_inputs = request.POST.getlist('installment_due_date[]')

    submitted_rows = []
    for amount_raw, due_date_raw in zip(amount_inputs, due_date_inputs):
        amount_text = (amount_raw or '').strip()
        due_date_text = (due_date_raw or '').strip()

        if not amount_text and not due_date_text:
            continue

        if not amount_text or not due_date_text:
            messages.error(request, "Each installment row needs both amount and due date.")
            return redirect('student_view', student_id=fee.student_id)

        try:
            amount = Decimal(amount_text)
        except Exception:
            messages.error(request, "Installment amount must be a valid number.")
            return redirect('student_view', student_id=fee.student_id)

        if amount <= 0:
            messages.error(request, "Installment amount must be greater than zero.")
            return redirect('student_view', student_id=fee.student_id)

        due_date = parse_date(due_date_text)
        if due_date is None:
            messages.error(request, "Installment due date is invalid.")
            return redirect('student_view', student_id=fee.student_id)

        submitted_rows.append({
            'amount': amount.quantize(Decimal('0.01')),
            'due_date': due_date,
        })

    locked_installments = []
    locked_pending_total = Decimal('0.00')
    max_locked_number = 0
    for installment in fee.installments.all().order_by('installment_number'):
        if installment.payments.exists():
            locked_installments.append(installment)
            locked_pending_total += installment.pending_amount or Decimal('0.00')
            max_locked_number = max(max_locked_number, installment.installment_number)

    target_amount = (fee.pending_amount or Decimal('0.00')) - locked_pending_total
    if target_amount < Decimal('0.00'):
        target_amount = Decimal('0.00')

    submitted_total = sum((row['amount'] for row in submitted_rows), Decimal('0.00')).quantize(Decimal('0.01'))
    target_amount = target_amount.quantize(Decimal('0.01'))

    if target_amount > 0 and not submitted_rows:
        messages.error(request, "Add at least one installment for the remaining pending amount.")
        return redirect('student_view', student_id=fee.student_id)

    if submitted_total != target_amount:
        messages.error(
            request,
            f"The editable installments must total ₹{target_amount}. You entered ₹{submitted_total}."
        )
        return redirect('student_view', student_id=fee.student_id)

    with transaction.atomic():
        for installment in fee.installments.all():
            if installment.payments.exists():
                continue
            installment.delete()

        next_number = max_locked_number
        for row in submitted_rows:
            next_number += 1
            FeeInstallment.objects.create(
                fee=fee,
                installment_number=next_number,
                amount=row['amount'],
                due_date=row['due_date'],
            )

        fee.custom_installment_plan = True
        fee.save(update_fields=['custom_installment_plan'])
        fee.recalculate_payment_status()

    messages.success(request, f"Custom installment plan saved for {fee.course.course_name}.")
    return redirect('student_view', student_id=fee.student_id)


@require_POST
def bulk_update_student_status(request):
    status = request.POST.get('status')
    student_ids = request.POST.getlist('student_ids')

    if not status or not student_ids:
        return JsonResponse({'success': False, 'error': 'Invalid data'})

    Student.objects.filter(id__in=student_ids).update(status=status)

    if status == 'dropped':
        _clear_student_batch_assignments(student_ids)

    return JsonResponse({'success': True})


@require_POST
def bulk_delete_students(request):
    student_ids = request.POST.getlist('student_ids')

    if not student_ids:
        return JsonResponse({'success': False, 'error': 'No students selected'})

    students = Student.objects.filter(id__in=student_ids)
    students_by_id = {str(student.id): student for student in students}

    ordered_students = []
    for student_id in student_ids:
        student = students_by_id.get(str(student_id))
        if not student:
            return JsonResponse({'success': False, 'error': 'One or more selected students were not found'})
        ordered_students.append(student)

    for student in ordered_students:
        block_reason = _get_student_delete_block_reason(student)
        if block_reason:
            return JsonResponse({
                'success': False,
                'error': f"{student.student_name}: {block_reason}"
            })

    Student.objects.filter(id__in=[student.id for student in ordered_students]).delete()

    return JsonResponse({'success': True})
#=======================================SETTINGS=========================================================================

#============DEPARTMENT==================

def department_view(request):
    if request.method == "POST" and 'department_name' in request.POST:
        department_name = request.POST.get('department_name').strip()

        if not department_name:
            messages.error(request, "Department name cannot be empty.")
            return redirect('department')

        try:
            Department.objects.create(
                department_name=department_name
            )
            messages.success(request, "Department added successfully.")
        except IntegrityError:
            messages.error(
                request,
                "This department already exists."
            )

        return redirect('department')

    departments = Department.objects.all()

    return render(
        request,
        'department.html',
        {'departments': departments}
    )



def delete_department(request, pk):
    department = get_object_or_404(Department, pk=pk)

    try:
        department.delete()
        messages.success(request, "Department deleted successfully.")
    except ProtectedError:
        messages.error(
            request,
            "Cannot delete this department because it is in use."
        )

    return redirect('department')

#============COURSE==================

# Add Course (Department FIRST)
from django.db import IntegrityError

def course_view(request):
    departments = Department.objects.all()

    if request.method == "POST":
        department_id = request.POST.get('department')
        course_name = request.POST.get('course_name')
        fees = request.POST.get('fees')
        advance_payment_amount = request.POST.get('advance_payment_amount')
        duration_months = request.POST.get('duration_months')
        installment_enabled = request.POST.get('installment_enabled') == 'on'
        installment_count = request.POST.get('installment_count', 1)
        installment_interval_days = request.POST.get('installment_interval_days', 30)

        if not department_id or not course_name:
            messages.error(request, "Department and course name are required.")
            return redirect('course')

        course_name = course_name.strip()

        try:
            Course.objects.create(
                course_name=course_name,
                department_id=department_id,
                fees=fees if fees else None,
                advance_payment_amount=advance_payment_amount if advance_payment_amount else '1000.00',
                duration_months=int(duration_months) if duration_months else 3,
                installment_enabled=installment_enabled,
                installment_count=int(installment_count) if installment_enabled else 1,
                installment_interval_days=int(installment_interval_days) if installment_enabled else 30
            )
            messages.success(request, "Course added successfully.")

        except IntegrityError:
            messages.error(
                request,
                "This course already exists in the selected department."
            )

        return redirect('course')

    courses = Course.objects.select_related('department')

    return render(
        request,
        'course.html',
        {
            'departments': departments,
            'courses': courses
        }
    )


@login_required
def edit_course(request):
    if request.method == "POST":
        course_id = request.POST.get('course_id')
        department_id = request.POST.get('department')
        course_name = request.POST.get('course_name')
        fees = request.POST.get('fees')
        advance_payment_amount = request.POST.get('advance_payment_amount')
        duration_months = request.POST.get('duration_months')
        installment_enabled = request.POST.get('installment_enabled') == 'on'
        installment_count = request.POST.get('installment_count', 1)
        installment_interval_days = request.POST.get('installment_interval_days', 30)
        
        if not course_id or not department_id or not course_name:
            messages.error(request, "All required fields must be filled.")
            return redirect('course')
        
        try:
            course = Course.objects.get(id=course_id)
            
            existing_course = Course.objects.filter(
                course_name__iexact=course_name.strip(),
                department_id=department_id
            ).exclude(id=course_id).first()
            
            if existing_course:
                messages.error(request, "A course with this name already exists in the selected department.")
                return redirect('course')
            
            course.course_name = course_name.strip()
            course.department_id = department_id
            course.fees = fees if fees else None
            course.advance_payment_amount = advance_payment_amount if advance_payment_amount else '1000.00'
            course.duration_months = int(duration_months) if duration_months else 3
            course.installment_enabled = installment_enabled
            course.installment_count = int(installment_count) if installment_enabled else 1
            course.installment_interval_days = int(installment_interval_days) if installment_enabled else 30
            course.save()

            for fee in Fee.objects.filter(course=course):
                fee.recalculate_payment_status()
            
            messages.success(request, "Course updated successfully.")
            
        except Course.DoesNotExist:
            messages.error(request, "Course not found.")
        except IntegrityError:
            messages.error(request, "An error occurred while updating the course.")
        
        return redirect('course')
    
    return redirect('course')


from django.db.models.deletion import ProtectedError

def delete_course(request, pk):
    course = get_object_or_404(Course, pk=pk)

    # Optional: explicit check (clear message)
    if LeadCourse.objects.filter(course=course).exists():
        messages.error(
            request,
            "Cannot delete this course because it is already in use."
        )
        return redirect('course')

    try:
        course.delete()
        messages.success(request, "Course deleted successfully.")
    except ProtectedError:
        messages.error(
            request,
            "Cannot delete this course because it is in use."
        )

    return redirect('course')

def get_courses_by_department(request):
    department_id = request.GET.get('department_id')
    courses = Course.objects.filter(department_id=department_id).values('id', 'course_name')
    return JsonResponse(list(courses), safe=False)

def get_modules_by_course(request):
    course_id = request.GET.get('course_id')
    modules = Module.objects.filter(course_id=course_id).values('id', 'module_name')
    return JsonResponse(list(modules), safe=False)

#============SOURCE==================


def source_view(request):
    if request.method == 'POST':
        source_name = _normalize_source_name(request.POST.get('source_name'))
        edit_source_id = request.POST.get('edit_source_id')

        if source_name:
            if edit_source_id:
                source = get_object_or_404(Source, pk=edit_source_id)
                duplicate_exists = Source.objects.filter(
                    source_name__iexact=source_name
                ).exclude(pk=source.pk).exists()

                if duplicate_exists:
                    messages.error(request, "A source with this name already exists.")
                else:
                    old_name = source.source_name
                    source.source_name = source_name
                    source.save(update_fields=['source_name'])
                    messages.success(
                        request,
                        f'Source "{old_name}" updated to "{source_name}" successfully.'
                    )
            else:
                source, created = Source.objects.get_or_create(source_name=source_name)
                if created:
                    messages.success(request, "Source added successfully.")
                else:
                    messages.info(request, "This source already exists.")

        return redirect('source')

    context = {
        'sources': Source.objects.all(),
    }
    return render(request, 'source.html', context)


def delete_source(request, pk):
    source = get_object_or_404(Source, pk=pk)

    try:
        source.delete()
        messages.success(request, "Source deleted successfully.")
    except ProtectedError:
        messages.error(
            request,
            "Cannot delete this source because it is used in Student Enquiries."
        )

    return redirect('source')


#============STATUS==================


def status_view(request):
    if request.method == 'POST':
        status_name = (request.POST.get('status_name') or '').strip()
        edit_status_id = request.POST.get('edit_status_id')

        if status_name:
            if edit_status_id:
                status = get_object_or_404(Status, pk=edit_status_id)
                duplicate_exists = Status.objects.filter(
                    status_name__iexact=status_name
                ).exclude(pk=status.pk).exists()

                if duplicate_exists:
                    messages.error(request, "A status with this name already exists.")
                else:
                    old_name = status.status_name
                    status.status_name = status_name
                    status.save(update_fields=['status_name'])
                    messages.success(
                        request,
                        f'Status "{old_name}" updated to "{status_name}" successfully.'
                    )
            else:
                status, created = Status.objects.get_or_create(status_name=status_name)
                if created:
                    messages.success(request, "Status added successfully.")
                else:
                    messages.info(request, "This status already exists.")

        return redirect('status')   # redirect after POST (good practice)

    statuses = Status.objects.all()

    context = {
        'statuses': statuses
    }
    return render(request, 'status.html', context)

def delete_status(request, pk):
    status = get_object_or_404(Status, pk=pk)

    try:
        status.delete()
        messages.success(request, "Status deleted successfully.")
    except ProtectedError:
        messages.error(
            request,
            "Cannot delete this status because it is used in Student Enquiries."
        )

    return redirect('status')




#============BATCH==================


@login_required
def get_eligible_trainers(request):
    """
    API endpoint to fetch trainers who are specialized in the given department and course.
    Only returns trainers whose specialization matches both department and course.
    """
    department_id = request.GET.get('department_id')
    course_id = request.GET.get('course_id')
    
    if not department_id or not course_id:
        return JsonResponse({'trainers': []})
    
    # Get trainers who have specialization in BOTH the selected department AND course
    eligible_trainers = UserProfile.objects.filter(
        role='trainer',
        specializations__department_id=department_id,
        specializations__course_id=course_id
    ).select_related('user').distinct()
    
    trainer_data = [
        {
            'id': trainer.id,
            'name': trainer.user.get_full_name() or trainer.user.username,
            'username': trainer.user.username
        }
        for trainer in eligible_trainers
    ]
    
    return JsonResponse({'trainers': trainer_data})

@login_required
def batch_view(request):
    from django.db.models import Count
    user_profile = request.user.userprofile
    
    batches = Batch.objects.select_related(
        'department', 'course', 'trainer__user'
    ).annotate(
        enrolled_count=Count('studentcourse', distinct=True)
    ).order_by('-created_at')
    
    # Trainer sees only his batches
    if user_profile.role == "trainer":
        batches = batches.filter(trainer=user_profile)

    search = request.GET.get('search')
    if search:
        batches = batches.filter(
            Q(batch_name__icontains=search) |
            Q(course__course_name__icontains=search) |
            Q(department__department_name__icontains=search)
        )

    department_id = request.GET.get('department')
    if department_id:
        batches = batches.filter(department_id=department_id)

    course_id = request.GET.get('course')
    if course_id:
        batches = batches.filter(course_id=course_id)
    
    # Get courses grouped by department for the dropdown
    departments_with_courses = []
    for dept in Department.objects.all():
        courses = Course.objects.filter(department=dept)
        if courses.exists():
            departments_with_courses.append({
                'department': dept,
                'courses': courses
            })

    context = {
        'batches': batches,
        'departments': Department.objects.all(),
        'courses': Course.objects.select_related('department'),
        'departments_with_course':departments_with_courses,
        'trainers': UserProfile.objects.filter(role='trainer'),
    }
    return render(request, 'batch.html', context)

@login_required
def add_batch(request):
    user_profile = request.user.userprofile
    
    # Prevent trainers from adding batches
    if user_profile.role == "trainer":
        messages.error(request, "You don't have permission to add batches.")
        return redirect('batch')
    
    if request.method == 'POST':
        try:
            batch_name = request.POST.get('batch_name').strip()
            department_id = request.POST.get('department')
            course_id = request.POST.get('course')
            trainer_id = request.POST.get('trainer')

            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')

            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')

            mode = request.POST.get('mode')
            max_students = request.POST.get('max_students', 30)

            if not all([batch_name, department_id, course_id, trainer_id, start_date, start_time, mode]):
                messages.error(request, "All mandatory fields are required.")
                return redirect('add_batch')

            # Validate that the trainer is specialized in the selected department and course
            trainer = get_object_or_404(UserProfile, id=trainer_id, role='trainer')
            has_specialization = TrainerSpecialization.objects.filter(
                trainer=trainer,
                department_id=department_id,
                course_id=course_id
            ).exists()
            
            if not has_specialization:
                messages.error(request, f"Selected trainer is not specialized in {Course.objects.get(id=course_id).course_name} ({Department.objects.get(id=department_id).department_name}). Please select a qualified trainer.")
                return redirect('add_batch')
       
            Batch.objects.create(
                batch_name=batch_name,
                department_id=department_id,
                course_id=course_id,
                trainer_id=trainer_id,
                start_date=start_date,
                end_date=end_date,
                start_time=start_time,
                end_time=end_time,
                mode=mode,
                no_of_students=max_students,
                status='Active'
            )

            messages.success(request, "Batch added successfully.")

        except IntegrityError:
            messages.error(request, "Batch name already exists.")

        return redirect('add_batch')
    context = {
        'departments': Department.objects.all(),
        'courses': Course.objects.select_related('department'),
        'trainers': UserProfile.objects.filter(role='trainer'),
    }
    return render(request, 'add_batch.html', context)

@login_required
def edit_batch(request, batch_id):
    user_profile = request.user.userprofile
    
    # Prevent trainers from editing batches
    if user_profile.role == "trainer":
        messages.error(request, "You don't have permission to edit batches.")
        return redirect('batch')
    
    batch = get_object_or_404(Batch, id=batch_id)
    if request.method == 'POST':
        department_id = request.POST.get('department')
        course_id = request.POST.get('course')
        trainer_id = request.POST.get('trainer')
        
        # Validate that the trainer is specialized in the selected department and course
        if trainer_id:
            trainer = get_object_or_404(UserProfile, id=trainer_id, role='trainer')
            has_specialization = TrainerSpecialization.objects.filter(
                trainer=trainer,
                department_id=department_id,
                course_id=course_id
            ).exists()
            
            if not has_specialization:
                messages.error(request, f"Selected trainer is not specialized in {Course.objects.get(id=course_id).course_name} ({Department.objects.get(id=department_id).department_name}). Please select a qualified trainer.")
                return redirect('edit_batch', batch_id=batch_id)
        
        batch.batch_name = request.POST.get('batch_name')
        batch.department_id = department_id
        batch.course_id = course_id
        batch.trainer_id = trainer_id
        batch.start_date = request.POST.get('start_date')
        batch.end_date = request.POST.get('end_date')
        batch.start_time = request.POST.get('start_time')
        batch.end_time = request.POST.get('end_time') or None
        batch.mode = request.POST.get('mode')
        batch.no_of_students = request.POST.get('max_students', 30)
        batch.status = request.POST.get('status')
        batch.save()
        messages.success(request, "Batch updated successfully.")
        return redirect('batch')

    context = {
        'batch': batch,
        'departments': Department.objects.all(),
        'courses': Course.objects.select_related('department'),
        'trainers': UserProfile.objects.filter(role='trainer'),
    }
    return render(request, 'add_batch.html', context)

@login_required
def delete_batch(request, batch_id):
    user_profile = request.user.userprofile
    
    # Prevent trainers from deleting batches
    if user_profile.role == "trainer":
        messages.error(request, "You don't have permission to delete batches.")
        return redirect('batch')
    
    batch = get_object_or_404(Batch, id=batch_id)
    batch.delete()
    messages.success(request, "Batch deleted successfully.")
    return redirect('batch')

@login_required
def batch_students(request, batch_id):
    batch = get_object_or_404(Batch, id=batch_id)
    students = Student.objects.filter(
        student_courses__batch=batch,
        student_courses__course=batch.course
    ).distinct()
    return render(request, 'batch_students.html', {'batch': batch, 'students': students})


def _queue_batch_assignment_mail(student, batch):
    if not student or not batch:
        return

    email = (student.email or "").strip() or None
    pending_mail, created = PendingMail.objects.get_or_create(
        student=student,
        mail_type='batch',
        batch=batch,
        defaults={
            'email': email,
            'last_error': None,
        }
    )

    if not created:
        pending_mail.email = email
        pending_mail.last_error = None
        pending_mail.save(update_fields=['email', 'last_error'])


def _queue_payment_success_mail(student, amount, payment_date, fee=None):
    if not student:
        return

    PendingMail.objects.create(
        student=student,
        email=(student.email or "").strip() or None,
        mail_type='payment',
        amount=amount,
        payment_date=payment_date,
        fee=fee,
    )


def _get_pending_mail_redirect_name(mail_type):
    redirect_map = {
        'batch': 'mail_pending_list',
        'payment': 'payment_email_pending_list',
        'overdue': 'overdue_email_pending_list',
    }
    return redirect_map.get(mail_type, 'mail_pending_list')


def _get_list_per_page(request, default=20):
    try:
        per_page = int(request.GET.get('per_page', default))
    except (TypeError, ValueError):
        per_page = default

    if per_page not in {10, 20, 50, 100}:
        per_page = default
    return per_page


def _get_overdue_fee_queryset():
    today = timezone.localdate()
    batch_assignment_subquery = StudentCourse.objects.filter(
        student_id=OuterRef('student_id'),
        course_id=OuterRef('course_id'),
        batch__isnull=False,
    ).order_by('-id')
    latest_sent_subquery = MailLog.objects.filter(
        fee_id=OuterRef('pk'),
        mail_type='overdue',
        status='sent',
    ).order_by('-sent_at', '-id')
    latest_payment_subquery = Payment.objects.filter(
        fee_id=OuterRef('pk')
    ).order_by('-payment_date', '-id')

    return Fee.objects.select_related('student', 'course', 'course__department').annotate(
        has_batch_assigned=Exists(batch_assignment_subquery),
        current_batch_id=Subquery(batch_assignment_subquery.values('batch_id')[:1]),
        latest_overdue_sent_at=Subquery(latest_sent_subquery.values('sent_at')[:1]),
        latest_payment_date=Subquery(latest_payment_subquery.values('payment_date')[:1]),
    ).filter(
        has_batch_assigned=True,
        due_date__lt=today,
        total_fee__gt=F('paid_amount'),
    )


def _sync_overdue_pending_mails(fee_ids=None):
    normalized_fee_ids = None
    if fee_ids is not None:
        normalized_fee_ids = [fee_id for fee_id in fee_ids if fee_id]
        if not normalized_fee_ids:
            return

    overdue_fees_qs = _get_overdue_fee_queryset()
    if normalized_fee_ids is not None:
        overdue_fees_qs = overdue_fees_qs.filter(id__in=normalized_fee_ids)

    overdue_fees = list(overdue_fees_qs)
    eligible_fee_ids = {fee.id for fee in overdue_fees}

    existing_qs = PendingMail.objects.select_related('fee').filter(mail_type='overdue')
    if normalized_fee_ids is not None:
        existing_qs = existing_qs.filter(fee_id__in=normalized_fee_ids)

    if eligible_fee_ids:
        existing_qs.exclude(fee_id__in=eligible_fee_ids).delete()
    else:
        existing_qs.delete()

    existing_by_fee_id = {
        item.fee_id: item
        for item in existing_qs.filter(fee_id__in=eligible_fee_ids)
    }

    batch_ids = {fee.current_batch_id for fee in overdue_fees if fee.current_batch_id}
    batches_by_id = Batch.objects.in_bulk(batch_ids)

    for fee in overdue_fees:
        existing_pending = existing_by_fee_id.get(fee.id)
        batch = batches_by_id.get(fee.current_batch_id)
        email = (fee.student.email or "").strip() or None

        if existing_pending:
            update_fields = []
            if existing_pending.student_id != fee.student_id:
                existing_pending.student = fee.student
                update_fields.append('student')
            if existing_pending.email != email:
                existing_pending.email = email
                update_fields.append('email')
            if existing_pending.batch_id != fee.current_batch_id:
                existing_pending.batch = batch
                update_fields.append('batch')
            if existing_pending.amount != fee.pending_amount:
                existing_pending.amount = fee.pending_amount
                update_fields.append('amount')
            if existing_pending.fee_id != fee.id:
                existing_pending.fee = fee
                update_fields.append('fee')
            if update_fields:
                existing_pending.save(update_fields=update_fields)
            continue

        latest_sent_at = getattr(fee, 'latest_overdue_sent_at', None)
        latest_payment_date = getattr(fee, 'latest_payment_date', None)
        should_queue = not latest_sent_at
        if latest_sent_at and latest_payment_date and latest_payment_date > latest_sent_at.date():
            should_queue = True

        if not should_queue:
            continue

        PendingMail.objects.create(
            student=fee.student,
            email=email,
            mail_type='overdue',
            batch=batch,
            amount=fee.pending_amount,
            fee=fee,
        )


def _build_batch_email_content(pending_mail):
    student = pending_mail.student
    batch = pending_mail.batch
    company_name = getattr(settings, "MAIL_COMPANY_NAME", "") or getattr(settings, "COMPANY_NAME", "") or "FlowDesk CRM"

    student_course = None
    if student and batch and batch.course_id:
        student_course = StudentCourse.objects.select_related('course', 'course__department').filter(
            student=student,
            course=batch.course
        ).first()
    elif student:
        student_course = StudentCourse.objects.select_related('course', 'course__department').filter(
            student=student
        ).order_by('-id').first()

    course = None
    if batch and batch.course:
        course = batch.course
    elif student_course:
        course = student_course.course

    department = None
    if batch and batch.department:
        department = batch.department
    elif course and getattr(course, 'department', None):
        department = course.department

    fee_record = None
    if student and course:
        fee_record = Fee.objects.select_related('course').filter(student=student, course=course).order_by('-id').first()
    elif student:
        fee_record = Fee.objects.select_related('course').filter(student=student).order_by('-id').first()

    if fee_record:
        installments = fee_record.installments.all().order_by('installment_number')
    else:
        installments = []
        fee_record = {
            'total_fee': 0,
            'paid_amount': 0,
            'pending_amount': 0,
            'due_date': None,
            'get_fee_status_display': 'Pending',
        }

    modules = Module.objects.none()
    if course:
        modules = Module.objects.filter(course=course).prefetch_related('topic_set').order_by('id')

    context = {
        'company_name': company_name,
        'student': student,
        'batch': batch,
        'course': course,
        'department': department,
        'fee_record': fee_record,
        'installments': installments,
        'modules': modules,
    }

    subject = "Batch Assignment Confirmation"
    html_body = render_to_string('email/batch_assignment.html', context)
    text_body = strip_tags(html_body).strip() or "Batch assignment confirmation"
    return subject, text_body, html_body


def _build_payment_email_content(pending_mail):
    student = pending_mail.student
    company_name = getattr(settings, "MAIL_COMPANY_NAME", "") or getattr(settings, "COMPANY_NAME", "") or "FlowDesk CRM"
    fee_record = getattr(pending_mail, 'fee', None)

    payment_qs = Payment.objects.select_related('fee', 'fee__course').filter(
        fee__student=student
    ).order_by('-payment_date', '-id')
    if fee_record:
        payment_qs = payment_qs.filter(fee=fee_record)
    if pending_mail.amount is not None:
        payment_qs = payment_qs.filter(amount_paid=pending_mail.amount)
    if pending_mail.payment_date:
        payment_qs = payment_qs.filter(payment_date=pending_mail.payment_date)
    payment = payment_qs.first()

    if not payment:
        payment = Payment.objects.select_related('fee', 'fee__course').filter(
            fee__student=student
        ).order_by('-payment_date', '-id').first()

    fee_record = payment.fee if payment else fee_record
    if not fee_record:
        fee_record = Fee.objects.select_related('course').filter(student=student).order_by('-id').first()

    if not fee_record:
        fee_record = {
            'course': {'course_name': 'Course'},
            'total_fee': 0,
            'paid_amount': 0,
            'pending_amount': 0,
        }

    if not payment:
        payment = {
            'amount_paid': pending_mail.amount if pending_mail.amount is not None else 0,
            'payment_date': pending_mail.payment_date,
            'get_payment_mode_display': 'N/A',
            'remarks': '',
        }

    context = {
        'company_name': company_name,
        'student': student,
        'payment': payment,
        'fee_record': fee_record,
    }

    subject = "Payment Confirmation"
    html_body = render_to_string('email/payment_confirmation.html', context)
    text_body = strip_tags(html_body).strip() or "Payment confirmation"
    return subject, text_body, html_body


def _build_overdue_email_content(pending_mail):
    student = pending_mail.student
    company_name = getattr(settings, "MAIL_COMPANY_NAME", "") or getattr(settings, "COMPANY_NAME", "") or "FlowDesk CRM"
    today = timezone.localdate()

    fee_record = getattr(pending_mail, 'fee', None)
    if not fee_record and student:
        fee_record = _get_overdue_fee_queryset().filter(student=student).order_by('due_date', 'id').first()

    if not fee_record and student:
        fee_record = Fee.objects.select_related('course', 'course__department').filter(
            student=student,
            total_fee__gt=F('paid_amount'),
        ).order_by('due_date', 'id').first()

    student_course = None
    if student and fee_record and fee_record.course_id:
        student_course = StudentCourse.objects.select_related('batch', 'department', 'course').filter(
            student=student,
            course=fee_record.course,
            batch__isnull=False,
        ).order_by('-id').first()

    overdue_installment = None
    installments = []
    if fee_record and hasattr(fee_record, 'installments'):
        installments = list(fee_record.installments.all().order_by('installment_number'))
        for installment in installments:
            if installment.pending_amount > 0 and installment.due_date and installment.due_date < today:
                overdue_installment = installment
                break

    if not fee_record:
        fee_record = {
            'course': {'course_name': 'Course'},
            'total_fee': 0,
            'paid_amount': 0,
            'pending_amount': getattr(pending_mail, 'amount', 0) or 0,
            'due_date': None,
            'get_fee_status_display': 'Overdue',
        }

    due_date = getattr(fee_record, 'due_date', None)
    days_overdue = 0
    if due_date and due_date < today:
        days_overdue = (today - due_date).days

    context = {
        'company_name': company_name,
        'student': student,
        'fee_record': fee_record,
        'installment': overdue_installment,
        'batch': getattr(student_course, 'batch', None) or getattr(pending_mail, 'batch', None),
        'course': getattr(fee_record, 'course', None),
        'days_overdue': days_overdue,
        'is_overdue': True,
    }

    subject = "Overdue Fee Reminder"
    html_body = render_to_string('email/fee_reminder.html', context)
    text_body = strip_tags(html_body).strip() or "Overdue fee reminder"
    return subject, text_body, html_body


def _send_pending_mail_now(pending_mail):
    student = pending_mail.student
    email = (pending_mail.email or "").strip()
    fee = getattr(pending_mail, 'fee', None)

    if pending_mail.mail_type == 'batch':
        subject, text_body, html_body = _build_batch_email_content(pending_mail)
    elif pending_mail.mail_type == 'payment':
        subject, text_body, html_body = _build_payment_email_content(pending_mail)
    elif pending_mail.mail_type == 'overdue':
        subject, text_body, html_body = _build_overdue_email_content(pending_mail)
    else:
        error_message = "Unsupported mail type."
        MailLog.objects.create(
            student=student,
            fee=fee,
            email=email or "",
            mail_type=pending_mail.mail_type,
            status='failed',
            message=error_message,
        )
        pending_mail.last_error = error_message
        pending_mail.save(update_fields=['last_error'])
        return False, error_message

    if not email:
        error_message = "Student email address is missing."
        MailLog.objects.create(
            student=student,
            fee=fee,
            email="",
            mail_type=pending_mail.mail_type,
            status='failed',
            message=error_message,
        )
        pending_mail.last_error = error_message
        pending_mail.save(update_fields=['last_error'])
        return False, error_message

    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", "") or "no-reply@flowdesk.local"

    try:
        message = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=from_email,
            to=[email],
        )
        message.attach_alternative(html_body, "text/html")
        message.send(fail_silently=False)
        MailLog.objects.create(
            student=student,
            fee=fee,
            email=email,
            mail_type=pending_mail.mail_type,
            status='sent',
            message='Email sent manually from Mail Management.',
        )
        pending_mail.delete()
        return True, "Mail sent successfully."
    except Exception as exc:
        raw_error = str(exc)
        error_message = raw_error
        if "10061" in raw_error or "Connection refused" in raw_error:
            error_message = (
                "Could not connect to the email server. "
                "Please configure SMTP values in .env "
                "(EMAIL_BACKEND, EMAIL_HOST, EMAIL_PORT, EMAIL_HOST_USER, "
                "EMAIL_HOST_PASSWORD, EMAIL_USE_TLS)."
            )
        MailLog.objects.create(
            student=student,
            fee=fee,
            email=email,
            mail_type=pending_mail.mail_type,
            status='failed',
            message=raw_error,
        )
        pending_mail.last_error = raw_error
        pending_mail.save(update_fields=['last_error'])
        return False, error_message


def _get_mail_log_student(mail_log):
    if mail_log.student:
        return mail_log.student

    if mail_log.fee and mail_log.fee.student:
        return mail_log.fee.student

    email = (mail_log.email or "").strip()
    if email:
        return Student.objects.filter(email__iexact=email).first()

    return None


@login_required
def mail_management(request):
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        messages.error(request, "Only admin can access Mail Management.")
        return redirect('dashboard')

    batch_name_subquery = StudentCourse.objects.filter(
        student_id=OuterRef('student_id'),
        batch__isnull=False
    ).order_by('-id').values('batch__batch_name')[:1]

    selected_department = request.GET.get('department', '')
    selected_course = request.GET.get('course', '')
    selected_batch = request.GET.get('batch', '')
    query = request.GET.get('q', '').strip()
    has_email = request.GET.get('has_email', '')
    batch_status = request.GET.get('batch_status', '')

    departments = Department.objects.order_by('department_name')
    courses = Course.objects.select_related('department').order_by('course_name')
    batches = Batch.objects.select_related('department', 'course').order_by('batch_name')

    mail_logs = MailLog.objects.select_related(
        'student', 'fee', 'fee__student', 'fee__course', 'fee__course__department'
    ).annotate(
        student_batch_name=Subquery(batch_name_subquery)
    ).all()
    
    # Apply filters
    mail_type = request.GET.get('mail_type')
    if mail_type:
        mail_logs = mail_logs.filter(mail_type=mail_type)
    
    status = request.GET.get('status')
    if status:
        mail_logs = mail_logs.filter(status=status)
    
    date_from = request.GET.get('date_from')
    if date_from:
        mail_logs = mail_logs.filter(sent_at__date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        mail_logs = mail_logs.filter(sent_at__date__lte=date_to)

    if selected_department:
        mail_logs = mail_logs.filter(
            Q(fee__course__department_id=selected_department) |
            Q(student__student_courses__department_id=selected_department)
        )

    if selected_course:
        mail_logs = mail_logs.filter(
            Q(fee__course_id=selected_course) |
            Q(student__student_courses__course_id=selected_course)
        )

    if selected_batch:
        mail_logs = mail_logs.filter(student__student_courses__batch_id=selected_batch)

    if query:
        mail_logs = mail_logs.filter(
            Q(student__student_name__icontains=query) |
            Q(fee__student__student_name__icontains=query) |
            Q(email__icontains=query)
        )

    if has_email == 'yes':
        mail_logs = mail_logs.exclude(email__isnull=True).exclude(email='')
    elif has_email == 'no':
        mail_logs = mail_logs.filter(Q(email__isnull=True) | Q(email=''))

    if batch_status == 'with_batch':
        mail_logs = mail_logs.filter(student_batch_name__isnull=False)
    elif batch_status == 'without_batch':
        mail_logs = mail_logs.filter(student_batch_name__isnull=True)

    mail_logs = mail_logs.distinct()
    
    # Calculate filtered counts
    filtered_total = mail_logs.count()
    filtered_sent = mail_logs.filter(status='sent').count()
    filtered_failed = mail_logs.filter(status='failed').count()

    per_page = _get_list_per_page(request, 20)
    paginator = Paginator(mail_logs.order_by('-sent_at'), per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    email_students = {}
    missing_student_emails = {
        (mail_log.email or "").strip().lower()
        for mail_log in page_obj
        if not mail_log.student and not (mail_log.fee and mail_log.fee.student) and (mail_log.email or "").strip()
    }
    if missing_student_emails:
        email_query = Q()
        for email in missing_student_emails:
            email_query |= Q(email__iexact=email)

        email_students = {
            student.email.lower(): student
            for student in Student.objects.filter(email_query)
            if student.email
        }

    for mail_log in page_obj:
        display_student = mail_log.student
        if not display_student and mail_log.fee:
            display_student = mail_log.fee.student
        if not display_student and mail_log.email:
            display_student = email_students.get(mail_log.email.strip().lower())

        mail_log.display_student = display_student
        mail_log.display_student_name = display_student.student_name if display_student else ''
        mail_log.can_preview = bool(display_student)
    
    email_backend = getattr(settings, 'EMAIL_BACKEND', '')

    return render(request, 'email/mail_management.html', {
        'mail_logs': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_mail_logs': filtered_total,
        'sent_mail_logs': filtered_sent,
        'failed_mail_logs': filtered_failed,
        'departments': departments,
        'courses': courses,
        'batches': batches,
        'selected_department': selected_department,
        'selected_course': selected_course,
        'selected_batch': selected_batch,
        'email_backend': email_backend,
        'is_console_email_backend': email_backend.endswith('console.EmailBackend'),
    })


@login_required
def mail_pending_list(request):
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        messages.error(request, "Only admin can access Mail Management.")
        return redirect('dashboard')

    selected_department = request.GET.get('department', '')
    selected_course = request.GET.get('course', '')
    selected_batch = request.GET.get('batch', '')
    query = request.GET.get('q', '').strip()
    has_email = request.GET.get('has_email', '')
    batch_status = request.GET.get('batch_status', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    departments = Department.objects.order_by('department_name')
    courses = Course.objects.select_related('department').order_by('course_name')
    batches = Batch.objects.select_related('department', 'course').order_by('batch_name')

    # Base queryset
    base_queryset = PendingMail.objects.select_related(
        'student', 'batch', 'batch__department', 'batch__course'
    ).filter(mail_type='batch')

    filtered_base_queryset = base_queryset

    if selected_department:
        filtered_base_queryset = filtered_base_queryset.filter(batch__department_id=selected_department)

    if selected_course:
        filtered_base_queryset = filtered_base_queryset.filter(batch__course_id=selected_course)

    if selected_batch:
        filtered_base_queryset = filtered_base_queryset.filter(batch_id=selected_batch)

    if query:
        filtered_base_queryset = filtered_base_queryset.filter(
            Q(student__student_name__icontains=query) |
            Q(email__icontains=query) |
            Q(batch__batch_name__icontains=query)
        )

    if date_from:
        filtered_base_queryset = filtered_base_queryset.filter(created_at__date__gte=date_from)

    if date_to:
        filtered_base_queryset = filtered_base_queryset.filter(created_at__date__lte=date_to)

    if batch_status == 'with_batch':
        filtered_base_queryset = filtered_base_queryset.filter(batch__isnull=False)
    elif batch_status == 'without_batch':
        filtered_base_queryset = filtered_base_queryset.filter(batch__isnull=True)

    filtered_base_queryset = filtered_base_queryset.distinct()
    batch_pending_list = filtered_base_queryset

    if has_email == 'yes':
        batch_pending_list = batch_pending_list.exclude(email__isnull=True).exclude(email='')
    elif has_email == 'no':
        batch_pending_list = batch_pending_list.filter(Q(email__isnull=True) | Q(email=''))
    
    # Calculate counts from filtered queryset
    total_pending = batch_pending_list.count()
    with_email_count = filtered_base_queryset.exclude(email__isnull=True).exclude(email='').count()
    without_email_count = filtered_base_queryset.filter(Q(email__isnull=True) | Q(email='')).count()

    per_page = _get_list_per_page(request, 20)
    paginator = Paginator(batch_pending_list.order_by('-created_at'), per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    email_backend = getattr(settings, 'EMAIL_BACKEND', '')

    return render(request, 'email/mail_pending_list.html', {
        'batch_pending_list': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_pending': total_pending,
        'with_email_count': with_email_count,
        'without_email_count': without_email_count,
        'departments': departments,
        'courses': courses,
        'batches': batches,
        'selected_department': selected_department,
        'selected_course': selected_course,
        'selected_batch': selected_batch,
        'email_backend': email_backend,
        'is_console_email_backend': email_backend.endswith('console.EmailBackend'),
    })


@login_required
def payment_email_pending_list(request):
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        messages.error(request, "Only admin can access Mail Management.")
        return redirect('dashboard')

    selected_department = request.GET.get('department', '')
    selected_course = request.GET.get('course', '')
    selected_batch = request.GET.get('batch', '')

    departments = Department.objects.order_by('department_name')
    courses = Course.objects.select_related('department').order_by('course_name')
    batches = Batch.objects.select_related('department', 'course').order_by('batch_name')

    # Base queryset
    base_queryset = PendingMail.objects.select_related(
        'student', 'fee', 'fee__course', 'fee__course__department'
    ).filter(mail_type='payment')

    filtered_base_queryset = base_queryset
    
    # Apply filters
    amount_range = request.GET.get('amount_range')
    if amount_range:
        if amount_range == '0-1000':
            filtered_base_queryset = filtered_base_queryset.filter(amount__gte=0, amount__lte=1000)
        elif amount_range == '1000-5000':
            filtered_base_queryset = filtered_base_queryset.filter(amount__gte=1000, amount__lte=5000)
        elif amount_range == '5000+':
            filtered_base_queryset = filtered_base_queryset.filter(amount__gte=5000)

    if selected_department:
        filtered_base_queryset = filtered_base_queryset.filter(
            Q(fee__course__department_id=selected_department) |
            Q(student__student_courses__department_id=selected_department)
        )

    if selected_course:
        filtered_base_queryset = filtered_base_queryset.filter(
            Q(fee__course_id=selected_course) |
            Q(student__student_courses__course_id=selected_course)
        )

    if selected_batch:
        filtered_base_queryset = filtered_base_queryset.filter(
            student__student_courses__batch_id=selected_batch
        )

    filtered_base_queryset = filtered_base_queryset.distinct()
    payment_pending_list = filtered_base_queryset
    
    has_email = request.GET.get('has_email')
    if has_email == 'yes':
        payment_pending_list = payment_pending_list.exclude(email__isnull=True).exclude(email='')
    elif has_email == 'no':
        payment_pending_list = payment_pending_list.filter(Q(email__isnull=True) | Q(email=''))
    
    # Calculate counts from filtered queryset
    total_pending = payment_pending_list.count()
    with_email_count = filtered_base_queryset.exclude(email__isnull=True).exclude(email='').count()
    without_email_count = filtered_base_queryset.filter(Q(email__isnull=True) | Q(email='')).count()

    per_page = _get_list_per_page(request, 20)
    paginator = Paginator(payment_pending_list.order_by('-created_at'), per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    email_backend = getattr(settings, 'EMAIL_BACKEND', '')

    return render(request, 'email/payment_email _Pending_List.html', {
        'payment_pending_list': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_pending': total_pending,
        'with_email_count': with_email_count,
        'without_email_count': without_email_count,
        'departments': departments,
        'courses': courses,
        'batches': batches,
        'selected_department': selected_department,
        'selected_course': selected_course,
        'selected_batch': selected_batch,
        'email_backend': email_backend,
        'is_console_email_backend': email_backend.endswith('console.EmailBackend'),
    })


@login_required
def overdue_email_pending_list(request):
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        messages.error(request, "Only admin can access Mail Management.")
        return redirect('dashboard')

    _sync_overdue_pending_mails()

    selected_department = request.GET.get('department', '')
    selected_course = request.GET.get('course', '')
    selected_batch = request.GET.get('batch', '')
    query = request.GET.get('q', '').strip()
    has_email = request.GET.get('has_email', '')
    amount_range = request.GET.get('amount_range', '')
    due_from = request.GET.get('due_from', '')
    due_to = request.GET.get('due_to', '')

    departments = Department.objects.order_by('department_name')
    courses = Course.objects.select_related('department').order_by('course_name')
    batches = Batch.objects.select_related('department', 'course').order_by('batch_name')

    base_queryset = PendingMail.objects.select_related(
        'student', 'batch', 'batch__department', 'batch__course', 'fee', 'fee__course', 'fee__course__department'
    ).filter(
        mail_type='overdue',
        fee__isnull=False,
    )

    filtered_base_queryset = base_queryset

    if selected_department:
        filtered_base_queryset = filtered_base_queryset.filter(fee__course__department_id=selected_department)

    if selected_course:
        filtered_base_queryset = filtered_base_queryset.filter(fee__course_id=selected_course)

    if selected_batch:
        filtered_base_queryset = filtered_base_queryset.filter(batch_id=selected_batch)

    if query:
        filtered_base_queryset = filtered_base_queryset.filter(
            Q(student__student_name__icontains=query) |
            Q(email__icontains=query) |
            Q(fee__course__course_name__icontains=query) |
            Q(batch__batch_name__icontains=query)
        )

    if amount_range:
        if amount_range == '0-1000':
            filtered_base_queryset = filtered_base_queryset.filter(amount__gte=0, amount__lte=1000)
        elif amount_range == '1000-5000':
            filtered_base_queryset = filtered_base_queryset.filter(amount__gte=1000, amount__lte=5000)
        elif amount_range == '5000+':
            filtered_base_queryset = filtered_base_queryset.filter(amount__gte=5000)

    if due_from:
        filtered_base_queryset = filtered_base_queryset.filter(fee__due_date__gte=due_from)

    if due_to:
        filtered_base_queryset = filtered_base_queryset.filter(fee__due_date__lte=due_to)

    filtered_base_queryset = filtered_base_queryset.distinct()
    overdue_pending_list = filtered_base_queryset

    if has_email == 'yes':
        overdue_pending_list = overdue_pending_list.exclude(email__isnull=True).exclude(email='')
    elif has_email == 'no':
        overdue_pending_list = overdue_pending_list.filter(Q(email__isnull=True) | Q(email=''))

    total_pending = overdue_pending_list.count()
    with_email_count = filtered_base_queryset.exclude(email__isnull=True).exclude(email='').count()
    without_email_count = filtered_base_queryset.filter(Q(email__isnull=True) | Q(email='')).count()

    today = timezone.localdate()
    per_page = _get_list_per_page(request, 20)
    paginator = Paginator(overdue_pending_list.order_by('fee__due_date', '-created_at'), per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    for item in page_obj:
        item.days_overdue = 0
        if item.fee and item.fee.due_date and item.fee.due_date < today:
            item.days_overdue = (today - item.fee.due_date).days

    email_backend = getattr(settings, 'EMAIL_BACKEND', '')

    return render(request, 'email/overdue_email_pending_list.html', {
        'overdue_pending_list': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_pending': total_pending,
        'with_email_count': with_email_count,
        'without_email_count': without_email_count,
        'departments': departments,
        'courses': courses,
        'batches': batches,
        'selected_department': selected_department,
        'selected_course': selected_course,
        'selected_batch': selected_batch,
        'email_backend': email_backend,
        'is_console_email_backend': email_backend.endswith('console.EmailBackend'),
    })


@login_required
@require_POST
def send_pending_mail_action(request, pending_id):
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"success": False, "message": "Only admin can send mail."}, status=403)
        messages.error(request, "Only admin can send mail.")
        return redirect('mail_pending_list')

    pending_mail = get_object_or_404(
        PendingMail.objects.select_related('student', 'batch', 'fee', 'fee__course'),
        id=pending_id
    )
    success, message = _send_pending_mail_now(pending_mail)
    redirect_name = _get_pending_mail_redirect_name(pending_mail.mail_type)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({"success": success, "message": message})

    if success:
        messages.success(request, message)
    else:
        messages.error(request, f"Mail send failed: {message}")
    return redirect(redirect_name)


def _handle_bulk_pending_mail_action(request, mail_type):
    redirect_name = _get_pending_mail_redirect_name(mail_type)

    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"success": False, "message": "Only admin can send mail."}, status=403)
        messages.error(request, "Only admin can send mail.")
        return redirect(redirect_name)

    pending_ids = request.POST.getlist('pending_ids')
    if not pending_ids:
        messages.warning(request, "Select at least one pending email to send.")
        return redirect(redirect_name)

    pending_queryset = PendingMail.objects.select_related('student', 'batch', 'fee', 'fee__course').filter(
        id__in=pending_ids,
        mail_type=mail_type
    )
    pending_mails = list(pending_queryset)

    if not pending_mails:
        messages.warning(request, "Selected pending emails were not found.")
        return redirect(redirect_name)

    sent_count = 0
    failed_count = 0

    for pending_mail in pending_mails:
        success, _ = _send_pending_mail_now(pending_mail)
        if success:
            sent_count += 1
        else:
            failed_count += 1

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            "success": failed_count == 0,
            "message": f"Bulk send completed. Sent: {sent_count}, Failed: {failed_count}.",
            "sent_count": sent_count,
            "failed_count": failed_count,
        })

    if sent_count and not failed_count:
        messages.success(request, f"Bulk send completed successfully. Sent: {sent_count}.")
    elif sent_count and failed_count:
        messages.warning(request, f"Bulk send completed with partial failures. Sent: {sent_count}, Failed: {failed_count}.")
    else:
        messages.error(request, f"Bulk send failed for all selected emails ({failed_count}).")
    return redirect(redirect_name)


@login_required
@require_POST
def send_bulk_pending_mail_action(request):
    return _handle_bulk_pending_mail_action(request, 'batch')


@login_required
@require_POST
def send_bulk_payment_pending_mail_action(request):
    return _handle_bulk_pending_mail_action(request, 'payment')


@login_required
@require_POST
def send_bulk_overdue_pending_mail_action(request):
    return _handle_bulk_pending_mail_action(request, 'overdue')


def _preview_pending_mail_action(request, pending_id, mail_type, builder):
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        return JsonResponse({"success": False, "message": "Only admin can preview mail."}, status=403)

    pending_mail = get_object_or_404(
        PendingMail.objects.select_related('student', 'batch', 'fee', 'fee__course'),
        id=pending_id,
        mail_type=mail_type
    )

    try:
        subject, text_body, html_body = builder(pending_mail)
    except Exception as exc:
        return JsonResponse({"success": False, "message": str(exc)}, status=500)

    return JsonResponse({
        "success": True,
        "subject": subject,
        "text_body": text_body,
        "html_body": html_body,
    })


@login_required
def preview_batch_pending_mail_action(request, pending_id):
    return _preview_pending_mail_action(request, pending_id, 'batch', _build_batch_email_content)


@login_required
def preview_payment_pending_mail_action(request, pending_id):
    return _preview_pending_mail_action(request, pending_id, 'payment', _build_payment_email_content)


@login_required
def preview_overdue_pending_mail_action(request, pending_id):
    return _preview_pending_mail_action(request, pending_id, 'overdue', _build_overdue_email_content)


@login_required
def preview_mail_log_action(request, log_id):
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != "admin":
        return JsonResponse({"success": False, "message": "Only admin can preview mail."}, status=403)

    mail_log = get_object_or_404(
        MailLog.objects.select_related('student', 'fee', 'fee__student', 'fee__course'),
        id=log_id
    )
    student = _get_mail_log_student(mail_log)
    if not student:
        return JsonResponse({"success": False, "message": "Student record not available for this log."}, status=400)

    try:
        if mail_log.mail_type == 'batch':
            latest_student_course = StudentCourse.objects.select_related('batch').filter(
                student=student,
                batch__isnull=False
            ).order_by('-id').first()
            pending_mail = SimpleNamespace(
                student=student,
                batch=latest_student_course.batch if latest_student_course else None,
                amount=None,
                payment_date=mail_log.sent_at.date() if mail_log.sent_at else None,
                mail_type='batch',
            )
            subject, text_body, html_body = _build_batch_email_content(pending_mail)
        elif mail_log.mail_type == 'payment':
            payment = Payment.objects.select_related('fee').filter(
                fee__student=student
            ).order_by('-payment_date', '-id').first()
            pending_mail = SimpleNamespace(
                student=student,
                batch=None,
                amount=payment.amount_paid if payment else None,
                payment_date=payment.payment_date if payment else (mail_log.sent_at.date() if mail_log.sent_at else None),
                mail_type='payment',
                fee=mail_log.fee or (payment.fee if payment else None),
            )
            subject, text_body, html_body = _build_payment_email_content(pending_mail)
        elif mail_log.mail_type == 'overdue':
            pending_mail = SimpleNamespace(
                student=student,
                batch=None,
                amount=mail_log.fee.pending_amount if mail_log.fee else None,
                payment_date=None,
                mail_type='overdue',
                fee=mail_log.fee,
            )
            subject, text_body, html_body = _build_overdue_email_content(pending_mail)
        else:
            return JsonResponse({"success": False, "message": "Unsupported mail type."}, status=400)
    except Exception as exc:
        return JsonResponse({"success": False, "message": str(exc)}, status=500)

    return JsonResponse({
        "success": True,
        "subject": subject,
        "text_body": text_body,
        "html_body": html_body,
        "mail_type": mail_log.mail_type,
    })


@login_required
def assign_students_batch(request):
    if request.method == "POST":
        batch_id = request.POST.get("batch_id")
        student_ids = request.POST.getlist("student_ids")

        if not batch_id or not student_ids:
            return JsonResponse({"success": False, "error": "Invalid data"})

        batch = Batch.objects.get(id=batch_id)
        students, student_courses, error = _get_batch_assignment_candidates(batch, student_ids)
        if error:
            return JsonResponse({"success": False, "error": error})

        current_count = StudentCourse.objects.filter(batch=batch, course=batch.course).count()
        available_seats = batch.no_of_students - current_count
        if len(students) > available_seats:
            return JsonResponse({
                "success": False,
                "error": f"Only {available_seats} seats left"
            })

        _assign_students_to_batch(batch, students, student_courses)

        return JsonResponse({"success": True})
    
@login_required
def get_batch_students(request, batch_id):
    batch = Batch.objects.get(id=batch_id)

    capacity = batch.no_of_students
    
    # Students assigned to this batch for this course
    assigned = Student.objects.filter(
        student_courses__batch=batch,
        student_courses__course=batch.course
    ).distinct()
    
    current = assigned.count()
    seats_left = capacity - current

    # Students who have this course but no batch assigned for it
    available = Student.objects.filter(
        student_courses__course=batch.course,
        student_courses__batch__isnull=True
    ).exclude(
        status='dropped'
    ).distinct()

    eligible_students = []
    blocked_students = []
    for student in available:
        block_reason = _get_batch_assignment_block_reason(student, batch.course)
        student_payload = {
            "id": student.id,
            "name": student.student_name,
        }
        if block_reason:
            student_payload["reason"] = block_reason
            blocked_students.append(student_payload)
        else:
            eligible_students.append(student_payload)

    return JsonResponse({
        "assigned":[{"id":s.id,"name":s.student_name} for s in assigned],
        "available": eligible_students,
        "blocked": blocked_students,
        "capacity": capacity,
        "current": current,
        "seats_left": seats_left
    })

def assign_students_to_batch(request):
    if request.method == "POST":
        student_ids = request.POST.getlist("student_ids")
        batch_id = request.POST.get("batch_id")

        batch = Batch.objects.get(id=batch_id)

        students, student_courses, error = _get_batch_assignment_candidates(batch, student_ids)
        if error:
            return JsonResponse({"success": False, "error": error})

        current_count = StudentCourse.objects.filter(batch=batch, course=batch.course).count()
        available_seats = batch.no_of_students - current_count

        if len(students) > available_seats:
            return JsonResponse({
                "success": False,
                "error": f"Only {available_seats} seats left"
            })

        _assign_students_to_batch(batch, students, student_courses)

        return JsonResponse({"success": True})


@require_POST
def assign_single_student_batch(request):
    user_profile = request.user.userprofile
    
    # Prevent trainers from assigning students
    if user_profile.role == "trainer":
        return JsonResponse({"success": False, "error": "Trainers cannot assign students to batches"})
    
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        batch_id = request.POST.get("batch_id")

        try:
            batch = Batch.objects.get(id=batch_id)
            current_count = StudentCourse.objects.filter(batch=batch, course=batch.course).count()

            if current_count >= batch.no_of_students:
                return JsonResponse({
                    "success": False,
                    "error": "Batch full"
                })

            students, student_courses, error = _get_batch_assignment_candidates(batch, [student_id])
            if error:
                return JsonResponse({"success": False, "error": error})

            _assign_students_to_batch(batch, students, student_courses)

            return JsonResponse({"success": True})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

@require_POST
def remove_student_batch(request):
    user_profile = request.user.userprofile
    
    # Prevent trainers from removing students
    if user_profile.role == "trainer":
        return JsonResponse({"success": False, "error": "Trainers cannot remove students from batches"})
    
    student_id = request.POST.get("student_id")
    batch_id = request.POST.get("batch_id")
    
    student_course = StudentCourse.objects.filter(student_id=student_id, batch_id=batch_id).first()
    if student_course:
        student_course.batch = None
        student_course.save(update_fields=['batch'])
        sync_student_fee(student_course.student)
    return JsonResponse({"success":True})

@require_POST
def assign_multiple_students_batch(request):
    user_profile = request.user.userprofile
    
    # Prevent trainers from assigning students
    if user_profile.role == "trainer":
        return JsonResponse({"success": False, "error": "Trainers cannot assign students to batches"})
    
    ids = request.POST.get('student_ids', '').split(',')
    batch_id = request.POST.get('batch_id')
    batch = Batch.objects.get(id=batch_id)

    students, student_courses, error = _get_batch_assignment_candidates(batch, ids)
    if error:
        return JsonResponse({'success': False, 'error': error})

    current_count = StudentCourse.objects.filter(batch=batch, course=batch.course).count()
    available_seats = batch.no_of_students - current_count
    if len(students) > available_seats:
        return JsonResponse({
            'success': False,
            'error': f"Only {available_seats} seats left"
        })

    _assign_students_to_batch(batch, students, student_courses)

    return JsonResponse({'success':True})

@require_POST
def remove_multiple_students_batch(request):
    user_profile = request.user.userprofile
    
    # Prevent trainers from removing students
    if user_profile.role == "trainer":
        return JsonResponse({"success": False, "error": "Trainers cannot remove students from batches"})
    
    ids = request.POST.get('student_ids').split(',')
    batch_id = request.POST.get('batch_id')

    student_courses = list(
        StudentCourse.objects.filter(student_id__in=ids, batch_id=batch_id).select_related('student')
    )
    for student_course in student_courses:
        student_course.batch = None
        student_course.save(update_fields=['batch'])
        sync_student_fee(student_course.student)

    return JsonResponse({'success':True})

@login_required
def batch_students(request, batch_id):
    batch = get_object_or_404(Batch, id=batch_id)
    students = Student.objects.filter(
        student_courses__batch=batch,
        student_courses__course=batch.course
    ).distinct()
    return render(request, 'batch_students.html', {
    'batch': batch,
    'students': students
})
#============MODULE==================

@login_required
def syllabus_view(request):
    courses = Course.objects.all()
    modules = Module.objects.select_related('course').prefetch_related('topic_set').order_by('course', 'id')
    
    filter_course = request.GET.get('filter_course')
    if filter_course:
        modules = modules.filter(course_id=filter_course)
    
    return render(request, 'syllabus.html', {'courses': courses, 'modules': modules})

@login_required
def add_module(request):
    if request.method == "POST":
        course_id = request.POST.get('course')
        module_name = request.POST.get('module_name')
        if not course_id or not module_name:
            messages.error(request, "Course and module name are required.")
            return redirect('syllabus')
        
        Module.objects.create(course_id=course_id, module_name=module_name.strip())
        messages.success(request, "Module added successfully.")
        return redirect('syllabus')
    return redirect('syllabus')

@login_required
def add_topics(request):
    if request.method == "POST":
        module_id = request.POST.get('module')
        topic_names = request.POST.getlist('topic_names[]')
        
        if not module_id or not topic_names:
            messages.error(request, "Module and at least one topic name are required.")
            return redirect('syllabus')
        
        created_count = 0
        for topic_name in topic_names:
            if topic_name.strip():
                Topic.objects.create(module_id=module_id, topic_name=topic_name.strip())
                created_count += 1
        
        messages.success(request, f"{created_count} topic(s) added successfully.")
        return redirect('syllabus')
    return redirect('syllabus')

@login_required
def trainer_syllabus(request):
    user_profile = request.user.userprofile
    
    if user_profile.role == 'trainer':
        batches = Batch.objects.filter(trainer=user_profile, status='Active').select_related('course')
    elif user_profile.role == 'admin':
        batches = Batch.objects.filter(status='Active').select_related('course', 'trainer__user')
    else:
        messages.error(request, "Access denied.")
        return redirect('dashboard')
    
    batch_id = request.GET.get('batch')
    selected_batch = None
    modules = []
    total_topics = 0
    completed_topics = 0
    partial_topics = 0
    not_started_topics = 0
    
    if batch_id:
        if user_profile.role == 'trainer':
            selected_batch = get_object_or_404(Batch, id=batch_id, trainer=user_profile)
        else:
            selected_batch = get_object_or_404(Batch, id=batch_id)
        
        modules = Module.objects.filter(course=selected_batch.course).prefetch_related('topic_set').order_by('id')
        
        for module in modules:
            module_total_topics = 0
            module_completed_topics = 0
            module_partial_topics = 0

            for topic in module.topic_set.all():
                topic.progress = TopicProgress.objects.filter(batch=selected_batch, topic=topic).first()
                module_total_topics += 1
                total_topics += 1
                if topic.progress:
                    if topic.progress.status == 'completed':
                        module_completed_topics += 1
                        completed_topics += 1
                    elif topic.progress.status == 'partial':
                        module_partial_topics += 1
                        partial_topics += 1
                    else:
                        not_started_topics += 1
                else:
                    not_started_topics += 1

            module.not_started_topics = module_total_topics - (module_completed_topics + module_partial_topics)
            module.total_topics = module_total_topics
            module.completed_topics = module_completed_topics
            module.partial_topics = module_partial_topics
            module.progress_percent = round((module_completed_topics / module_total_topics) * 100) if module_total_topics else 0
    
    return render(request, 'trainer_syllabus.html', {
        'batches': batches,
        'selected_batch': selected_batch,
        'modules': modules,
        'total_topics': total_topics,
        'completed_topics': completed_topics,
        'partial_topics': partial_topics,
        'not_started_topics': not_started_topics,
    })

@login_required
def module_view(request):
    courses = Course.objects.all()
    if request.method == "POST":
        course_id = request.POST.get('course')
        module_name = request.POST.get('module_name')
        if not course_id or not module_name:
            messages.error(request, "Course and module name are required.")
            return redirect('module')
        
        Module.objects.create(course_id=course_id, module_name=module_name.strip())
        messages.success(request, "Module added successfully.")
        return redirect('module')

    modules = Module.objects.select_related('course')
    return render(request, 'module.html', {'courses': courses, 'modules': modules})

@login_required
def edit_module(request):
    if request.method == "POST":
        module_id = request.POST.get('module_id')
        course_id = request.POST.get('course')
        module_name = request.POST.get('module_name')
        
        if not module_id or not course_id or not module_name:
            messages.error(request, "All fields are required.")
            return redirect('syllabus')
            
        try:
            module = Module.objects.get(id=module_id)
            module.course_id = course_id
            module.module_name = module_name.strip()
            module.save()
            messages.success(request, "Module updated successfully.")
        except Module.DoesNotExist:
            messages.error(request, "Module not found.")
            
        return redirect('syllabus')
    return redirect('syllabus')

@login_required
def delete_module(request, pk):
    module = get_object_or_404(Module, pk=pk)
    try:
        module.delete()
        messages.success(request, "Module deleted successfully.")
    except ProtectedError:
        messages.error(request, "Cannot delete this module because it is in use.")
    return redirect('syllabus')

#============TOPIC==================

@login_required
def topic_view(request):
    courses = Course.objects.all()
    if request.method == "POST":
        module_id = request.POST.get('module')
        topic_names = request.POST.getlist('topic_names[]')
        
        if not module_id or not topic_names:
            messages.error(request, "Module and at least one topic name are required.")
            return redirect('topic')
        
        created_count = 0
        for topic_name in topic_names:
            if topic_name.strip():
                Topic.objects.create(module_id=module_id, topic_name=topic_name.strip())
                created_count += 1
        
        messages.success(request, f"{created_count} topic(s) added successfully.")
        return redirect('topic')

    # Filter topics by course
    topics = Topic.objects.select_related('module__course').order_by('module__course', 'module', 'id')
    
    filter_course = request.GET.get('filter_course')
    if filter_course:
        topics = topics.filter(module__course_id=filter_course)
    
    return render(request, 'topic.html', {'courses': courses, 'topics': topics})

@login_required
def edit_topic(request):
    if request.method == "POST":
        topic_id = request.POST.get('topic_id')
        module_id = request.POST.get('module')
        topic_name = request.POST.get('topic_name')
        
        if not topic_id or not module_id or not topic_name:
            messages.error(request, "All fields are required.")
            return redirect('syllabus')
            
        try:
            topic = Topic.objects.get(id=topic_id)
            topic.module_id = module_id
            topic_name = topic_name.strip()
            topic.topic_name = topic_name
            topic.save()
            messages.success(request, "Topic updated successfully.")
        except Topic.DoesNotExist:
            messages.error(request, "Topic not found.")
            
        return redirect('syllabus')
    return redirect('syllabus')

@login_required
def delete_topic(request, pk):
    topic = get_object_or_404(Topic, pk=pk)
    try:
        topic.delete()
        messages.success(request, "Topic deleted successfully.")
    except ProtectedError:
        messages.error(request, "Cannot delete this topic because it is in use.")
    return redirect('syllabus')
#=======================================USERS========================================================================

# CREATE USER (ADMIN)
def create_user(request):
    from django.contrib.auth.password_validation import validate_password
    from django.contrib.auth.validators import UnicodeUsernameValidator
    from django.core.exceptions import ValidationError
    from django.core.validators import validate_email
    from django.db import transaction
    import re

    departments = Department.objects.prefetch_related('course_set').order_by('department_name')
    field_errors = {}
    form_data = {
        "username": "",
        "first_name": "",
        "last_name": "",
        "email": "",
        "mobile": "",
        "role": "",
        "specialized_departments": [],
        "specialized_courses": [],
    }

    if request.method == "POST":
        username = (request.POST.get('username') or "").strip()
        password = request.POST.get('password') or ""
        first_name = (request.POST.get('first_name') or "").strip()
        last_name = (request.POST.get('last_name') or "").strip()
        email = (request.POST.get('email') or "").strip()
        mobile = (request.POST.get('mobile') or "").strip()
        role = (request.POST.get('role') or "").strip()
        specialized_departments = [value for value in request.POST.getlist('specialized_departments') if value]
        specialized_courses = [value for value in request.POST.getlist('specialized_courses') if value]

        form_data.update(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            mobile=mobile,
            role=role,
            specialized_departments=specialized_departments,
            specialized_courses=specialized_courses,
        )

        username_validator = UnicodeUsernameValidator()
        name_pattern = re.compile(r"^[A-Za-z][A-Za-z .'-]*$")
        mobile_pattern = re.compile(r"^\+?\d{10,15}$")
        valid_roles = {choice[0] for choice in UserProfile.ROLE_CHOICES}

        if not username:
            field_errors["username"] = "Username is required."
        elif len(username) < 3:
            field_errors["username"] = "Username must be at least 3 characters."
        else:
            try:
                username_validator(username)
            except ValidationError:
                field_errors["username"] = "Use only letters, digits, and @/./+/-/_ characters."
            if User.objects.filter(username__iexact=username).exists():
                field_errors["username"] = "Username already exists."

        if not password:
            field_errors["password"] = "Password is required."
        else:
            temp_user = User(username=username, email=email, first_name=first_name, last_name=last_name)
            try:
                validate_password(password, temp_user)
            except ValidationError as exc:
                field_errors["password"] = " ".join(exc.messages)

        if not first_name:
            field_errors["first_name"] = "First name is required."
        elif len(first_name) > 150:
            field_errors["first_name"] = "First name must be 150 characters or fewer."
        elif not name_pattern.match(first_name):
            field_errors["first_name"] = "First name contains invalid characters."

        if not last_name:
            field_errors["last_name"] = "Last name is required."
        elif len(last_name) > 150:
            field_errors["last_name"] = "Last name must be 150 characters or fewer."
        elif not name_pattern.match(last_name):
            field_errors["last_name"] = "Last name contains invalid characters."

        if not email:
            field_errors["email"] = "Email is required."
        else:
            try:
                validate_email(email)
            except ValidationError:
                field_errors["email"] = "Enter a valid email address."
            if User.objects.filter(email__iexact=email).exists():
                field_errors["email"] = "Email already exists."

        if not mobile:
            field_errors["mobile"] = "Mobile is required."
        elif not mobile_pattern.match(mobile):
            field_errors["mobile"] = "Enter a valid mobile number (10 to 15 digits, optional +)."
        elif UserProfile.objects.filter(mobile=mobile).exists():
            field_errors["mobile"] = "Mobile number already exists."

        if not role:
            field_errors["role"] = "Role is required."
        elif role not in valid_roles:
            field_errors["role"] = "Select a valid role."

        specialization_pairs = []
        if role == "trainer":
            department_ids = {department.id for department in departments}
            selected_department_ids = set()
            for department_id in specialized_departments:
                try:
                    parsed_department_id = int(department_id)
                except (TypeError, ValueError):
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                if parsed_department_id not in department_ids:
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                selected_department_ids.add(parsed_department_id)

            selected_courses = Course.objects.filter(id__in=specialized_courses).select_related('department')
            selected_course_map = {course.id: course for course in selected_courses}

            if not selected_department_ids:
                field_errors["specialization"] = "Select at least one specialized department."
            if not specialized_courses:
                field_errors["specialization"] = "Select at least one course for the trainer."

            department_course_selection = {department_id: [] for department_id in selected_department_ids}
            for course_id in specialized_courses:
                try:
                    parsed_course_id = int(course_id)
                except (TypeError, ValueError):
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                course = selected_course_map.get(parsed_course_id)
                if not course or course.department_id not in selected_department_ids:
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                department_course_selection[course.department_id].append(course.id)

            if not field_errors.get("specialization"):
                missing_course_departments = [
                    department_id for department_id, course_ids in department_course_selection.items() if not course_ids
                ]
                if missing_course_departments:
                    field_errors["specialization"] = "Select at least one course under each specialized department."
                else:
                    specialization_pairs = [
                        (department_id, course_id)
                        for department_id, course_ids in department_course_selection.items()
                        for course_id in course_ids
                    ]

        if field_errors:
            messages.error(request, "Please correct the highlighted fields.")
            return render(
                request,
                "user.html",
                {"field_errors": field_errors, "form_data": form_data, "departments": departments},
                status=400,
            )

        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                password=password,
                email=email.lower(),
                first_name=first_name,
                last_name=last_name
            )
            user_profile = UserProfile.objects.create(
                user=user,
                mobile=mobile,
                role=role
            )
            if role == "trainer":
                TrainerSpecialization.objects.bulk_create(
                    [
                        TrainerSpecialization(
                            trainer=user_profile,
                            department_id=department_id,
                            course_id=course_id,
                        )
                        for department_id, course_id in specialization_pairs
                    ]
                )

        display_name = first_name or username
        messages.success(request, f"Welcome {display_name}! User created successfully.")
        return redirect('user_list')

    return render(
        request,
        "user.html",
        {"field_errors": field_errors, "form_data": form_data, "departments": departments},
    )



def user_list(request):
    current_profile = UserProfile.objects.filter(user=request.user).first()
    if not current_profile or current_profile.role != 'admin':
        messages.error(request, "Only admin can access the user list.")
        return redirect('dashboard')

    users = User.objects.select_related('userprofile').filter(is_superuser=False)

    # Search
    q = request.GET.get('q')
    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(email__icontains=q)
        )

    # Role filter
    role = request.GET.get('role')
    if role:
        users = users.filter(userprofile__role=role)

    # Date filters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date:
        users = users.filter(date_joined__date__gte=from_date)

    if to_date:
        users = users.filter(date_joined__date__lte=to_date)

    return render(request, 'user_view.html', {
        'users': users
    })

def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile = get_object_or_404(UserProfile, user=user)
    current_profile = get_object_or_404(UserProfile, user=request.user)
    is_admin_user = current_profile.role == 'admin'
    is_self_profile = request.user.id == user.id
    departments = Department.objects.prefetch_related('course_set').order_by('department_name')
    existing_specializations = list(
        TrainerSpecialization.objects.filter(trainer=profile).select_related('department', 'course')
    )
    initial_specialized_departments = sorted(
        {specialization.department_id for specialization in existing_specializations}
    )
    initial_specialized_courses = [specialization.course_id for specialization in existing_specializations]
    field_errors = {}
    form_data = {
        "username": user.username,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "email": user.email,
        "mobile": profile.mobile,
        "specialized_departments": initial_specialized_departments,
        "specialized_courses": initial_specialized_courses,
    }

    if not is_admin_user and not is_self_profile:
        messages.error(request, "You can only edit your own profile.")
        return redirect('user_detail', user_id=request.user.id)

    if request.method == "POST":
        username = (request.POST.get('username') or "").strip()
        user.first_name = (request.POST.get('first_name') or "").strip()
        user.last_name = (request.POST.get('last_name') or "").strip()
        user.email = (request.POST.get('email') or "").strip()
        profile.mobile = (request.POST.get('mobile') or "").strip()
        if is_admin_user and profile.role != 'admin':
            profile.role = (request.POST.get('role') or "").strip()

        specialized_departments = [value for value in request.POST.getlist('specialized_departments') if value]
        specialized_courses = [value for value in request.POST.getlist('specialized_courses') if value]
        form_data.update(
            username=username,
            first_name=user.first_name,
            last_name=user.last_name,
            email=user.email,
            mobile=profile.mobile,
            specialized_departments=specialized_departments,
            specialized_courses=specialized_courses,
        )

        username_validator = UnicodeUsernameValidator()
        if not username:
            field_errors["username"] = "Username is required."
        elif len(username) < 3:
            field_errors["username"] = "Username must be at least 3 characters."
        else:
            try:
                username_validator(username)
            except ValidationError:
                field_errors["username"] = "Use only letters, digits, and @/./+/-/_ characters."
            if User.objects.filter(username__iexact=username).exclude(id=user.id).exists():
                field_errors["username"] = "Username already exists."

        user.username = username

        specialization_pairs = []
        if is_admin_user and profile.role == "trainer":
            department_ids = {department.id for department in departments}
            selected_department_ids = set()

            for department_id in specialized_departments:
                try:
                    parsed_department_id = int(department_id)
                except (TypeError, ValueError):
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                if parsed_department_id not in department_ids:
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                selected_department_ids.add(parsed_department_id)

            selected_courses = Course.objects.filter(id__in=specialized_courses).select_related('department')
            selected_course_map = {course.id: course for course in selected_courses}

            if not selected_department_ids:
                field_errors["specialization"] = "Select at least one specialized department."
            if not specialized_courses:
                field_errors["specialization"] = "Select at least one course for the trainer."

            department_course_selection = {department_id: [] for department_id in selected_department_ids}
            for course_id in specialized_courses:
                try:
                    parsed_course_id = int(course_id)
                except (TypeError, ValueError):
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                course = selected_course_map.get(parsed_course_id)
                if not course or course.department_id not in selected_department_ids:
                    field_errors["specialization"] = "Select valid departments and courses for the trainer."
                    continue
                department_course_selection[course.department_id].append(course.id)

            if not field_errors.get("specialization"):
                missing_course_departments = [
                    department_id for department_id, course_ids in department_course_selection.items() if not course_ids
                ]
                if missing_course_departments:
                    field_errors["specialization"] = "Select at least one course under each specialized department."
                else:
                    specialization_pairs = [
                        (department_id, course_id)
                        for department_id, course_ids in department_course_selection.items()
                        for course_id in course_ids
                    ]

        if field_errors:
            messages.error(request, "Please correct the highlighted fields.")
            return render(request, 'edit_user.html', {
                'user_obj': user,
                'profile': profile,
                'is_self_profile': is_self_profile,
                'can_manage_users': is_admin_user,
                'departments': departments,
                'field_errors': field_errors,
                'form_data': form_data,
            }, status=400)

        password = request.POST.get('password')
        if password:
            user.set_password(password)

        with transaction.atomic():
            user.save()
            profile.save()
            if is_admin_user:
                TrainerSpecialization.objects.filter(trainer=profile).delete()
                if profile.role == "trainer" and specialization_pairs:
                    TrainerSpecialization.objects.bulk_create(
                        [
                            TrainerSpecialization(
                                trainer=profile,
                                department_id=department_id,
                                course_id=course_id,
                            )
                            for department_id, course_id in specialization_pairs
                        ]
                    )

        messages.success(request, "User updated successfully")
        return redirect('user_detail', user_id=user.id)

    return render(request, 'edit_user.html', {
        'user_obj': user,
        'profile': profile,
        'is_self_profile': is_self_profile,
        'can_manage_users': is_admin_user,
        'departments': departments,
        'field_errors': field_errors,
        'form_data': form_data,
    })



def delete_user(request, user_id):
    from django.db import IntegrityError
    
    user_to_delete = get_object_or_404(User, id=user_id)

    #  Prevent deleting yourself
    if user_to_delete == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('user_list')

    #  Prevent deleting super admin
    if user_to_delete.is_superuser:
        messages.error(request, "Super admin cannot be deleted.")
        return redirect('user_list')

    try:
        user_to_delete.delete()
        messages.success(request, "User deleted successfully")
    except (ProtectedError, IntegrityError):
        messages.error(request, "Cannot delete this user because they have associated records (leads, students, notifications, etc.).")

    return redirect('user_list')

def user_live_search(request):
    query = request.GET.get('q', '')

    users = User.objects.select_related('userprofile').filter(is_superuser=False)

    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query)
        )

    data = []
    for user in users:
        data.append({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": user.userprofile.role if hasattr(user, 'userprofile') else "",
            "date_joined": user.date_joined.strftime("%Y-%m-%d"),
        })

    return JsonResponse({"users": data})


@login_required
def user_detail(request, user_id):
    user = get_object_or_404(
        User.objects.select_related('userprofile'),
        id=user_id,
    )
    profile = get_object_or_404(UserProfile, user=user)
    trainer_specializations = TrainerSpecialization.objects.filter(
        trainer=profile
    ).select_related('department', 'course').order_by('department__department_name', 'course__course_name')

    specialization_map = {}
    for specialization in trainer_specializations:
        specialization_map.setdefault(
            specialization.department.department_name,
            []
        ).append(specialization.course.course_name)

    return render(request, 'user_detail.html', {
        'user_obj': user,
        'profile': profile,
        'specialization_map': specialization_map,
    })


#=======================================TASK=========================================================================
def add_task(request, lead_id):
    lead = get_object_or_404(StudentEnquiry, id=lead_id)

    if request.method == "POST":
        LeadTask.objects.create(
            lead=lead,
            task_name=request.POST.get("task_name"),
            category=request.POST.get("category"),
            date=request.POST.get("date"),
            time=request.POST.get("time"),
            assigned_to_id=request.POST.get("assigned_to"),
            description=request.POST.get("description"),
        )

        # create activity entry
        LeadActivity.objects.create(
            lead=lead,
            action="Task Added",
            new_value=request.POST.get("task_name")
        )

        messages.success(request, "Task added successfully")

    return redirect("lead_profile", id=lead.id)


#=======================================FOLLOWUPS=========================================================================

@login_required
def followups(request):
    # Get all follow-ups from FollowUp table
    followups_qs = FollowUp.objects.select_related(
    'lead__status', 'assigned_to__user'
    ).prefetch_related(
        'lead__lead_courses__course',
        'lead__lead_courses__department'
    ).order_by('-followup_date', '-followup_time')

    # Filter by assigned CRE
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role == 'cre':
        followups_qs = followups_qs.filter(assigned_to=user_profile)

    # Search
    q = request.GET.get('q')
    if q:
        followups_qs = followups_qs.filter(
            Q(lead__full_name__icontains=q) |
            Q(lead__email__icontains=q) |
            Q(lead__mobile__icontains=q) |
            Q(title__icontains=q)
        )

    # Filter by date
    date_from = request.GET.get('date_from')
    if date_from:
        followups_qs = followups_qs.filter(followup_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        followups_qs = followups_qs.filter(followup_date__lte=date_to)

    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        followups_qs = followups_qs.filter(status=status_filter)

    # Filter by assigned user
    assigned_to = request.GET.get('assigned_to')
    if assigned_to:
        followups_qs = followups_qs.filter(assigned_to_id=assigned_to)

    # Pagination
    per_page = int(request.GET.get('per_page', 10))
    paginator = Paginator(followups_qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'followups': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'users': UserProfile.objects.select_related('user').filter(role='cre'),
    }
    return render(request, 'followups.html', context)


@login_required
def followup_live_search(request):
    query = (request.GET.get('q') or '').strip()

    followups_qs = FollowUp.objects.select_related(
        'lead', 'lead__status', 'assigned_to__user'
    ).prefetch_related(
        'lead__lead_courses__course',
        'lead__lead_courses__department'
    )

    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role == 'cre':
        followups_qs = followups_qs.filter(assigned_to=user_profile)

    if query:
        followups_qs = followups_qs.filter(
            Q(lead__full_name__icontains=query) |
            Q(lead__email__icontains=query) |
            Q(lead__mobile__icontains=query) |
            Q(title__icontains=query) |
            Q(notes__icontains=query)
        )

    followups_qs = followups_qs.order_by('-followup_date', '-followup_time')[:50]

    data = []
    for followup in followups_qs:
        lead_courses = [
            {
                "course": lead_course.course.course_name if lead_course.course else "",
                "department": lead_course.department.department_name if lead_course.department else "",
            }
            for lead_course in followup.lead.lead_courses.all()
        ]

        data.append({
            "id": followup.id,
            "lead_id": followup.lead.id,
            "lead_name": followup.lead.full_name,
            "lead_mobile": followup.lead.mobile or "",
            "lead_email": followup.lead.email or "",
            "courses": lead_courses,
            "status": followup.status,
            "status_display": followup.get_status_display(),
            "followup_date": followup.followup_date.strftime("%d %b %Y"),
            "followup_time": followup.followup_time.strftime("%H:%M") if followup.followup_time else "",
            "title": followup.title,
            "notes": followup.notes or "",
            "assigned_to": (
                followup.assigned_to.user.get_full_name() or followup.assigned_to.user.username
                if followup.assigned_to else ""
            ),
        })

    return JsonResponse({"followups": data})


@login_required
def complete_followup(request, followup_id):
    followup = get_object_or_404(FollowUp, id=followup_id)
    
    # Check permission: admin or assigned person
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role != 'admin' and followup.assigned_to != user_profile:
        messages.error(request, "You don't have permission to complete this follow-up")
        return redirect('followups')
    
    if request.method == "POST":
        notes = request.POST.get('notes', '')
        add_new_followup = request.POST.get('add_new_followup') == 'yes'
        
        # Update followup status
        followup.status = 'completed'
        if notes:
            followup.notes = notes
        followup.save()
        
        # Create activity log
        LeadActivity.objects.create(
            lead=followup.lead,
            user=request.user,
            action="Completed Follow-up",
            new_value=f"{followup.title} on {followup.followup_date}"
        )
        
        messages.success(request, "Follow-up marked as completed")
        
        # Redirect to add new follow-up if requested
        if add_new_followup:
            return redirect('add_followup', lead_id=followup.lead.id)
        
        return redirect('followups')
    
    return render(request, 'complete_followup.html', {'followup': followup})


@login_required
def cancel_followup(request, followup_id):
    followup = get_object_or_404(FollowUp, id=followup_id)
    
    # Check permission: admin or assigned person
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role != 'admin' and followup.assigned_to != user_profile:
        messages.error(request, "You don't have permission to cancel this follow-up")
        return redirect('followups')
    
    if request.method == "POST":
        notes = request.POST.get('notes', '')
        
        # Update followup status
        followup.status = 'cancelled'
        if notes:
            followup.notes = notes
        followup.save()
        
        # Create activity log
        LeadActivity.objects.create(
            lead=followup.lead,
            user=request.user,
            action="Cancelled Follow-up",
            new_value=f"{followup.title} on {followup.followup_date}"
        )
        
        messages.success(request, "Follow-up cancelled")
        return redirect('followups')
    
    return render(request, 'cancel_followup.html', {'followup': followup})


@login_required
def undo_followup(request, followup_id):
    followup = get_object_or_404(FollowUp, id=followup_id)
    
    # Check permission: admin or assigned person
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role != 'admin' and followup.assigned_to != user_profile:
        messages.error(request, "You don't have permission to undo this follow-up")
        return redirect('followups')
    
    # Change status back to pending (no activity log)
    followup.status = 'pending'
    followup.save()
    
    messages.success(request, "Follow-up status reset to pending")
    return redirect('followups')


@login_required
def add_followup(request, lead_id):
    lead = get_object_or_404(StudentEnquiry, id=lead_id)
    
    if request.method == "POST":
        title = request.POST.get('title')
        followup_date = request.POST.get('followup_date')
        followup_time = request.POST.get('followup_time')
        notes = request.POST.get('notes')
        
        if title and followup_date:
            # Create new follow-up
            FollowUp.objects.create(
                lead=lead,
                title=title,
                followup_date=followup_date,
                followup_time=followup_time or None,
                notes=notes or '',
                assigned_to=lead.assigned,
                created_by=request.user,
                status='pending'
            )
            
            # Create activity log
            LeadActivity.objects.create(
                lead=lead,
                user=request.user,
                action="Follow-up Added",
                new_value=f"{title} on {followup_date}"
            )
            
            messages.success(request, "New follow-up added successfully")
            return redirect('followups')
        else:
            messages.error(request, "Title and date are required")
    
    return render(request, 'add_followup.html', {'lead': lead})

#=======================================PAYMENT=========================================================================

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import timedelta
from django.core.exceptions import ValidationError
from .models import Fee, Payment

@login_required
def payments(request):
    def get_effective_fee_status(fee, student_course):
        if fee.paid_amount >= fee.total_fee:
            return 'paid'
        if student_course and student_course.batch:
            return fee.fee_status
        return 'partial' if fee.paid_amount > 0 else 'pending'

    fees_qs = Fee.objects.select_related(
        'student',
        'course'
    ).prefetch_related(
        Prefetch('installments', queryset=FeeInstallment.objects.order_by('installment_number'))
    )

    # Apply filters
    course_id = request.GET.get('course')
    if course_id:
        fees_qs = fees_qs.filter(course_id=course_id)
    
    batch_id = request.GET.get('batch')
    if batch_id:
        fees_qs = fees_qs.filter(
            student__student_courses__batch_id=batch_id
        ).distinct()
    
    status_filter = request.GET.get('status')

    latest_payment_qs = Payment.objects.filter(
        fee_id=OuterRef('pk')
    ).order_by('-payment_date', '-id')

    fees = list(fees_qs.annotate(
        latest_payment_id=Subquery(latest_payment_qs.values('id')[:1]),
        latest_payment_date=Subquery(latest_payment_qs.values('payment_date')[:1]),
    ).order_by('-due_date', '-id'))

    undo_cutoff_date = timezone.localdate() - timedelta(days=1)
    today = timezone.localdate()
    
    for fee in fees:
        fee.can_undo_latest_payment = bool(
            fee.latest_payment_id and
            fee.latest_payment_date and
            fee.latest_payment_date >= undo_cutoff_date
        )
        fee.has_installments = fee.installments.exists()

        # Calculate due date based on batch assignment
        student_course = StudentCourse.objects.filter(
            student=fee.student,
            course=fee.course
        ).first()
        
        if student_course and student_course.batch:
            fee.calculated_due_date = fee.due_date
            for inst in fee.installments.all():
                inst.update_status()
                inst.display_due_date = inst.due_date
                inst.display_status = inst.status
            
            # Update status to overdue if past due date and not fully paid
            if fee.calculated_due_date < today and fee.pending_amount > 0 and fee.fee_status != 'paid':
                fee.fee_status = 'overdue'
                fee.save(update_fields=['fee_status'])
        else:
            # No batch assigned
            fee.calculated_due_date = None
            for inst in fee.installments.all():
                inst.display_due_date = None
                inst.display_status = 'partial' if inst.paid_amount > 0 else 'pending'
                if inst.status == 'overdue' and inst.paid_amount <= 0:
                    inst.status = 'pending'
                    inst.save(update_fields=['status'])
            if fee.fee_status == 'overdue':
                fee.fee_status = 'pending' if fee.paid_amount <= 0 else 'partial'
                fee.save(update_fields=['fee_status'])

        fee.effective_fee_status = get_effective_fee_status(fee, student_course)

    if status_filter:
        fees = [fee for fee in fees if fee.effective_fee_status == status_filter]

    total_fees = len(fees)
    paid_count = sum(1 for fee in fees if fee.effective_fee_status == 'paid')
    partial_count = sum(1 for fee in fees if fee.effective_fee_status == 'partial')
    pending_count = sum(1 for fee in fees if fee.effective_fee_status == 'pending')
    overdue_count = sum(1 for fee in fees if fee.effective_fee_status == 'overdue')

    return render(request, 'payment.html', {
        'fees': fees,
        'total_fees': total_fees,
        'paid_count': paid_count,
        'partial_count': partial_count,
        'pending_count': pending_count,
        'overdue_count': overdue_count,
        'courses': Course.objects.all(),
        'batches': Batch.objects.select_related('course'),
    })


@login_required
def add_payment(request, fee_id=None):
    fee = None
    if fee_id:
        fee = get_object_or_404(
            Fee.objects.select_related('student', 'course').prefetch_related('installments'),
            id=fee_id
        )
        sync_student_fee(fee.student)
        fee = get_object_or_404(
            Fee.objects.select_related('student', 'course').prefetch_related('installments'),
            id=fee_id
        )

    if request.method == "POST":
        selected_fee_id = request.POST.get('fee_id')
        payment_mode = request.POST.get('payment_mode')
        remarks = request.POST.get('remarks', '')
        installment_id = request.POST.get('installment_id')

        if not fee:
            fee = get_object_or_404(
                Fee.objects.select_related('student', 'course').prefetch_related('installments'),
                id=selected_fee_id
            )
            sync_student_fee(fee.student)
            fee = get_object_or_404(
                Fee.objects.select_related('student', 'course').prefetch_related('installments'),
                id=selected_fee_id
            )

        try:
            amount_paid = Decimal(request.POST.get('amount_paid', '0').strip())
        except (InvalidOperation, AttributeError):
            messages.error(request, "Enter a valid payment amount.")
            return redirect('add_payment_for_fee', fee_id=fee.id)

        if amount_paid <= 0:
            messages.error(request, "Payment amount must be greater than zero.")
            return redirect('add_payment_for_fee', fee_id=fee.id)

        try:
            payment = Payment(
                fee=fee,
                amount_paid=amount_paid,
                payment_mode=payment_mode,
                remarks=remarks,
            )
            
            if installment_id:
                from .models import FeeInstallment
                try:
                    installment = FeeInstallment.objects.get(id=installment_id, fee=fee)
                    if amount_paid > installment.pending_amount:
                        messages.error(request, f"Payment amount (₹{amount_paid}) exceeds installment pending amount (₹{installment.pending_amount}).")
                        return redirect('add_payment_for_fee', fee_id=fee.id)
                    payment.installment = installment
                except FeeInstallment.DoesNotExist:
                    messages.error(request, "Invalid installment selected.")
                    return redirect('add_payment_for_fee', fee_id=fee.id)
            else:
                # For non-installment specific payments, check total pending
                if amount_paid > fee.pending_amount:
                    messages.error(request, f"Payment amount (₹{amount_paid}) exceeds total pending amount (₹{fee.pending_amount}).")
                    return redirect('add_payment_for_fee', fee_id=fee.id)
            
            payment.save()
            fee.refresh_from_db()
            _queue_payment_success_mail(
                student=fee.student,
                amount=amount_paid,
                payment_date=timezone.localdate(),
                fee=fee,
            )
            _sync_overdue_pending_mails([fee.id])
            
            # Check if payment was auto-allocated
            if not installment_id and fee.installments.exists():
                messages.success(
                    request,
                    f"Payment of ₹{amount_paid} added and auto-allocated to installments. Remaining: ₹{fee.pending_amount}"
                )
            else:
                messages.success(
                    request,
                    f"Payment of ₹{amount_paid} added successfully. Remaining: ₹{fee.pending_amount}"
                )
            return redirect('invoice_view_detail', fee_id=fee.id)
        except ValidationError as e:
            messages.error(request, e.messages[0] if e.messages else "Invalid payment.")
            return redirect('add_payment_for_fee', fee_id=fee.id)
        except Exception as e:
            messages.error(request, f"Error processing payment: {str(e)}")
            return redirect('add_payment_for_fee', fee_id=fee.id)

    fees = list(
        Fee.objects.select_related('student', 'course')
        .prefetch_related(
            Prefetch(
                'student__student_courses',
                queryset=StudentCourse.objects.select_related('course', 'batch').order_by('-id'),
            )
        )
        .order_by('-id')
    )

    for fee_item in fees:
        matched_student_course = next(
            (
                student_course
                for student_course in fee_item.student.student_courses.all()
                if student_course.course_id == fee_item.course_id and student_course.batch_id
            ),
            None,
        )
        if not matched_student_course:
            matched_student_course = next(
                (
                    student_course
                    for student_course in fee_item.student.student_courses.all()
                    if student_course.course_id == fee_item.course_id
                ),
                None,
            )

        matched_batch = matched_student_course.batch if matched_student_course else None
        course_name = fee_item.course.course_name if fee_item.course else 'No course assigned'
        batch_name = matched_batch.batch_name if matched_batch else 'No batch assigned'
        batch_key = str(matched_batch.id) if matched_batch else f"unassigned-{fee_item.course_id or 'na'}"

        fee_item.picker_course_id = str(fee_item.course_id or '')
        fee_item.picker_course_name = course_name
        fee_item.picker_batch_id = str(matched_batch.id) if matched_batch else ''
        fee_item.picker_batch_key = batch_key
        fee_item.picker_batch_name = batch_name
        fee_item.picker_student_id = str(fee_item.student_id)
        fee_item.picker_student_name = fee_item.student.student_name
        fee_item.picker_search = " ".join(
            filter(
                None,
                [
                    fee_item.student.student_name,
                    str(fee_item.student_id),
                    course_name,
                    batch_name,
                ],
            )
        ).lower()

    return render(request, 'invoice-create.html', {
        'fee': fee,
        'fees': fees,
        'payment_modes': Payment.PAYMENT_MODE_CHOICES,
    })


@login_required
def undo_payment(request, payment_id):
    if request.method != "POST":
        return redirect('payments')

    payment = get_object_or_404(
        Payment.objects.select_related('fee', 'fee__student', 'fee__course', 'installment'),
        id=payment_id
    )
    fee = payment.fee
    latest_payment = fee.payments.order_by('-payment_date', '-id').first()
    undo_cutoff_date = timezone.localdate() - timedelta(days=1)

    if not latest_payment or latest_payment.id != payment.id:
        messages.error(request, "Only the latest payment can be undone.")
        return redirect('payments')

    if payment.payment_date < undo_cutoff_date:
        messages.error(request, "Undo is available only for recent payments.")
        return redirect('payments')

    amount = payment.amount_paid
    student_name = fee.student.student_name
    installment = payment.installment

    # Delete payment first
    payment.delete()
    
    # Update installment status if linked
    if installment:
        installment.update_status()
    
    # Update fee status
    fee.recalculate_payment_status()
    _sync_overdue_pending_mails([fee.id])

    messages.success(request, f"Payment of ₹{amount} for {student_name} was undone.")
    return redirect('payments')


@login_required
def invoice_view(request, fee_id=None):
    if fee_id is None:
        fee = Fee.objects.select_related('student', 'course').prefetch_related('installments').order_by('-id').first()
        if not fee:
            messages.info(request, "No invoices found yet.")
            return redirect('payments')
        return redirect('invoice_view_detail', fee_id=fee.id)

    fee = get_object_or_404(
        Fee.objects.select_related('student', 'course').prefetch_related('installments'),
        id=fee_id
    )
    payment_history = fee.payments.all().order_by('-payment_date', '-id')
    latest_payment = payment_history.first()
    profile = UserProfile.objects.filter(user=request.user).first()
    today = timezone.localdate()
    current_paid_amount = latest_payment.amount_paid if latest_payment else fee.paid_amount
    pending_amount = fee.pending_amount
    issue_date = latest_payment.payment_date if latest_payment else today

    invoice_items = [
        {
            'title': fee.course.course_name if fee.course else 'Course Fee',
            'description': f"Current payment for {fee.student.student_name}",
            'qty': Decimal('1.00'),
            'price': current_paid_amount,
        }
    ]

    # Build address from separate fields
    student_address = ", ".join(filter(None, [
        fee.student.house_name,
        fee.student.place,
        fee.student.district,
        fee.student.state
    ])) or 'Not provided'

    return render(request, 'make_invoice.html', {
        'fee': fee,
        'payment_history': payment_history,
        'invoice_number': f"INV-{today.year}-{fee.id:04d}",
        'invoice_label': 'Tuition Fee Invoice',
        'invoice_product': fee.course.course_name if fee.course else 'Course Fee',
        'issue_date': issue_date,
        'from_name': 'FlowDesk CRM',
        'from_email': request.user.email or '',
        'from_phone': profile.mobile if profile else '',
        'from_address': 'FlowDesk CRM',
        'to_name': fee.student.student_name,
        'to_email': fee.student.email or '',
        'to_phone': fee.student.mobile or '',
        'to_address': student_address,
        'invoice_note': (
            f"Thank you {fee.student.student_name}. "
            f"Current payment received is {current_paid_amount}. "
            f"Pending fee amount is {pending_amount}."
        ),
        'current_paid_amount': current_paid_amount,
        'pending_amount': pending_amount,
        'invoice_items': invoice_items,
    })


#=======================================ATTENDANCE=========================================================================
from datetime import date, datetime
from .models import Attendance, Student, Batch
from django.http import JsonResponse

def mark_attendance(request, batch_id):

    batch = get_object_or_404(Batch, id=batch_id)
    students = Student.objects.filter(
        student_courses__batch=batch,
        student_courses__course=batch.course
    ).distinct()

    today = date.today()
    selected_date = request.GET.get('date', str(today))

    if request.method == "POST":
        att_date = request.POST.get("date")

        for student in students:
            status = request.POST.get(f"status_{student.id}")
            duration = request.POST.get(f"duration_{student.id}")
            remarks = request.POST.get(f"remarks_{student.id}")

            Attendance.objects.update_or_create(
                student=student,
                date=att_date,
                defaults={
                    "batch": batch,
                    "status": status,
                    "duration": duration,
                    "remarks": remarks,
                    "marked_by": request.user
                }
            )

        SessionUpdate.objects.update_or_create(
            batch=batch,
            session_date=att_date,
            defaults={
                "attendance_marked": True,
                "updated_by": request.user.userprofile,
            }
        )

        messages.success(request, "Attendance saved successfully")
        return redirect(f"{request.path}?date={att_date}")

    existing_attendance = {}
    for att in Attendance.objects.filter(batch=batch, date=selected_date):
        existing_attendance[att.student_id] = att

    students_with_attendance = []
    for student in students:
        att = existing_attendance.get(student.id)
        students_with_attendance.append({
            'student': student,
            'status': att.status if att else 'present',
            'duration': att.duration if att else 'full',
            'remarks': att.remarks if att else ''
        })

    return render(request, "attendance.html", {
        "students_with_attendance": students_with_attendance,
        "batch": batch,
        "today": today,
        "selected_date": selected_date
    })


@login_required
def view_attendance(request):
    user_profile = UserProfile.objects.filter(user=request.user).first()
    batches = Batch.objects.select_related('course').order_by('batch_name')
    is_session_view = request.GET.get("session_view") == "1"

    if user_profile and user_profile.role == "trainer":
        batches = batches.filter(trainer=user_profile)

    selected_batch = None
    selected_date = request.GET.get("date")
    attendance_records = Attendance.objects.none()

    batch_id = request.GET.get("batch")
    if batch_id:
        selected_batch = batches.filter(id=batch_id).first()
        if selected_batch and selected_date:
            attendance_records = Attendance.objects.select_related(
                'student', 'batch', 'marked_by'
            ).filter(
                batch=selected_batch,
                date=selected_date
            ).order_by('student__student_name')

    present_count = attendance_records.filter(status='present').count()
    absent_count = attendance_records.filter(status='absent').count()
    leave_count = attendance_records.filter(status='leave').count()
    late_count = attendance_records.filter(status='late').count()

    return render(request, "view_attendance.html", {
        "batches": batches,
        "selected_batch": selected_batch,
        "selected_date": selected_date,
        "attendance_records": attendance_records,
        "is_session_view": is_session_view,
        "present_count": present_count,
        "absent_count": absent_count,
        "leave_count": leave_count,
        "late_count": late_count,
    })

@login_required
def monthly_attendance_view(request):
    from calendar import monthrange
    
    batch_id = request.GET.get('batch')
    month_str = request.GET.get('month')
    
    if not batch_id or not month_str:
        return JsonResponse({'error': 'Batch and month are required'}, status=400)
    
    try:
        year, month = map(int, month_str.split('-'))
        days_in_month = monthrange(year, month)[1]
    except:
        return JsonResponse({'error': 'Invalid month format'}, status=400)
    
    batch = get_object_or_404(Batch, id=batch_id)
    students = Student.objects.filter(
        student_courses__batch=batch,
        student_courses__course=batch.course
    ).distinct().order_by('student_name')
    
    attendance_data = Attendance.objects.filter(
        batch=batch,
        date__year=year,
        date__month=month
    ).select_related('student')
    
    student_list = []
    for student in students:
        attendance_dict = {}
        for att in attendance_data.filter(student=student):
            attendance_dict[att.date.day] = att.status
        
        student_list.append({
            'name': student.student_name,
            'attendance': attendance_dict
        })
    
    return JsonResponse({
        'days_in_month': days_in_month,
        'students': student_list,
        'batch_name': batch.batch_name,
        'month': month_str
    })


@login_required
@role_required(['admin', 'trainer', 'management'])
def trainer_attendance_report(request):
    from django.db.models import Count
    from django.db.models.functions import TruncMonth

    user_profile = request.user.userprofile

    sessions = SessionUpdate.objects.select_related(
        'batch',
        'batch__course',
        'updated_by',
        'updated_by__user'
    ).filter(updated_by__role='trainer')

    trainers = UserProfile.objects.select_related('user').filter(role='trainer').order_by(
        'user__first_name', 'user__username'
    )
    batches = Batch.objects.select_related('course', 'trainer', 'trainer__user').order_by('batch_name')

    if user_profile.role == 'trainer':
        sessions = sessions.filter(updated_by=user_profile)
        trainers = trainers.filter(id=user_profile.id)
        batches = batches.filter(trainer=user_profile)

    selected_trainer = (request.GET.get('trainer') or '').strip()
    selected_batch = (request.GET.get('batch') or '').strip()
    selected_day = (request.GET.get('day') or '').strip()
    selected_month = (request.GET.get('month') or '').strip()
    selected_from_date = (request.GET.get('from_date') or '').strip()
    selected_to_date = (request.GET.get('to_date') or '').strip()
    selected_attendance = (request.GET.get('attendance') or 'all').strip()
    has_active_filters = any([
        bool(selected_trainer),
        bool(selected_batch),
        bool(selected_day),
        bool(selected_month),
        bool(selected_from_date),
        bool(selected_to_date),
        selected_attendance in ('marked', 'not_marked'),
    ])

    def parse_iso_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None

    if selected_trainer and user_profile.role != 'trainer':
        sessions = sessions.filter(updated_by_id=selected_trainer)

    if selected_batch:
        sessions = sessions.filter(batch_id=selected_batch)

    if selected_attendance == 'marked':
        sessions = sessions.filter(attendance_marked=True)
    elif selected_attendance == 'not_marked':
        sessions = sessions.filter(attendance_marked=False)

    day_value = parse_iso_date(selected_day)
    if day_value:
        sessions = sessions.filter(session_date=day_value)

    if selected_month:
        try:
            month_year, month_number = selected_month.split('-')
            sessions = sessions.filter(
                session_date__year=int(month_year),
                session_date__month=int(month_number),
            )
        except (ValueError, TypeError):
            selected_month = ''

    from_date_value = parse_iso_date(selected_from_date)
    to_date_value = parse_iso_date(selected_to_date)

    if from_date_value:
        sessions = sessions.filter(session_date__gte=from_date_value)

    if to_date_value:
        sessions = sessions.filter(session_date__lte=to_date_value)

    sessions = sessions.order_by('-session_date', 'batch__batch_name')

    summary = sessions.aggregate(
        total_updates=Count('id'),
        marked_updates=Count('id', filter=Q(attendance_marked=True)),
        not_marked_updates=Count('id', filter=Q(attendance_marked=False)),
        active_trainers=Count('updated_by', distinct=True),
        active_days=Count('session_date', distinct=True),
    )

    def trainer_name(first_name, last_name, username):
        full_name = f"{(first_name or '').strip()} {(last_name or '').strip()}".strip()
        return full_name if full_name else (username or '-')

    monthly_summary_rows = []
    monthly_summary_qs = sessions.annotate(month=TruncMonth('session_date')).values(
        'month',
        'updated_by__user__first_name',
        'updated_by__user__last_name',
        'updated_by__user__username',
    ).annotate(
        total_updates=Count('id'),
        marked_updates=Count('id', filter=Q(attendance_marked=True)),
        not_marked_updates=Count('id', filter=Q(attendance_marked=False)),
        working_days=Count('session_date', distinct=True),
    ).order_by('-month', 'updated_by__user__first_name', 'updated_by__user__username')

    for row in monthly_summary_qs:
        total_updates = row['total_updates'] or 0
        marked_updates = row['marked_updates'] or 0
        marked_percent = round((marked_updates / total_updates) * 100, 1) if total_updates else 0
        monthly_summary_rows.append({
            'month': row['month'],
            'trainer': trainer_name(
                row['updated_by__user__first_name'],
                row['updated_by__user__last_name'],
                row['updated_by__user__username'],
            ),
            'working_days': row['working_days'],
            'total_updates': total_updates,
            'marked_updates': marked_updates,
            'not_marked_updates': row['not_marked_updates'] or 0,
            'marked_percent': marked_percent,
        })

    day_summary_rows = []
    day_summary_qs = sessions.values(
        'session_date',
        'updated_by__user__first_name',
        'updated_by__user__last_name',
        'updated_by__user__username',
    ).annotate(
        batch_count=Count('batch_id', distinct=True),
        total_updates=Count('id'),
        marked_updates=Count('id', filter=Q(attendance_marked=True)),
        not_marked_updates=Count('id', filter=Q(attendance_marked=False)),
    ).order_by('-session_date', 'updated_by__user__first_name', 'updated_by__user__username')

    for row in day_summary_qs:
        total_updates = row['total_updates'] or 0
        marked_updates = row['marked_updates'] or 0
        marked_percent = round((marked_updates / total_updates) * 100, 1) if total_updates else 0
        day_summary_rows.append({
            'session_date': row['session_date'],
            'trainer': trainer_name(
                row['updated_by__user__first_name'],
                row['updated_by__user__last_name'],
                row['updated_by__user__username'],
            ),
            'batch_count': row['batch_count'],
            'total_updates': total_updates,
            'marked_updates': marked_updates,
            'not_marked_updates': row['not_marked_updates'] or 0,
            'marked_percent': marked_percent,
        })

    detail_limit = 300
    detail_sessions = list(sessions[:detail_limit])
    is_detail_trimmed = (summary.get('total_updates') or 0) > detail_limit

    return render(request, "trainer_attendance_report.html", {
        "trainers": trainers,
        "batches": batches,
        "sessions": detail_sessions,
        "summary": summary,
        "monthly_summary_rows": monthly_summary_rows,
        "day_summary_rows": day_summary_rows,
        "is_detail_trimmed": is_detail_trimmed,
        "detail_limit": detail_limit,
        "selected_trainer": selected_trainer,
        "selected_batch": selected_batch,
        "selected_day": selected_day,
        "selected_month": selected_month,
        "selected_from_date": selected_from_date,
        "selected_to_date": selected_to_date,
        "selected_attendance": selected_attendance,
        "has_active_filters": has_active_filters,
    })

#=======================================SeSSION_UPDATE=========================================================================

from .models import SessionUpdate, TopicProgress, Topic, Assignment, AssignmentSubmission

@login_required
def session_update(request):

    user_profile = request.user.userprofile

    # Trainer sees only his batch
    if user_profile.role == "trainer":
        batches = Batch.objects.filter(trainer=user_profile, status="Active")
    else:
        batches = Batch.objects.filter(status="Active")

    if request.method == "POST":

        batch_id = request.POST.get("batch")
        session_date = request.POST.get("session_date")
        attendance_marked = True if request.POST.get("attendance_marked") else False
        assignment_given = (request.POST.get("assignment_given") or "").strip()
        assignment_marks_raw = (request.POST.get("assignment_marks") or "").strip()
        remarks = request.POST.get("remarks")

        batch = Batch.objects.get(id=batch_id)
        assignment_marks = None
        if assignment_marks_raw:
            try:
                assignment_marks = int(assignment_marks_raw)
            except ValueError:
                messages.error(request, "Assignment marks must be a valid number.")
                return redirect("session_update")
            if assignment_marks < 0:
                messages.error(request, "Assignment marks cannot be negative.")
                return redirect("session_update")

        if assignment_given:
            normalized_assignment_title = _normalize_title_key(assignment_given)
            existing_assignment_titles = Assignment.objects.filter(
                batch=batch
            ).values_list("title", flat=True)
            if any(_normalize_title_key(existing_title) == normalized_assignment_title for existing_title in existing_assignment_titles):
                messages.error(request, "An assignment with this name already exists for the selected batch.")
                return redirect("session_update")

        #  Block duplicate session for same batch + date
        if SessionUpdate.objects.filter(batch=batch, session_date=session_date).exists():
            messages.error(request, f"A session update for {batch.batch_name} on {session_date} already exists. You cannot add another session for the same batch and date.")
            return redirect("session_update")

        #  Create SessionUpdate
        session_obj = SessionUpdate.objects.create(
            batch=batch,
            session_date=session_date,
            attendance_marked=attendance_marked,
            assignment_given=assignment_given,
            assignment_marks=assignment_marks,
            remarks=remarks,
            updated_by=user_profile,
        )

        #  Update TopicProgress
        for key in request.POST:
            if key.startswith("status_"):
                topic_id = key.split("_")[1]
                status = request.POST.get(key)
                topic_remarks = request.POST.get(f"remarks_{topic_id}")

                topic = Topic.objects.get(id=topic_id)

                TopicProgress.objects.update_or_create(
                    batch=batch,
                    topic=topic,
                    defaults={
                        "status": status,
                        "remarks": topic_remarks,
                        "updated_by": user_profile,
                    }
                )

        # Create Assignment if given
        if assignment_given:
            assignment_desc = request.POST.get("assignment_description", "")
            due_date = request.POST.get("assignment_due_date")
            
            assignment = Assignment.objects.create(
                batch=batch,
                session=session_obj,
                title=assignment_given,
                description=assignment_desc,
                due_date=due_date if due_date else session_date,
                created_by=user_profile
            )
            
            # Create submission records for all students in batch
            students = Student.objects.filter(
                student_courses__batch=batch,
                student_courses__course=batch.course
            ).distinct()
            for student in students:
                AssignmentSubmission.objects.create(
                    assignment=assignment,
                    student=student,
                    status='pending'
                )

        messages.success(request, "Session update saved successfully!")
        return redirect("session_update")

    context = {
        "batches": batches,
    }

    return render(request, "session_update.html", context)



@login_required
def get_batch_topics(request):

    batch_id = request.GET.get("batch_id")

    if not batch_id:
        return JsonResponse({"error": "Batch ID missing"}, status=400)

    batch = get_object_or_404(Batch, id=batch_id)

    user_profile = request.user.userprofile

    #  Security: Trainer should only access his batch
    if user_profile.role == "trainer":
        if batch.trainer != user_profile:
            return JsonResponse({"error": "Unauthorized"}, status=403)

    # Get modules of that batch's course
    modules = Module.objects.filter(course=batch.course)

    topics_data = []

    for module in modules:
        topics = Topic.objects.filter(module=module)

        for topic in topics:

            progress = TopicProgress.objects.filter(
                batch=batch,
                topic=topic
            ).first()

            topics_data.append({
                "id": topic.id,
                "name": topic.topic_name,
                "module": module.module_name,
                "status": progress.status if progress else "not_started",
                "remarks": progress.remarks if progress else ""
            })

    return JsonResponse({"topics": topics_data})


from .models import SessionUpdate, Batch, Attendance


@login_required
def session_update_list(request):

    user_profile = request.user.userprofile
    is_trainer_user = user_profile.role == "trainer"

    sessions = SessionUpdate.objects.select_related(
        "batch",
        "batch__course",
        "updated_by"
    ).annotate(
        has_attendance_records=Exists(
            Attendance.objects.filter(
                batch_id=OuterRef("batch_id"),
                date=OuterRef("session_date"),
            )
        )
    ).order_by("-session_date")

    #  ROLE BASED FILTERING

    # Trainer → only his batches
    if is_trainer_user:
        sessions = sessions.filter(batch__trainer=user_profile)

    # CRE → no access (optional)
    if user_profile.role == "cre":
        sessions = SessionUpdate.objects.none()

    #  Filters
    batch_id = request.GET.get("batch")
    trainer_id = request.GET.get("trainer")
    course_id = request.GET.get("course")
    date = request.GET.get("date")

    if batch_id:
        sessions = sessions.filter(batch_id=batch_id)

    if trainer_id:
        sessions = sessions.filter(batch__trainer_id=trainer_id)

    if course_id:
        sessions = sessions.filter(batch__course_id=course_id)

    if date:
        sessions = sessions.filter(session_date=date)

    if is_trainer_user:
        batches = Batch.objects.filter(status="Active", trainer=user_profile).select_related("course").order_by("batch_name")
        trainers = UserProfile.objects.filter(pk=user_profile.pk).select_related("user")
        courses = Course.objects.filter(batch__status="Active", batch__trainer=user_profile).select_related("department").distinct().order_by("course_name", "department__department_name")
        selected_trainer_id = str(user_profile.pk)
    else:
        batches = Batch.objects.filter(status="Active").select_related("course").order_by("batch_name")
        trainers = UserProfile.objects.filter(role="trainer").select_related("user").order_by("user__first_name", "user__username")
        courses = Course.objects.select_related("department").order_by("course_name", "department__department_name")
        selected_trainer_id = trainer_id or ""

    context = {
        "sessions": sessions,
        "batches": batches,
        "trainers": trainers,
        "courses": courses,
        "is_trainer_user": is_trainer_user,
        "selected_trainer_id": selected_trainer_id,
    }

    return render(request, "session_update_list.html", context)




@login_required
def edit_session_update(request, session_id):
    from .models import Assignment, AssignmentSubmission
    user_profile = request.user.userprofile
    session_obj = get_object_or_404(SessionUpdate, id=session_id)

    if user_profile.role == 'trainer' and session_obj.batch.trainer != user_profile:
        messages.error(request, "You don't have permission to edit this session.")
        return redirect('session_update_list')

    existing_assignment = Assignment.objects.filter(
        batch=session_obj.batch, session=session_obj
    ).first()

    if request.method == 'POST':
        new_title = (request.POST.get('assignment_given') or '').strip()
        old_title = (session_obj.assignment_given or '').strip()
        new_description = (request.POST.get('assignment_description') or '').strip()
        new_due_date = (request.POST.get('assignment_due_date') or '').strip() or None
        assignment_marks_raw = (request.POST.get('assignment_marks') or '').strip()
        new_marks = None
        if assignment_marks_raw:
            try:
                new_marks = int(assignment_marks_raw)
            except ValueError:
                messages.error(request, "Assignment marks must be a valid number.")
                return redirect('edit_session_update', session_id=session_id)

        session_obj.assignment_given = new_title
        session_obj.assignment_marks = new_marks
        session_obj.save(update_fields=['assignment_given', 'assignment_marks'])

        if old_title and new_title:
            # Update existing assignment record
            Assignment.objects.filter(
                batch=session_obj.batch, session=session_obj
            ).update(
                title=new_title,
                description=new_description,
                due_date=new_due_date or session_obj.session_date,
            )
        elif old_title and not new_title:
            Assignment.objects.filter(
                batch=session_obj.batch, session=session_obj
            ).delete()
        elif not old_title and new_title:
            normalized = _normalize_title_key(new_title)
            existing_titles = Assignment.objects.filter(batch=session_obj.batch).values_list('title', flat=True)
            if not any(_normalize_title_key(t) == normalized for t in existing_titles):
                assignment = Assignment.objects.create(
                    batch=session_obj.batch,
                    session=session_obj,
                    title=new_title,
                    description=new_description,
                    due_date=new_due_date or session_obj.session_date,
                    created_by=user_profile,
                )
                students = Student.objects.filter(
                    student_courses__batch=session_obj.batch,
                    student_courses__course=session_obj.batch.course
                ).distinct()
                for student in students:
                    AssignmentSubmission.objects.get_or_create(
                        assignment=assignment, student=student, defaults={'status': 'pending'}
                    )

        for key in request.POST:
            if key.startswith('status_'):
                topic_id = key.split('_')[1]
                try:
                    topic = Topic.objects.get(id=topic_id)
                    TopicProgress.objects.update_or_create(
                        batch=session_obj.batch,
                        topic=topic,
                        defaults={
                            'status': request.POST.get(key),
                            'remarks': request.POST.get(f'remarks_{topic_id}', ''),
                            'updated_by': user_profile,
                        }
                    )
                except Topic.DoesNotExist:
                    pass

        messages.success(request, "Session update edited successfully.")
        return redirect('session_update_list')

    topics_data = []
    for module in Module.objects.filter(course=session_obj.batch.course):
        for topic in Topic.objects.filter(module=module):
            progress = TopicProgress.objects.filter(batch=session_obj.batch, topic=topic).first()
            topics_data.append({
                'id': topic.id,
                'name': topic.topic_name,
                'module': module.module_name,
                'status': progress.status if progress else 'not_started',
                'remarks': progress.remarks if progress else '',
            })

    return render(request, 'edit_session_update.html', {
        'session': session_obj,
        'existing_assignment': existing_assignment,
        'topics_data': topics_data,
    })


@login_required
def delete_session_update(request, session_id):
    user_profile = request.user.userprofile
    if user_profile.role != 'admin':
        messages.error(request, "Only admin can delete session updates.")
        return redirect('session_update_list')
    session_obj = get_object_or_404(SessionUpdate, id=session_id)
    session_obj.delete()
    messages.success(request, "Session update deleted successfully.")
    return redirect('session_update_list')


#=======================================ASSIGNMENT TRACKING=========================================================================

@login_required
def assignment_list(request):
    user_profile = request.user.userprofile
    can_filter_created_by = user_profile.role != "trainer"
    
    assignments = Assignment.objects.select_related(
        'batch', 'batch__course', 'session', 'created_by', 'created_by__user'
    ).order_by('-created_at')
    batches = Batch.objects.select_related('course').filter(status='Active').order_by('batch_name')
    
    if user_profile.role == "trainer":
        assignments = assignments.filter(batch__trainer=user_profile)
        batches = batches.filter(trainer=user_profile)

    creator_pool = assignments

    q = (request.GET.get('q') or '').strip()
    if q:
        assignments = assignments.filter(
            Q(title__icontains=q) |
            Q(description__icontains=q) |
            Q(batch__batch_name__icontains=q) |
            Q(created_by__user__first_name__icontains=q) |
            Q(created_by__user__last_name__icontains=q) |
            Q(created_by__user__username__icontains=q)
        )

    date_from = request.GET.get('date_from')
    if date_from:
        assignments = assignments.filter(due_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        assignments = assignments.filter(due_date__lte=date_to)
    
    batch_id = request.GET.get('batch')
    if batch_id:
        assignments = assignments.filter(batch_id=batch_id)

    course_id = request.GET.get('course')
    if course_id:
        assignments = assignments.filter(batch__course_id=course_id)

    if can_filter_created_by:
        created_by_id = request.GET.get('created_by')
        if created_by_id:
            assignments = assignments.filter(created_by_id=created_by_id)

    courses = Course.objects.filter(
        id__in=batches.values_list('course_id', flat=True)
    ).select_related('department').distinct().order_by('course_name')

    if can_filter_created_by:
        assignment_creators = UserProfile.objects.select_related('user').filter(
            id__in=creator_pool.values_list('created_by_id', flat=True).distinct()
        ).order_by('user__first_name', 'user__username')
    else:
        assignment_creators = UserProfile.objects.none()

    try:
        per_page = int(request.GET.get('per_page', 10))
    except (TypeError, ValueError):
        per_page = 10

    paginator = Paginator(assignments, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    context = {
        'assignments': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'batches': batches,
        'courses': courses,
        'assignment_creators': assignment_creators,
        'can_filter_created_by': can_filter_created_by,
    }
    return render(request, 'assignment_list.html', context)

@login_required
def assignment_detail(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)
    submissions = AssignmentSubmission.objects.filter(
        assignment=assignment
    ).select_related('student')
    
    if request.method == 'POST':
        for submission in submissions:
            status_key = f'status_{submission.id}'
            remarks_key = f'remarks_{submission.id}'
            
            if status_key in request.POST:
                submission.status = request.POST.get(status_key)
                submission.remarks = request.POST.get(remarks_key, '')
                if submission.status == 'submitted':
                    submission.submitted_date = timezone.now()
                submission.save()
        
        messages.success(request, 'Assignment submissions updated')
        return redirect('assignment_detail', assignment_id=assignment_id)
    
    context = {
        'assignment': assignment,
        'submissions': submissions
    }
    return render(request, 'assignment_detail.html', context)


@login_required
def delete_assignment(request, assignment_id):
    if request.method != 'POST':
        return redirect('assignment_list')

    user_profile = request.user.userprofile
    if user_profile.role != "admin":
        messages.error(request, "Only admin users can delete assignments.")
        return redirect('assignment_list')

    assignment = get_object_or_404(
        Assignment.objects.select_related('batch'),
        id=assignment_id
    )

    assignment_title = assignment.title
    assignment.delete()
    messages.success(request, f'Assignment "{assignment_title}" deleted successfully.')
    return redirect('assignment_list')


@login_required
def exam_list(request):
    user_profile = request.user.userprofile

    batches = Batch.objects.filter(status='Active')
    if user_profile.role == "trainer":
        batches = batches.filter(trainer=user_profile)

    if request.method == 'POST':
        batch_id = request.POST.get('batch')
        title = request.POST.get('title', '').strip()
        exam_date = request.POST.get('exam_date')
        max_marks = request.POST.get('max_marks') or 100

        if not batch_id or not title or not exam_date:
            messages.error(request, "Batch, exam title, and exam date are required.")
            return redirect('exam_list')

        batch = get_object_or_404(Batch, id=batch_id)
        if user_profile.role == "trainer" and batch.trainer != user_profile:
            messages.error(request, "You can only create exams for your own batches.")
            return redirect('exam_list')

        normalized_exam_title = _normalize_title_key(title)
        existing_exam_titles = Exam.objects.filter(batch=batch).values_list('title', flat=True)
        if any(_normalize_title_key(existing_title) == normalized_exam_title for existing_title in existing_exam_titles):
            messages.error(request, "An exam with this name already exists for the selected batch.")
            return redirect('exam_list')

        Exam.objects.create(
            batch=batch,
            title=title,
            exam_date=exam_date,
            max_marks=max_marks,
            created_by=user_profile
        )
        messages.success(request, "Exam created successfully.")
        return redirect('exam_list')

    exams = Exam.objects.select_related('batch', 'created_by', 'created_by__user').order_by('-exam_date', '-created_at')

    if user_profile.role == "trainer":
        exams = exams.filter(batch__trainer=user_profile)

    batch_id = request.GET.get('batch')
    if batch_id:
        exams = exams.filter(batch_id=batch_id)

    try:
        per_page = int(request.GET.get('per_page', 10))
    except (TypeError, ValueError):
        per_page = 10

    paginator = Paginator(exams, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'exams': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'batches': batches,
    }
    return render(request, 'exam_list.html', context)


@login_required
def delete_exam(request, exam_id):
    if request.method != 'POST':
        return redirect('exam_list')

    user_profile = request.user.userprofile
    if user_profile.role != "admin":
        messages.error(request, "Only admin users can delete exams.")
        return redirect('exam_list')

    exam = get_object_or_404(
        Exam.objects.select_related('batch'),
        id=exam_id
    )

    exam_title = exam.title
    exam.delete()
    messages.success(request, f'Exam "{exam_title}" deleted successfully.')
    return redirect('exam_list')


@login_required
def exam_detail(request, exam_id):
    exam = get_object_or_404(Exam.objects.select_related('batch'), id=exam_id)
    user_profile = request.user.userprofile

    if user_profile.role == "trainer" and exam.batch.trainer != user_profile:
        messages.error(request, "You can only view exams from your own batches.")
        return redirect('exam_list')

    batch_students = Student.objects.filter(
        student_courses__batch=exam.batch,
        student_courses__course=exam.batch.course
    ).distinct().order_by('student_name')
    for student in batch_students:
        ExamPerformance.objects.get_or_create(exam=exam, student=student)

    performances = ExamPerformance.objects.filter(exam=exam).select_related('student').order_by('student__student_name')

    if request.method == 'POST':
        today = timezone.localdate()
        if exam.exam_date > today:
            messages.error(request, "Cannot fill exam details before the exam date.")
            return redirect('exam_detail', exam_id=exam_id)

        for performance in performances:
            score_key = f'score_{performance.id}'
            remarks_key = f'remarks_{performance.id}'

            if score_key in request.POST:
                raw_score = request.POST.get(score_key, '').strip()
                score_value = None
                if raw_score:
                    try:
                        score_value = Decimal(raw_score)
                    except InvalidOperation:
                        messages.error(request, f"Invalid score for {performance.student.student_name}.")
                        return redirect('exam_detail', exam_id=exam_id)

                performance.score = score_value
                performance.remarks = request.POST.get(remarks_key, '').strip()
                performance.save()

        messages.success(request, "Exam performance updated.")
        return redirect('exam_detail', exam_id=exam_id)

    context = {
        'exam': exam,
        'performances': performances,
    }
    return render(request, 'exam_detail.html', context)



#=======================================ADMIN=========================================================================

# ==================== CRE VIEWS ====================

def cre_dashboard(request):
    return render(request, 'cre_base.html', {'page': 'dashboard'})


@login_required
def cre_leads(request):
    user_profile = UserProfile.objects.filter(user=request.user).first()
    
    leads_qs = StudentEnquiry.objects.select_related(
        'status', 'source', 'assigned', 'created_by'
    ).filter(assigned=user_profile).order_by('-enquiry_date')

    # Search
    q = request.GET.get('q')
    if q:
        leads_qs = leads_qs.filter(
            Q(full_name__icontains=q) |
            Q(email__icontains=q) |
            Q(mobile__icontains=q)
        )

    # Date filter
    date_from = request.GET.get('date_from')
    if date_from:
        leads_qs = leads_qs.filter(enquiry_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        leads_qs = leads_qs.filter(enquiry_date__lte=date_to)

    # Status filter
    status = request.GET.get('status')
    if status:
        leads_qs = leads_qs.filter(status_id=status)

    # Source filter
    source = request.GET.get('source')
    if source:
        leads_qs = leads_qs.filter(source_id=source)

    # Pagination
    per_page = int(request.GET.get('per_page', 10))
    paginator = Paginator(leads_qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'leads': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'statuses': Status.objects.all(),
        'sources': Source.objects.all(),
        'lead': StudentEnquiry,
    }
    return render(request, 'cre/leads.html', context)


def cre_create_lead(request):
    return render(request, 'cre/create_lead.html')


@login_required
@role_required(['cre'])
def cre_followups(request):
    user_profile = UserProfile.objects.filter(user=request.user).first()
    context = {
        'role': 'cre',
        'show_payment_cards': True,
        **_get_cre_followup_context(user_profile),
    }
    return render(request, 'dashboard.html', context)

#===========================================================REPORTS=============================================
@role_required(['admin', 'management'])
def lead_report(request):
    from .models import StudentEnquiry

    leads_qs = StudentEnquiry.objects.select_related(
        'status', 'source', 'assigned__user', 'created_by'
    ).prefetch_related(
        'lead_courses__department',
        'lead_courses__course'
    ).order_by('-enquiry_date', '-id')

    q = request.GET.get('q', '').strip()
    if q:
        leads_qs = leads_qs.filter(
            Q(full_name__icontains=q) |
            Q(email__icontains=q) |
            Q(mobile__icontains=q)
        )

    date_from = request.GET.get('date_from')
    if date_from:
        leads_qs = leads_qs.filter(enquiry_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        leads_qs = leads_qs.filter(enquiry_date__lte=date_to)

    status_id = request.GET.get('status')
    if status_id:
        leads_qs = leads_qs.filter(status_id=status_id)

    department_id = request.GET.get('department')
    if department_id:
        leads_qs = leads_qs.filter(lead_courses__department_id=department_id).distinct()

    course_id = request.GET.get('course')
    if course_id:
        leads_qs = leads_qs.filter(lead_courses__course_id=course_id).distinct()

    assigned_to = request.GET.get('assigned_to')
    if assigned_to:
        leads_qs = leads_qs.filter(assigned_id=assigned_to)

    source = request.GET.get('source')
    if source:
        leads_qs = leads_qs.filter(source_id=source)

    mode = request.GET.get('mode')
    if mode:
        leads_qs = leads_qs.filter(mode=mode)

    created_by_id = request.GET.get('created_by')
    if created_by_id:
        leads_qs = leads_qs.filter(created_by_id=created_by_id)

    followup_state = request.GET.get('followup_state')
    today = timezone.localdate()
    if followup_state == 'with':
        leads_qs = leads_qs.filter(followup_date__isnull=False)
    elif followup_state == 'without':
        leads_qs = leads_qs.filter(followup_date__isnull=True)
    elif followup_state == 'today':
        leads_qs = leads_qs.filter(followup_date=today)
    elif followup_state == 'overdue':
        leads_qs = leads_qs.filter(followup_date__lt=today)
    elif followup_state == 'upcoming':
        leads_qs = leads_qs.filter(followup_date__gt=today)

    followup_from = request.GET.get('followup_from')
    if followup_from:
        leads_qs = leads_qs.filter(followup_date__gte=followup_from)

    followup_to = request.GET.get('followup_to')
    if followup_to:
        leads_qs = leads_qs.filter(followup_date__lte=followup_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(leads_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Calculate converted count (enrolled status)
    converted_count = StudentEnquiry.objects.filter(
        status__status_name__iexact='enrolled'
    ).count()

    # Calculate conversion rate percentage
    total_leads = StudentEnquiry.objects.count()
    conversion_rate = round((converted_count / total_leads * 100), 2) if total_leads > 0 else 0

    # Get upcoming follow-ups ordered by date (latest first)
    upcoming_followups = StudentEnquiry.objects.filter(
        followup_date__isnull=False
    ).select_related('status').prefetch_related(
        'lead_courses__course'
    ).order_by('-followup_date', '-followup_time')[:5]

    # Get latest 9 leads for Leads Status card
    latest_leads = StudentEnquiry.objects.select_related(
        'status', 'assigned__user'
    ).prefetch_related(
        'lead_courses__course'
    ).order_by('-enquiry_date', '-id')[:9]

    # Get status distribution for chart
    from django.db.models import Count
    status_distribution = leads_qs.values('status__status_name').annotate(
        count=Count('id')
    ).order_by('-count')

    # Get source distribution for chart
    source_distribution = leads_qs.values('source__source_name').annotate(
        count=Count('id')
    ).order_by('-count')

    # Calculate hot leads (last 7 days)
    from datetime import timedelta
    seven_days_ago = timezone.localdate() - timedelta(days=7)
    hot_leads_count = leads_qs.filter(enquiry_date__gte=seven_days_ago).count()

    # Calculate enrolled count (converted leads)
    enrolled_count = StudentEnquiry.objects.filter(
        status__status_name__iexact='enrolled'
    ).count()

    # Calculate conversion rate percentage
    total_leads = StudentEnquiry.objects.count()
    conversion_rate = round((enrolled_count / total_leads * 100), 2) if total_leads > 0 else 0

    context = {
        'leads': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_count': leads_qs.count(),
        'today_count': leads_qs.filter(enquiry_date=timezone.localdate()).count(),
        'followup_count': leads_qs.filter(followup_date__isnull=False).count(),
        'hot_leads_count': hot_leads_count,
        'enrolled_count': enrolled_count,
        'converted_count': conversion_rate,
        'upcoming_followups': upcoming_followups,
        'latest_leads': latest_leads,
        'status_distribution': status_distribution,
        'source_distribution': source_distribution,
        'statuses': Status.objects.all(),
        'sources': Source.objects.all(),
        'departments': Department.objects.all(),
        'courses': Course.objects.select_related('department'),
        'users': UserProfile.objects.select_related('user'),
        'mode_choices': StudentEnquiry.MODE_CHOICES,
        'created_by_users': User.objects.filter(
            id__in=StudentEnquiry.objects.exclude(created_by__isnull=True).values_list('created_by_id', flat=True).distinct()
        ).order_by('first_name', 'username'),
    }
    return render(request, 'reports/lead_report.html', context)

def _build_student_individual_report_data(student):
    from .models import Attendance, Assignment, AssignmentSubmission, Exam, ExamPerformance, Fee

    student_batches = student.student_courses.filter(batch__isnull=False).select_related('batch', 'course')
    report_data = []

    for sc in student_batches:
        batch = sc.batch
        course = sc.course

        attendance_qs = Attendance.objects.filter(student=student, batch=batch)
        total_attendance = attendance_qs.count()
        present_count = attendance_qs.filter(status__in=['present', 'late']).count()
        absent_count = attendance_qs.filter(status='absent').count()
        leave_count = attendance_qs.filter(status='leave').count()
        attendance_percent = round((present_count / total_attendance) * 100, 1) if total_attendance > 0 else 0

        assignments = Assignment.objects.filter(batch=batch)
        total_assignments = assignments.count()
        submissions = AssignmentSubmission.objects.filter(
            student=student,
            assignment__batch=batch
        )
        submitted_count = submissions.filter(status='submitted').count()
        pending_count = submissions.filter(status='pending').count()
        assignment_percent = round((submitted_count / total_assignments) * 100, 1) if total_assignments > 0 else 0

        exams = Exam.objects.filter(batch=batch)
        total_exams = exams.count()
        performances = ExamPerformance.objects.filter(
            student=student,
            exam__batch=batch
        ).select_related('exam')

        exam_list = []
        total_scored = 0
        total_max = 0
        for exam in exams:
            perf = performances.filter(exam=exam).first()
            score = perf.score if perf and perf.score else 0
            total_scored += float(score)
            total_max += exam.max_marks
            exam_list.append({
                'title': exam.title,
                'date': exam.exam_date,
                'max_marks': exam.max_marks,
                'score': score,
                'percentage': round((float(score) / exam.max_marks) * 100, 1) if exam.max_marks > 0 else 0
            })

        exam_percent = round((total_scored / total_max) * 100, 1) if total_max > 0 else 0

        fee = Fee.objects.filter(student=student, course=course).first()
        fee_data = None
        if fee:
            payments = fee.payments.all().order_by('-payment_date')
            fee_data = {
                'total_fee': fee.total_fee,
                'paid_amount': fee.paid_amount,
                'pending_amount': fee.pending_amount,
                'status': fee.fee_status,
                'payments': payments
            }

        report_data.append({
            'course': course,
            'batch': batch,
            'attendance': {
                'total': total_attendance,
                'present': present_count,
                'absent': absent_count,
                'leave': leave_count,
                'percentage': attendance_percent
            },
            'assignments': {
                'total': total_assignments,
                'submitted': submitted_count,
                'pending': pending_count,
                'percentage': assignment_percent
            },
            'exams': {
                'total': total_exams,
                'percentage': exam_percent,
                'exam_list': exam_list
            },
            'fees': fee_data
        })

    return report_data


@role_required(['admin', 'management'])
def student_individual_report(request, student_id):
    from .models import Student

    student = get_object_or_404(
        Student.objects.prefetch_related(
            'student_courses__course',
            'student_courses__batch',
            'fees__payments'
        ),
        id=student_id
    )
    report_data = _build_student_individual_report_data(student)
    context = {
        'student': student,
        'report_data': report_data
    }
    return render(request, 'reports/student_individual_report.html', context)


@role_required(['admin', 'management'])
def send_student_report_email(request, student_id):
    from .models import Student
    from django.core.exceptions import ValidationError
    from django.core.validators import validate_email

    if request.method != "POST":
        return redirect('student_individual_report', student_id=student_id)

    student = get_object_or_404(
        Student.objects.prefetch_related(
            'student_courses__course',
            'student_courses__batch',
            'fees__payments'
        ),
        id=student_id
    )

    recipient_type = (request.POST.get('recipient_type') or 'student').strip().lower()
    parent_email = (request.POST.get('parent_email') or '').strip()

    if recipient_type == 'parent':
        recipient_email = parent_email
    else:
        recipient_type = 'student'
        recipient_email = (student.email or '').strip()

    if not recipient_email:
        if recipient_type == 'student':
            messages.error(request, "Student email is missing. Please use parent email option.")
        else:
            messages.error(request, "Please enter a parent email address.")
        return redirect('student_individual_report', student_id=student_id)

    try:
        validate_email(recipient_email)
    except ValidationError:
        messages.error(request, "Enter a valid email address.")
        return redirect('student_individual_report', student_id=student_id)

    report_data = _build_student_individual_report_data(student)
    company_name = getattr(settings, "MAIL_COMPANY_NAME", "") or getattr(settings, "COMPANY_NAME", "") or "FlowDesk CRM"

    context = {
        'student': student,
        'report_data': report_data,
        'company_name': company_name,
        'generated_at': timezone.localtime(),
    }

    subject = f"{student.student_name} - Academic & Financial Report"
    html_body = render_to_string('email/student_report_email.html', context)
    text_body = strip_tags(html_body).strip() or "Student report"
    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", "") or "no-reply@flowdesk.local"

    try:
        message = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=from_email,
            to=[recipient_email],
        )
        message.attach_alternative(html_body, "text/html")
        message.send(fail_silently=False)
        messages.success(request, f"Student report sent successfully to {recipient_email}.")
    except Exception as exc:
        raw_error = str(exc)
        error_message = raw_error
        if "10061" in raw_error or "Connection refused" in raw_error:
            error_message = (
                "Could not connect to the email server. "
                "Please configure SMTP values in .env "
                "(EMAIL_BACKEND, EMAIL_HOST, EMAIL_PORT, EMAIL_HOST_USER, "
                "EMAIL_HOST_PASSWORD, EMAIL_USE_TLS)."
            )
        messages.error(request, error_message)

    return redirect('student_individual_report', student_id=student_id)


@role_required(['admin', 'management'])
def student_report(request):
    from .models import Student, Batch
    from datetime import timedelta
    from django.db.models import Count

    students_qs = Student.objects.prefetch_related(
        'student_courses__course', 'student_courses__department', 'student_courses__batch'
    ).order_by('-enrolled_date', '-id')

    q = request.GET.get('q', '').strip()
    if q:
        students_qs = students_qs.filter(
            Q(student_name__icontains=q) |
            Q(email__icontains=q) |
            Q(mobile__icontains=q)
        )

    department_id = request.GET.get('department')
    if department_id:
        students_qs = students_qs.filter(student_courses__department_id=department_id).distinct()

    course_id = request.GET.get('course')
    if course_id:
        students_qs = students_qs.filter(student_courses__course_id=course_id).distinct()

    batch_id = request.GET.get('batch')
    if batch_id:
        students_qs = students_qs.filter(student_courses__batch_id=batch_id).distinct()

    status = request.GET.get('status')
    if status:
        students_qs = students_qs.filter(status=status)

    enrolled_from = request.GET.get('enrolled_from')
    if enrolled_from:
        students_qs = students_qs.filter(enrolled_date__gte=enrolled_from)

    enrolled_to = request.GET.get('enrolled_to')
    if enrolled_to:
        students_qs = students_qs.filter(enrolled_date__lte=enrolled_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(students_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Calculate stats
    total_count = students_qs.count()
    active_count = students_qs.filter(status='active').count()
    completed_count = students_qs.filter(status='completed').count()
    with_batch_count = students_qs.filter(student_courses__batch__isnull=False).distinct().count()
    
    # Recent students (last 7 days)
    seven_days_ago = timezone.localdate() - timedelta(days=7)
    recent_count = students_qs.filter(enrolled_date__gte=seven_days_ago).count()
    
    # Today's enrollments
    today_count = students_qs.filter(enrolled_date=timezone.localdate()).count()

    # Get latest 9 students
    latest_students = Student.objects.prefetch_related(
        'student_courses__course', 'student_courses__batch'
    ).order_by('-enrolled_date', '-id')[:9]

    # Get status distribution for chart
    status_distribution = students_qs.values('status').annotate(
        count=Count('id')
    ).order_by('-count')

    # Get department distribution with course and batch details
    from .models import StudentCourse
    
    department_distribution = []
    for dept in Department.objects.all():
        courses_count = Course.objects.filter(department=dept).count()
        total_batches = Batch.objects.filter(department=dept).count()
        
        if courses_count > 0:
            department_distribution.append({
                'department_name': dept.department_name,
                'courses_count': courses_count,
                'total_batches': total_batches
            })

    context = {
        'students': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_count': total_count,
        'active_count': active_count,
        'completed_count': completed_count,
        'with_batch_count': with_batch_count,
        'recent_count': recent_count,
        'today_count': today_count,
        'latest_students': latest_students,
        'status_distribution': status_distribution,
        'department_distribution': department_distribution,
        'departments': Department.objects.all(),
        'courses': Course.objects.select_related('department'),
        'batches': Batch.objects.select_related('course').order_by('batch_name'),
        'status_choices': Student.STATUS_CHOICES,
    }
    return render(request, 'reports/student_report.html', context)

@role_required(['admin', 'management'])
def course_report(request):
    from django.db.models import Count
    from .models import Course

    courses_qs = Course.objects.select_related('department').annotate(
        total_enquiries=Count('leadcourse', distinct=True),
        total_students=Count('studentcourse__student', distinct=True),
        total_batches=Count('batch', distinct=True),
        active_batches=Count('batch', filter=Q(batch__status='Active'), distinct=True),
    ).order_by('department__department_name', 'course_name')

    q = request.GET.get('q', '').strip()
    if q:
        courses_qs = courses_qs.filter(course_name__icontains=q)

    department_id = request.GET.get('department')
    if department_id:
        courses_qs = courses_qs.filter(department_id=department_id)

    duration_from = request.GET.get('duration_from')
    if duration_from:
        courses_qs = courses_qs.filter(duration_months__gte=duration_from)

    duration_to = request.GET.get('duration_to')
    if duration_to:
        courses_qs = courses_qs.filter(duration_months__lte=duration_to)

    fee_from = request.GET.get('fee_from')
    if fee_from:
        courses_qs = courses_qs.filter(fees__gte=fee_from)

    fee_to = request.GET.get('fee_to')
    if fee_to:
        courses_qs = courses_qs.filter(fees__lte=fee_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(courses_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    total_batches = Batch.objects.count()
    
    # Use the full filtered queryset for graphs, not the paginated one
    department_distribution = courses_qs.values('department__department_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    duration_distribution = courses_qs.values('duration_months').annotate(
        count=Count('id')
    ).order_by('duration_months')
    
    context = {
        'courses': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_count': courses_qs.count(),
        'total_batches': total_batches,
        'departments': Department.objects.all(),
        'department_distribution': department_distribution,
        'duration_distribution': duration_distribution,
    }
    return render(request, 'reports/course_report.html', context)


@role_required(['admin', 'management'])
def batch_report(request):
    from django.db.models import Count
    from .models import Batch

    # Base queryset for all batches
    all_batches = Batch.objects.select_related(
        'department', 'course', 'trainer__user'
    )

    batches_qs = all_batches.order_by('-created_at', '-id')

    q = request.GET.get('q', '').strip()
    if q:
        batches_qs = batches_qs.filter(
            Q(batch_name__icontains=q) |
            Q(course__course_name__icontains=q) |
            Q(department__department_name__icontains=q)
        )

    department_id = request.GET.get('department')
    if department_id:
        batches_qs = batches_qs.filter(department_id=department_id)

    course_id = request.GET.get('course')
    if course_id:
        batches_qs = batches_qs.filter(course_id=course_id)

    trainer_id = request.GET.get('trainer')
    if trainer_id:
        batches_qs = batches_qs.filter(trainer_id=trainer_id)

    status = request.GET.get('status')
    if status:
        batches_qs = batches_qs.filter(status=status)

    mode = request.GET.get('mode')
    if mode:
        batches_qs = batches_qs.filter(mode=mode)

    start_from = request.GET.get('start_from')
    if start_from:
        batches_qs = batches_qs.filter(start_date__gte=start_from)

    start_to = request.GET.get('start_to')
    if start_to:
        batches_qs = batches_qs.filter(start_date__lte=start_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(batches_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    # Add enrolled_count and seats_left to each batch in page
    for batch in page_obj:
        batch.enrolled_count = batch.studentcourse_set.count()
        batch.seats_left_count = max(batch.no_of_students - batch.enrolled_count, 0)

    # Calculate card metrics from all batches (not filtered)
    total_students = sum(batch.studentcourse_set.count() for batch in all_batches)
    total_capacity = sum(batch.no_of_students for batch in all_batches)
    available_seats = total_capacity - total_students
    capacity_rate = (total_students / total_capacity * 100) if total_capacity > 0 else 0

    # Status distribution for chart (from all batches)
    status_distribution = all_batches.values('status').annotate(
        count=Count('id')
    ).order_by('-count')

    # Capacity utilization for chart (top 10 batches)
    capacity_utilization = []
    top_batches = list(batches_qs[:10])
    for batch in top_batches:
        enrolled = batch.studentcourse_set.count()
        utilization_rate = (enrolled / batch.no_of_students * 100) if batch.no_of_students > 0 else 0
        capacity_utilization.append({
            'batch_name': batch.batch_name,
            'enrolled': enrolled,
            'capacity': batch.no_of_students,
            'rate': round(utilization_rate, 1)
        })

    context = {
        'batches': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_count': all_batches.count(),
        'active_count': all_batches.filter(status='Active').count(),
        'completed_count': all_batches.filter(status='Completed').count(),
        'total_students': total_students,
        'available_seats': available_seats,
        'total_capacity': total_capacity,
        'capacity_rate': capacity_rate,
        'status_distribution': status_distribution,
        'capacity_utilization': capacity_utilization,
        'departments': Department.objects.all(),
        'courses': Course.objects.select_related('department'),
        'trainers': UserProfile.objects.select_related('user').filter(role='trainer'),
        'status_choices': Batch.STATUS_CHOICES,
        'mode_choices': Batch.MODE_CHOICES,
    }
    return render(request, 'reports/batch_report.html', context)


@role_required(['admin', 'management'])
def assignment_report(request):
    from django.db.models import Count, Sum
    from .models import Assignment, Batch
    from datetime import date

    assignments_qs = Assignment.objects.select_related(
        'batch', 'session', 'created_by__user'
    ).annotate(
        total_submissions=Count('submissions', distinct=True),
        submitted_submissions=Count(
            'submissions',
            filter=Q(submissions__status='submitted'),
            distinct=True
        )
    ).order_by('-due_date', '-created_at')

    q = request.GET.get('q', '').strip()
    if q:
        assignments_qs = assignments_qs.filter(title__icontains=q)

    batch_id = request.GET.get('batch')
    if batch_id:
        assignments_qs = assignments_qs.filter(batch_id=batch_id)

    due_from = request.GET.get('due_from')
    if due_from:
        assignments_qs = assignments_qs.filter(due_date__gte=due_from)

    due_to = request.GET.get('due_to')
    if due_to:
        assignments_qs = assignments_qs.filter(due_date__lte=due_to)

    created_by = request.GET.get('created_by')
    if created_by:
        assignments_qs = assignments_qs.filter(created_by_id=created_by)

    submission_state = request.GET.get('submission_state')
    if submission_state == 'completed':
        assignments_qs = assignments_qs.filter(submitted_submissions=F('total_submissions'))
    elif submission_state == 'pending':
        assignments_qs = assignments_qs.filter(submitted_submissions=0)
    elif submission_state == 'partial':
        assignments_qs = assignments_qs.filter(
            submitted_submissions__gt=0,
            submitted_submissions__lt=F('total_submissions')
        )

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(assignments_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    for assignment in page_obj:
        assignment.pending_submissions = max(
            assignment.total_submissions - assignment.submitted_submissions, 0
        )

    # Calculate metrics
    total_submissions_count = assignments_qs.aggregate(total=Sum('total_submissions'))['total'] or 0
    submitted_count = assignments_qs.aggregate(total=Sum('submitted_submissions'))['total'] or 0
    pending_count = total_submissions_count - submitted_count
    overdue_count = assignments_qs.filter(due_date__lt=date.today()).aggregate(
        total=Sum('total_submissions')
    )['total'] or 0
    completion_rate = (submitted_count / total_submissions_count * 100) if total_submissions_count > 0 else 0
    
    # Submission status distribution
    submission_status_distribution = [
        {'status': 'Submitted', 'count': submitted_count},
        {'status': 'Pending', 'count': pending_count},
    ]
    
    # Batch completion rate
    batch_completion_rate = []
    for assignment in assignments_qs:
        if assignment.batch:
            batch_name = assignment.batch.batch_name
            total = assignment.total_submissions
            submitted = assignment.submitted_submissions
            
            # Find or create batch entry
            batch_entry = next((b for b in batch_completion_rate if b['batch_name'] == batch_name), None)
            if batch_entry:
                batch_entry['total'] += total
                batch_entry['submitted'] += submitted
            else:
                batch_completion_rate.append({
                    'batch_name': batch_name,
                    'total': total,
                    'submitted': submitted
                })
    
    # Calculate completion rates
    for batch in batch_completion_rate:
        batch['completion_rate'] = round((batch['submitted'] / batch['total'] * 100), 1) if batch['total'] > 0 else 0
    
    batch_completion_rate = sorted(batch_completion_rate, key=lambda x: x['completion_rate'], reverse=True)[:10]

    context = {
        'assignments': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_count': assignments_qs.count(),
        'total_submissions': total_submissions_count,
        'submitted_count': submitted_count,
        'pending_count': pending_count,
        'overdue_count': overdue_count,
        'completion_rate': completion_rate,
        'submission_status_distribution': submission_status_distribution,
        'batch_completion_rate': batch_completion_rate,
        'batches': Batch.objects.order_by('batch_name'),
        'assignment_creators': UserProfile.objects.select_related('user').filter(
            id__in=Assignment.objects.exclude(created_by__isnull=True).values_list('created_by_id', flat=True).distinct()
        ).order_by('user__first_name', 'user__username'),
        'can_filter_created_by': request.user.userprofile.role in ['admin', 'management'],
    }
    return render(request, 'reports/assignment_report.html', context)

def _get_filtered_payment_report_querysets(request):
    from django.db.models import Exists

    payments_qs = Payment.objects.select_related(
        'fee',
        'fee__student',
        'fee__course',
        'fee__course__department',
    ).prefetch_related(
        'fee__student__student_courses__batch',
        'fee__student__student_courses__course',
    ).order_by('-payment_date', '-id')

    fees_qs = Fee.objects.select_related(
        'student',
        'course',
        'course__department',
     ).prefetch_related(
         'student__student_courses__batch',
         'student__student_courses__course',
     ).order_by('student__student_name', 'course__course_name')

    q = request.GET.get('q', '').strip()
    if q:
        student_filter = (
            Q(student__student_name__icontains=q) |
            Q(student__mobile__icontains=q)
        )
        payments_qs = payments_qs.filter(
            Q(fee__student__student_name__icontains=q) |
            Q(fee__student__mobile__icontains=q)
        )
        fees_qs = fees_qs.filter(student_filter)

    dept_id = request.GET.get('department')
    if dept_id:
        payments_qs = payments_qs.filter(fee__course__department_id=dept_id)
        fees_qs = fees_qs.filter(course__department_id=dept_id)

    course_id = request.GET.get('course')
    if course_id:
        payments_qs = payments_qs.filter(fee__course_id=course_id)
        fees_qs = fees_qs.filter(course_id=course_id)

    batch_id = request.GET.get('batch')
    if batch_id:
        payment_sc_match = StudentCourse.objects.filter(
            student=OuterRef('fee__student'),
            batch_id=batch_id,
            course=OuterRef('fee__course')
        )
        fee_sc_match = StudentCourse.objects.filter(
            student=OuterRef('student'),
            batch_id=batch_id,
            course=OuterRef('course')
        )
        payments_qs = payments_qs.filter(Exists(payment_sc_match))
        fees_qs = fees_qs.filter(Exists(fee_sc_match))

    fee_status = request.GET.get('fee_status')
    if fee_status:
        payments_qs = payments_qs.filter(fee__fee_status=fee_status)
        fees_qs = fees_qs.filter(fee_status=fee_status)

    payment_mode = request.GET.get('payment_mode')
    if payment_mode:
        payments_qs = payments_qs.filter(payment_mode=payment_mode)

    remarks_state = request.GET.get('remarks_state')
    if remarks_state == 'with':
        payments_qs = payments_qs.exclude(Q(remarks__isnull=True) | Q(remarks=''))
    elif remarks_state == 'without':
        payments_qs = payments_qs.filter(Q(remarks__isnull=True) | Q(remarks=''))

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from and not date_to:
        date_to = timezone.now().date().isoformat()
    if date_from:
        payments_qs = payments_qs.filter(payment_date__gte=date_from)
    if date_to:
        payments_qs = payments_qs.filter(payment_date__lte=date_to)

    amount_from = request.GET.get('amount_from')
    if amount_from:
        payments_qs = payments_qs.filter(amount_paid__gte=amount_from)

    amount_to = request.GET.get('amount_to')
    if amount_to:
        payments_qs = payments_qs.filter(amount_paid__lte=amount_to)

    payment_filters_active = any([
        bool(payment_mode),
        bool(remarks_state),
        bool(date_from),
        bool(date_to),
        bool(amount_from),
        bool(amount_to),
    ])
    if payment_filters_active:
        filtered_fee_ids = payments_qs.values_list('fee_id', flat=True).distinct()
        fees_qs = fees_qs.filter(id__in=filtered_fee_ids)

    return payments_qs.distinct(), fees_qs.distinct()


@role_required(['admin', 'management'])
def payment_report(request):
    from django.db.models import Sum, Count, Max, Q
    from django.db.models.functions import TruncMonth

    payments_qs, fees_qs = _get_filtered_payment_report_querysets(request)

    payment_mode = request.GET.get('payment_mode')
    remarks_state = request.GET.get('remarks_state')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from and not date_to:
        date_to = timezone.now().date().isoformat()
    amount_from = request.GET.get('amount_from')
    amount_to = request.GET.get('amount_to')

    payment_filters_active = any([
        bool(payment_mode),
        bool(remarks_state),
        bool(date_from),
        bool(date_to),
        bool(amount_from),
        bool(amount_to),
    ])

    fee_payment_filter = Q()
    if payment_mode:
        fee_payment_filter &= Q(payments__payment_mode=payment_mode)
    if remarks_state == 'with':
        fee_payment_filter &= ~Q(payments__remarks__isnull=True) & ~Q(payments__remarks='')
    elif remarks_state == 'without':
        fee_payment_filter &= Q(payments__remarks__isnull=True) | Q(payments__remarks='')
    if date_from:
        fee_payment_filter &= Q(payments__payment_date__gte=date_from)
    if date_to:
        fee_payment_filter &= Q(payments__payment_date__lte=date_to)
    if amount_from:
        fee_payment_filter &= Q(payments__amount_paid__gte=amount_from)
    if amount_to:
        fee_payment_filter &= Q(payments__amount_paid__lte=amount_to)

    if payment_filters_active:
        fees_qs = fees_qs.annotate(
            last_payment_date=Max('payments__payment_date', filter=fee_payment_filter),
            transactions=Count('payments', filter=fee_payment_filter),
        )
    else:
        fees_qs = fees_qs.annotate(
            last_payment_date=Max('payments__payment_date'),
            transactions=Count('payments'),
        )

    # Show newest fee activity first in table.
    fees_qs = fees_qs.order_by('-last_payment_date', '-id')

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(fees_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    monthly_collected = payments_qs.annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        collected=Sum('amount_paid'),
        transactions=Count('id')
    ).order_by('month')[:12]

    context = {
        'fees': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_count': fees_qs.count(),
        'total_amount': payments_qs.aggregate(total=Sum('amount_paid')).get('total') or 0,
        'total_pending': fees_qs.aggregate(total=Sum(F('total_fee') - F('paid_amount'))).get('total') or 0,
        'paid_count': fees_qs.filter(fee_status='paid').count(),
        'partial_count': fees_qs.filter(fee_status='partial').count(),
        'pending_count': fees_qs.filter(fee_status='pending').count(),
        'overdue_count': fees_qs.filter(fee_status='overdue').count(),
        'monthly_collected': monthly_collected,
        'departments': Department.objects.order_by('department_name'),
        'courses': Course.objects.select_related('department').order_by('course_name'),
        'batches': Batch.objects.select_related('course').order_by('batch_name'),
        'payment_modes': Payment.PAYMENT_MODE_CHOICES,
        'fee_status_choices': Fee.FEE_STATUS_CHOICES,
    }
    return render(request, 'reports/payment_report.html', context)


def _build_excel_response(filename, sheet_title, headers, rows, top_rows=None):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    top_rows = top_rows or []
    for top_row in top_rows:
        sheet.append(top_row)

    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    header_row_index = len(top_rows) + 1
    for col_idx, _ in enumerate(headers, 1):
        header_cell = sheet.cell(row=header_row_index, column=col_idx)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, _ in enumerate(headers, 1):
        max_length = 0
        for cell in sheet[get_column_letter(col_idx)]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@role_required(['admin', 'management'])
def lead_report_export(request):
    from .models import StudentEnquiry

    leads_qs = StudentEnquiry.objects.select_related(
        'status', 'source', 'assigned__user', 'created_by'
    ).prefetch_related(
        'lead_courses__department',
        'lead_courses__course'
    ).order_by('-enquiry_date', '-id')

    q = request.GET.get('q', '').strip()
    if q:
        leads_qs = leads_qs.filter(
            Q(full_name__icontains=q) |
            Q(email__icontains=q) |
            Q(mobile__icontains=q)
        )

    date_from = request.GET.get('date_from')
    if date_from:
        leads_qs = leads_qs.filter(enquiry_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        leads_qs = leads_qs.filter(enquiry_date__lte=date_to)

    status_id = request.GET.get('status')
    if status_id:
        leads_qs = leads_qs.filter(status_id=status_id)

    department_id = request.GET.get('department')
    if department_id:
        leads_qs = leads_qs.filter(lead_courses__department_id=department_id).distinct()

    course_id = request.GET.get('course')
    if course_id:
        leads_qs = leads_qs.filter(lead_courses__course_id=course_id).distinct()

    assigned_to = request.GET.get('assigned_to')
    if assigned_to:
        leads_qs = leads_qs.filter(assigned_id=assigned_to)

    source = request.GET.get('source')
    if source:
        leads_qs = leads_qs.filter(source_id=source)

    mode = request.GET.get('mode')
    if mode:
        leads_qs = leads_qs.filter(mode=mode)

    created_by_id = request.GET.get('created_by')
    if created_by_id:
        leads_qs = leads_qs.filter(created_by_id=created_by_id)

    followup_state = request.GET.get('followup_state')
    today = timezone.localdate()
    if followup_state == 'with':
        leads_qs = leads_qs.filter(followup_date__isnull=False)
    elif followup_state == 'without':
        leads_qs = leads_qs.filter(followup_date__isnull=True)
    elif followup_state == 'today':
        leads_qs = leads_qs.filter(followup_date=today)
    elif followup_state == 'overdue':
        leads_qs = leads_qs.filter(followup_date__lt=today)
    elif followup_state == 'upcoming':
        leads_qs = leads_qs.filter(followup_date__gt=today)

    followup_from = request.GET.get('followup_from')
    if followup_from:
        leads_qs = leads_qs.filter(followup_date__gte=followup_from)

    followup_to = request.GET.get('followup_to')
    if followup_to:
        leads_qs = leads_qs.filter(followup_date__lte=followup_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    page_number = request.GET.get('page')
    if page_number:
        paginator = Paginator(leads_qs, per_page)
        leads_qs = paginator.get_page(page_number)

    column_definitions = {
        "name": (
            "Name",
            lambda lead: lead.full_name,
        ),
        "contact": (
            "Contact",
            lambda lead: f"{lead.mobile}{' | ' + lead.email if lead.email else ''}",
        ),
        "enquiry_date": (
            "Enquiry Date",
            lambda lead: lead.enquiry_date.strftime('%Y-%m-%d') if lead.enquiry_date else '',
        ),
        "dob": (
            "DOB",
            lambda lead: lead.dob.strftime('%Y-%m-%d') if lead.dob else '',
        ),
        "guardian_number": (
            "Guardian Number",
            lambda lead: lead.guardian_number or '',
        ),
        "year_of_passing": (
            "Year of Passing",
            lambda lead: str(lead.year_of_passing or ''),
        ),
        "department": (
            "Department",
            lambda lead: ', '.join([lc.department.department_name for lc in lead.lead_courses.all()]) if lead.lead_courses.exists() else '',
        ),
        "course": (
            "Course",
            lambda lead: ', '.join([lc.course.course_name for lc in lead.lead_courses.all()]) if lead.lead_courses.exists() else '',
        ),
        "status": (
            "Status",
            lambda lead: lead.status.status_name if lead.status else '',
        ),
        "source": (
            "Source",
            lambda lead: lead.source.source_name if lead.source else '',
        ),
        "assigned_to": (
            "Assigned To",
            lambda lead: lead.assigned.user.get_full_name() if lead.assigned and lead.assigned.user.get_full_name() else (
                lead.assigned.user.username if lead.assigned else ''
            ),
        ),
        "followup": (
            "Follow-up",
            lambda lead: lead.followup_date.strftime('%Y-%m-%d') if lead.followup_date else '',
        ),
    }

    requested_columns = request.GET.getlist('columns')
    selected_columns = [col for col in requested_columns if col in column_definitions]
    if not selected_columns:
        selected_columns = list(column_definitions.keys())

    headers = [column_definitions[col][0] for col in selected_columns]

    rows = []
    for lead in leads_qs:
        row = [column_definitions[col][1](lead) for col in selected_columns]
        rows.append(row)

    return _build_excel_response(
        "lead_report.xlsx",
        "Lead Report",
        headers,
        rows,
    )

@role_required(['admin', 'management'])
def student_report_export(request):
    from .models import Student

    students_qs = Student.objects.prefetch_related(
        'student_courses__course', 'student_courses__department'
    ).order_by('-enrolled_date', '-id')

    q = request.GET.get('q', '').strip()
    if q:
        students_qs = students_qs.filter(
            Q(student_name__icontains=q) |
            Q(email__icontains=q) |
            Q(mobile__icontains=q)
        )

    department_id = request.GET.get('department')
    if department_id:
        students_qs = students_qs.filter(student_courses__department_id=department_id).distinct()

    course_id = request.GET.get('course')
    if course_id:
        students_qs = students_qs.filter(student_courses__course_id=course_id).distinct()

    batch_id = request.GET.get('batch')
    if batch_id:
        students_qs = students_qs.filter(batch_id=batch_id)

    status = request.GET.get('status')
    if status:
        students_qs = students_qs.filter(status=status)

    batch_assignment = request.GET.get('batch_assignment')
    if batch_assignment == 'assigned':
        students_qs = students_qs.filter(batch__isnull=False)
    elif batch_assignment == 'unassigned':
        students_qs = students_qs.filter(batch__isnull=True)

    email_state = request.GET.get('email_state')
    if email_state == 'with':
        students_qs = students_qs.exclude(Q(email__isnull=True) | Q(email=''))
    elif email_state == 'without':
        students_qs = students_qs.filter(Q(email__isnull=True) | Q(email=''))

    enrollment_state = request.GET.get('enrollment_state')
    if enrollment_state == 'with_date':
        students_qs = students_qs.filter(enrolled_date__isnull=False)
    elif enrollment_state == 'without_date':
        students_qs = students_qs.filter(enrolled_date__isnull=True)

    enrolled_from = request.GET.get('enrolled_from')
    if enrolled_from:
        students_qs = students_qs.filter(enrolled_date__gte=enrolled_from)

    enrolled_to = request.GET.get('enrolled_to')
    if enrolled_to:
        students_qs = students_qs.filter(enrolled_date__lte=enrolled_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    page_number = request.GET.get('page')
    if page_number:
        paginator = Paginator(students_qs, per_page)
        students_qs = paginator.get_page(page_number)

    column_definitions = {
        "student_id": (
            "Student ID",
            lambda student: str(student.id),
        ),
        "name": (
            "Name",
            lambda student: student.student_name,
        ),
        "contact": (
            "Contact",
            lambda student: f"{student.mobile}{' | ' + student.email if student.email else ''}",
        ),
        "enrolled_date": (
            "Enrolled Date",
            lambda student: student.enrolled_date.strftime('%Y-%m-%d') if student.enrolled_date else '',
        ),
        "department": (
            "Department",
            lambda student: ', '.join([sc.department.department_name for sc in student.student_courses.all()]) if student.student_courses.exists() else '',
        ),
        "course": (
            "Course",
            lambda student: ', '.join([sc.course.course_name for sc in student.student_courses.all()]) if student.student_courses.exists() else '',
        ),
        "batch": (
            "Batch",
            lambda student: ', '.join([sc.batch.batch_name for sc in student.student_courses.all() if sc.batch]) if student.student_courses.exists() else '',
        ),
        "status": (
            "Status",
            lambda student: student.get_status_display(),
        ),
    }

    requested_columns = request.GET.getlist('columns')
    selected_columns = [col for col in requested_columns if col in column_definitions]
    if not selected_columns:
        selected_columns = list(column_definitions.keys())

    headers = [column_definitions[col][0] for col in selected_columns]

    rows = []
    for student in students_qs:
        row = [column_definitions[col][1](student) for col in selected_columns]
        rows.append(row)

    return _build_excel_response(
        "student_report.xlsx",
        "Student Report",
        headers,
        rows,
    )


@role_required(['admin', 'management'])
def course_report_export(request):
    from django.db.models import Count
    from .models import Course

    courses_qs = Course.objects.select_related('department').annotate(
        total_students=Count('studentcourse__student', distinct=True),
        total_batches=Count('batch', distinct=True),
        active_batches=Count('batch', filter=Q(batch__status='Active'), distinct=True),
        total_enquiries=Count('leadcourse', distinct=True),
    ).order_by('department__department_name', 'course_name')

    q = request.GET.get('q', '').strip()
    if q:
        courses_qs = courses_qs.filter(course_name__icontains=q)

    department_id = request.GET.get('department')
    if department_id:
        courses_qs = courses_qs.filter(department_id=department_id)

    duration_from = request.GET.get('duration_from')
    if duration_from:
        courses_qs = courses_qs.filter(duration_months__gte=duration_from)

    duration_to = request.GET.get('duration_to')
    if duration_to:
        courses_qs = courses_qs.filter(duration_months__lte=duration_to)

    fee_from = request.GET.get('fee_from')
    if fee_from:
        courses_qs = courses_qs.filter(fees__gte=fee_from)

    fee_to = request.GET.get('fee_to')
    if fee_to:
        courses_qs = courses_qs.filter(fees__lte=fee_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    page_number = request.GET.get('page')
    if page_number:
        paginator = Paginator(courses_qs, per_page)
        courses_qs = paginator.get_page(page_number)

    column_definitions = {
        "department": ("Department", lambda course: course.department.department_name if course.department else ''),
        "course": ("Course", lambda course: course.course_name),
        "fees": ("Fees", lambda course: str(course.fees) if course.fees is not None else ''),
        "duration": ("Duration", lambda course: f"{course.duration_months} months"),
        "enquiries": ("Enquiries", lambda course: course.total_enquiries),
        "students": ("Students", lambda course: course.total_students),
        "total_batches": ("Total Batches", lambda course: course.total_batches),
        "active_batches": ("Active Batches", lambda course: course.active_batches),
    }

    requested_columns = request.GET.getlist('columns')
    selected_columns = [col for col in requested_columns if col in column_definitions]
    if not selected_columns:
        selected_columns = list(column_definitions.keys())

    headers = [column_definitions[col][0] for col in selected_columns]

    rows = []
    for course in courses_qs:
        row = [column_definitions[col][1](course) for col in selected_columns]
        rows.append(row)

    return _build_excel_response(
        "course_report.xlsx",
        "Course Report",
        headers,
        rows,
    )


@role_required(['admin', 'management'])
def batch_report_export(request):
    from django.db.models import Count
    from .models import Batch

    batches_qs = Batch.objects.select_related(
        'department', 'course', 'trainer__user'
    ).annotate(
        enrolled_count=Count('studentcourse', distinct=True)
    ).order_by('-created_at', '-id')

    q = request.GET.get('q', '').strip()
    if q:
        batches_qs = batches_qs.filter(
            Q(batch_name__icontains=q) |
            Q(course__course_name__icontains=q) |
            Q(department__department_name__icontains=q)
        )

    department_id = request.GET.get('department')
    if department_id:
        batches_qs = batches_qs.filter(department_id=department_id)

    course_id = request.GET.get('course')
    if course_id:
        batches_qs = batches_qs.filter(course_id=course_id)

    trainer_id = request.GET.get('trainer')
    if trainer_id:
        batches_qs = batches_qs.filter(trainer_id=trainer_id)

    status = request.GET.get('status')
    if status:
        batches_qs = batches_qs.filter(status=status)

    mode = request.GET.get('mode')
    if mode:
        batches_qs = batches_qs.filter(mode=mode)

    start_from = request.GET.get('start_from')
    if start_from:
        batches_qs = batches_qs.filter(start_date__gte=start_from)

    start_to = request.GET.get('start_to')
    if start_to:
        batches_qs = batches_qs.filter(start_date__lte=start_to)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    page_number = request.GET.get('page')
    if page_number:
        paginator = Paginator(batches_qs, per_page)
        batches_qs = paginator.get_page(page_number)

    column_definitions = {
        "batch": ("Batch", lambda batch: batch.batch_name),
        "department": ("Department", lambda batch: batch.department.department_name if batch.department else ''),
        "course": ("Course", lambda batch: batch.course.course_name if batch.course else ''),
        "trainer": ("Trainer", lambda batch: batch.trainer.user.get_full_name() if batch.trainer and batch.trainer.user.get_full_name() else (batch.trainer.user.username if batch.trainer else '')),
        "duration": ("Duration", lambda batch: f"{batch.start_date.strftime('%Y-%m-%d') if batch.start_date else ''} to {batch.end_date.strftime('%Y-%m-%d') if batch.end_date else ''}".strip()),
        "mode": ("Mode", lambda batch: batch.mode),
        "status": ("Status", lambda batch: batch.status),
        "capacity": ("Capacity", lambda batch: batch.no_of_students),
        "enrolled": ("Enrolled", lambda batch: batch.enrolled_count),
        "seats_left": ("Seats Left", lambda batch: max(batch.no_of_students - batch.enrolled_count, 0)),
    }

    requested_columns = request.GET.getlist('columns')
    selected_columns = [col for col in requested_columns if col in column_definitions]
    if not selected_columns:
        selected_columns = list(column_definitions.keys())

    headers = [column_definitions[col][0] for col in selected_columns]

    rows = []
    for batch in batches_qs:
        row = [column_definitions[col][1](batch) for col in selected_columns]
        rows.append(row)

    return _build_excel_response(
        "batch_report.xlsx",
        "Batch Report",
        headers,
        rows,
    )


@role_required(['admin', 'management'])
def assignment_report_export(request):
    from django.db.models import Count
    from .models import Assignment

    assignments_qs = Assignment.objects.select_related(
        'batch', 'session', 'created_by__user'
    ).annotate(
        total_submissions=Count('submissions', distinct=True),
        submitted_submissions=Count(
            'submissions',
            filter=Q(submissions__status='submitted'),
            distinct=True
        )
    ).order_by('-due_date', '-created_at')

    q = request.GET.get('q', '').strip()
    if q:
        assignments_qs = assignments_qs.filter(title__icontains=q)

    batch_id = request.GET.get('batch')
    if batch_id:
        assignments_qs = assignments_qs.filter(batch_id=batch_id)

    due_from = request.GET.get('due_from')
    if due_from:
        assignments_qs = assignments_qs.filter(due_date__gte=due_from)

    due_to = request.GET.get('due_to')
    if due_to:
        assignments_qs = assignments_qs.filter(due_date__lte=due_to)

    created_by = request.GET.get('created_by')
    if created_by:
        assignments_qs = assignments_qs.filter(created_by_id=created_by)

    submission_state = request.GET.get('submission_state')
    if submission_state == 'completed':
        assignments_qs = assignments_qs.filter(submitted_submissions=F('total_submissions'))
    elif submission_state == 'pending':
        assignments_qs = assignments_qs.filter(submitted_submissions=0)
    elif submission_state == 'partial':
        assignments_qs = assignments_qs.filter(
            submitted_submissions__gt=0,
            submitted_submissions__lt=F('total_submissions')
        )

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    page_number = request.GET.get('page')
    if page_number:
        paginator = Paginator(assignments_qs, per_page)
        assignments_qs = paginator.get_page(page_number)

    column_definitions = {
        "assignment": ("Assignment", lambda assignment: assignment.title),
        "batch": ("Batch", lambda assignment: assignment.batch.batch_name if assignment.batch else ''),
        "session_date": ("Session Date", lambda assignment: assignment.session.session_date.strftime('%Y-%m-%d') if assignment.session else ''),
        "due_date": ("Due Date", lambda assignment: assignment.due_date.strftime('%Y-%m-%d') if assignment.due_date else ''),
        "created_by": ("Created By", lambda assignment: assignment.created_by.user.get_full_name() if assignment.created_by and assignment.created_by.user.get_full_name() else (assignment.created_by.user.username if assignment.created_by else '')),
        "total": ("Total", lambda assignment: assignment.total_submissions),
        "submitted": ("Submitted", lambda assignment: assignment.submitted_submissions),
        "pending": ("Pending", lambda assignment: max(assignment.total_submissions - assignment.submitted_submissions, 0)),
    }

    requested_columns = request.GET.getlist('columns')
    selected_columns = [col for col in requested_columns if col in column_definitions]
    if not selected_columns:
        selected_columns = list(column_definitions.keys())

    headers = [column_definitions[col][0] for col in selected_columns]

    rows = []
    for assignment in assignments_qs:
        row = [column_definitions[col][1](assignment) for col in selected_columns]
        rows.append(row)

    return _build_excel_response(
        "assignment_report.xlsx",
        "Assignment Report",
        headers,
        rows,
    )

@login_required
@role_required(['admin', 'management'])
def exam_report(request):
    from django.db.models import Count, Avg, Q
    from .models import Exam, Batch
    from datetime import date

    exams_qs = Exam.objects.select_related('batch', 'created_by__user').annotate(
        total_students=Count('batch__studentcourse', distinct=True),
        attempted_students=Count('performances', distinct=True),
        average_score=Avg('performances__score')
    ).order_by('-exam_date', '-created_at')

    q = request.GET.get('q', '').strip()
    if q:
        exams_qs = exams_qs.filter(title__icontains=q)

    batch_id = request.GET.get('batch')
    if batch_id:
        exams_qs = exams_qs.filter(batch_id=batch_id)

    exam_from = request.GET.get('exam_from')
    if exam_from:
        exams_qs = exams_qs.filter(exam_date__gte=exam_from)

    exam_to = request.GET.get('exam_to')
    if exam_to:
        exams_qs = exams_qs.filter(exam_date__lte=exam_to)

    status = request.GET.get('status')
    if status == 'completed':
        exams_qs = exams_qs.filter(exam_date__lt=date.today())
    elif status == 'pending':
        exams_qs = exams_qs.filter(exam_date=date.today())
    elif status == 'upcoming':
        exams_qs = exams_qs.filter(exam_date__gt=date.today())

    created_by = request.GET.get('created_by')
    if created_by:
        exams_qs = exams_qs.filter(created_by_id=created_by)

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    paginator = Paginator(exams_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    for exam in page_obj:
        exam.pending_students = max(exam.total_students - exam.attempted_students, 0)

    total_attempts = sum(e.attempted_students for e in exams_qs)
    completed_count = exams_qs.filter(exam_date__lt=date.today()).count()
    pending_count = exams_qs.filter(exam_date__gte=date.today()).count()
    average_score = exams_qs.aggregate(avg=Avg('average_score'))['avg'] or 0
    pass_rate = exams_qs.filter(average_score__gte=50).count() / exams_qs.count() * 100 if exams_qs.count() > 0 else 0

    status_distribution = [
        {'status': 'Completed', 'count': completed_count},
        {'status': 'Pending', 'count': pending_count},
    ]

    batch_performance = []
    for exam in exams_qs:
        if exam.batch and exam.average_score:
            batch_name = exam.batch.batch_name
            batch_entry = next((b for b in batch_performance if b['batch_name'] == batch_name), None)
            if batch_entry:
                batch_entry['total_score'] += exam.average_score
                batch_entry['count'] += 1
            else:
                batch_performance.append({
                    'batch_name': batch_name,
                    'total_score': exam.average_score,
                    'count': 1
                })

    for batch in batch_performance:
        batch['average_score'] = round(batch['total_score'] / batch['count'], 1) if batch['count'] > 0 else 0

    batch_performance = sorted(batch_performance, key=lambda x: x['average_score'], reverse=True)[:10]

    context = {
        'exams': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'total_count': exams_qs.count(),
        'total_attempts': total_attempts,
        'completed_count': completed_count,
        'pending_count': pending_count,
        'average_score': average_score,
        'pass_rate': pass_rate,
        'status_distribution': status_distribution,
        'batch_performance': batch_performance,
        'batches': Batch.objects.order_by('batch_name'),
        'exam_creators': UserProfile.objects.select_related('user').filter(
            id__in=Exam.objects.exclude(created_by__isnull=True).values_list('created_by_id', flat=True).distinct()
        ).order_by('user__first_name', 'user__username'),
        'can_filter_created_by': request.user.userprofile.role in ['admin', 'management'],
    }
    return render(request, 'reports/exam_report.html', context)


@login_required
@role_required(['admin', 'management'])
def exam_report_export(request):
    from django.db.models import Count, Avg
    from .models import Exam

    exams_qs = Exam.objects.select_related('batch', 'created_by__user').annotate(
        total_students=Count('batch__studentcourse', distinct=True),
        attempted_students=Count('performances', distinct=True),
        average_score=Avg('performances__score')
    ).order_by('-exam_date')

    requested_columns = request.GET.getlist('columns')
    column_definitions = {
        'exam': ('Exam', lambda e: e.title),
        'batch': ('Batch', lambda e: e.batch.batch_name if e.batch else '-'),
        'exam_date': ('Exam Date', lambda e: e.exam_date.strftime('%d %b %Y') if e.exam_date else '-'),
        'duration': ('Duration', lambda e: '-'),
        'total_marks': ('Total Marks', lambda e: e.max_marks or '-'),
        'created_by': ('Created By', lambda e: e.created_by.user.get_full_name() or e.created_by.user.username if e.created_by else '-'),
        'total_students': ('Total Students', lambda e: e.total_students),
        'attempted': ('Attempted', lambda e: e.attempted_students),
        'pending': ('Pending', lambda e: max(e.total_students - e.attempted_students, 0)),
        'avg_score': ('Avg Score', lambda e: f"{e.average_score:.1f}%" if e.average_score else '-'),
    }

    selected_columns = [col for col in requested_columns if col in column_definitions]
    if not selected_columns:
        selected_columns = list(column_definitions.keys())

    headers = [column_definitions[col][0] for col in selected_columns]
    rows = []
    for exam in exams_qs:
        row = [column_definitions[col][1](exam) for col in selected_columns]
        rows.append(row)

    return _build_excel_response("exam_report.xlsx", "Exam Report", headers, rows)

@role_required(['admin', 'management'])
def payment_report_export(request):
    from django.db.models import Max, Count, Q
    from decimal import Decimal

    payments_qs, fees_qs = _get_filtered_payment_report_querysets(request)

    payment_mode = request.GET.get('payment_mode')
    remarks_state = request.GET.get('remarks_state')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from and not date_to:
        from django.utils import timezone
        date_to = timezone.now().date().isoformat()
    amount_from = request.GET.get('amount_from')
    amount_to = request.GET.get('amount_to')

    payment_filters_active = any([
        bool(payment_mode),
        bool(remarks_state),
        bool(date_from),
        bool(date_to),
        bool(amount_from),
        bool(amount_to),
    ])

    fee_payment_filter = Q()
    if payment_mode:
        fee_payment_filter &= Q(payments__payment_mode=payment_mode)
    if remarks_state == 'with':
        fee_payment_filter &= ~Q(payments__remarks__isnull=True) & ~Q(payments__remarks='')
    elif remarks_state == 'without':
        fee_payment_filter &= Q(payments__remarks__isnull=True) | Q(payments__remarks='')
    if date_from:
        fee_payment_filter &= Q(payments__payment_date__gte=date_from)
    if date_to:
        fee_payment_filter &= Q(payments__payment_date__lte=date_to)
    if amount_from:
        fee_payment_filter &= Q(payments__amount_paid__gte=amount_from)
    if amount_to:
        fee_payment_filter &= Q(payments__amount_paid__lte=amount_to)

    if payment_filters_active:
        fees_qs = fees_qs.annotate(
            last_payment_date=Max('payments__payment_date', filter=fee_payment_filter),
            transactions=Count('payments', filter=fee_payment_filter),
        )
    else:
        fees_qs = fees_qs.annotate(
            last_payment_date=Max('payments__payment_date'),
            transactions=Count('payments'),
        )

    # Keep export order aligned with on-screen report: newest first.
    fees_qs = fees_qs.order_by('-last_payment_date', '-id')

    try:
        per_page = int(request.GET.get('per_page', 20))
    except (TypeError, ValueError):
        per_page = 20

    page_number = request.GET.get('page')
    if page_number:
        paginator = Paginator(fees_qs, per_page)
        fees_qs = paginator.get_page(page_number)

    column_definitions = {
        "student": (
            "Student",
            lambda fee: fee.student.student_name if fee.student else '',
        ),
        "mobile": (
            "Mobile",
            lambda fee: fee.student.mobile if fee.student else '',
        ),
        "course": (
            "Course",
            lambda fee: fee.course.course_name if fee.course else '',
        ),
        "batch": (
            "Batch",
            lambda fee: next(
                (sc.batch.batch_name for sc in fee.student.student_courses.all()
                 if sc.batch and sc.course == fee.course),
                ""
            ) if fee.student else "",
        ),
        "last_payment_date": (
            "Last Payment Date",
            lambda fee: fee.last_payment_date.strftime('%Y-%m-%d') if getattr(fee, 'last_payment_date', None) else '',
        ),
        "transactions": (
            "Transactions",
            lambda fee: str(getattr(fee, 'transactions', 0) or 0),
        ),
        "fee_status": (
            "Fee Status",
            lambda fee: fee.fee_status or '',
        ),
        "total_fee": (
            "Total Fee",
            lambda fee: str(fee.total_fee) if fee.total_fee is not None else '',
        ),
        "paid": (
            "Paid",
            lambda fee: str(fee.paid_amount) if fee.paid_amount is not None else '',
        ),
        "pending": (
            "Pending",
            lambda fee: str(fee.pending_amount) if fee else '',
        ),
        "due_date": (
            "Due Date",
            lambda fee: fee.due_date.strftime('%Y-%m-%d') if fee.due_date else '',
        ),
    }

    requested_columns = request.GET.getlist('columns')
    selected_columns = [col for col in requested_columns if col in column_definitions]
    if not selected_columns:
        selected_columns = list(column_definitions.keys())

    headers = [column_definitions[col][0] for col in selected_columns]

    rows = []
    total_fee_sum = Decimal('0.00')
    paid_sum = Decimal('0.00')
    pending_sum = Decimal('0.00')
    for fee in fees_qs:
        if getattr(fee, 'total_fee', None) is not None:
            total_fee_sum += fee.total_fee
        if getattr(fee, 'paid_amount', None) is not None:
            paid_sum += fee.paid_amount
        pending_sum += getattr(fee, 'pending_amount', Decimal('0.00')) or Decimal('0.00')

        row = [column_definitions[col][1](fee) for col in selected_columns]
        rows.append(row)

    return _build_excel_response(
        "payment_report.xlsx",
        "Payment Report",
        headers,
        rows,
        top_rows=[
            ["Total", str(total_fee_sum)],
            ["Paid", str(paid_sum)],
            ["Pending", str(pending_sum)],
            [],
        ],
    )



def student_live_search(request):
    query = request.GET.get('q', '')
    
    students = Student.objects.prefetch_related('student_courses__course', 'student_courses__batch')
    
    if query:
        students = students.filter(
            Q(student_name__icontains=query) |
            Q(email__icontains=query) |
            Q(mobile__icontains=query)
        )
    
    students = students.order_by('-id')[:50]
    
    data = []
    for student in students:
        courses = [
            f'<span class="badge bg-primary mb-1">{sc.course.course_name}</span>'
            for sc in student.student_courses.all()
        ]
        batch_names = [
            f'<span class="badge batch-badge">{sc.batch.batch_name}</span>'
            if sc.batch else
            '<span class="badge awaiting-batch-badge">Awaiting for batch</span>'
            for sc in student.student_courses.all()
        ]
        
        data.append({
            "id": student.id,
            "student_name": student.student_name,
            "email": student.email,
            "mobile": student.mobile,
            "course": '<br>'.join(courses) if courses else '<span class="text-muted">-</span>',
            "batch": '<br>'.join(batch_names) if batch_names else '<span class="badge awaiting-batch-badge">Awaiting for batch</span>',
            "enrolled_date": student.enrolled_date.strftime("%d %b %Y") if student.enrolled_date else "",
            "status": student.status,
        })
    
    return JsonResponse({"students": data})


@login_required
def get_trainer_notifications(request):
    """Get role-specific notifications: trainer, CRE, admin"""

    user_profile = request.user.userprofile
    role = user_profile.role

    today = timezone.localdate()
    tomorrow = today + timedelta(days=1)
    week_ahead = today + timedelta(days=7)

    notifications = []

    # === TRAINER NOTIFICATIONS ===
    if role == 'trainer':
        # Upcoming exams in trainer's batches
        my_batches = Batch.objects.filter(trainer=user_profile, status='Active')
        upcoming_exams = Exam.objects.filter(
            batch__in=my_batches,
            exam_date__gte=today,
            exam_date__lte=week_ahead
        ).select_related('batch').order_by('exam_date')[:5]
        
        for exam in upcoming_exams:
            is_today = exam.exam_date == today
            is_tomorrow = exam.exam_date == tomorrow
            
            notifications.append({
                'id': f'exam:{exam.id}',
                'type': 'exam',
                'title': f"{'Today' if is_today else 'Tomorrow' if is_tomorrow else 'Upcoming'} Exam",
                'message': f"{exam.title} - {exam.batch.batch_name}",
                'date': exam.exam_date.strftime('%d %b %Y'),
                'time': '',
                'priority': 'high' if is_today else 'medium' if is_tomorrow else 'low',
                'is_today': is_today,
                'url': f"/exams/{exam.id}/"
            })
        
        # Pending assignments (due soon)
        pending_assignments = Assignment.objects.filter(
            batch__in=my_batches,
            due_date__gte=today,
            due_date__lte=week_ahead
        ).select_related('batch').order_by('due_date')[:5]
        
        for assignment in pending_assignments:
            is_today = assignment.due_date == today
            is_tomorrow = assignment.due_date == tomorrow
            
            notifications.append({
                'id': f'assignment:{assignment.id}',
                'type': 'assignment',
                'title': f"Assignment Due {'Today' if is_today else 'Tomorrow' if is_tomorrow else 'Soon'}",
                'message': f"{assignment.title} - {assignment.batch.batch_name}",
                'date': assignment.due_date.strftime('%d %b %Y'),
                'time': '',
                'priority': 'high' if is_today else 'medium' if is_tomorrow else 'low',
                'is_today': is_today,
                'url': f"/assignments/{assignment.id}/"
            })

    # === CRE NOTIFICATIONS ===
    elif role == 'cre':
        # Pending followups
        pending_followups = FollowUp.objects.filter(
            assigned_to=user_profile,
            status='pending',
            followup_date__lte=week_ahead
        ).select_related('lead').order_by('followup_date', 'followup_time')
        
        for followup in pending_followups:
            is_overdue = followup.followup_date < today
            is_today = followup.followup_date == today
            
            notifications.append({
                'id': f'followup:{followup.id}',
                'type': 'followup',
                'title': f"{'Overdue' if is_overdue else 'Today' if is_today else 'Upcoming'} Follow-up",
                'message': f"{followup.title} - {followup.lead.full_name}",
                'date': followup.followup_date.strftime('%d %b %Y'),
                'time': followup.followup_time.strftime('%I:%M %p') if followup.followup_time else '',
                'priority': 'high' if is_overdue else 'medium' if is_today else 'low',
                'is_today': is_today,
                'url': reverse('lead_profile', args=[followup.lead.id])
            })

    # === ADMIN NOTIFICATIONS ===
    elif role == 'admin':
        # Today's pending followups only
        pending_followups = FollowUp.objects.filter(
            status='pending',
            followup_date=today
        ).select_related('lead', 'assigned_to__user').order_by('followup_time')[:20]
        
        for followup in pending_followups:
            assigned_name = followup.assigned_to.user.get_full_name() or followup.assigned_to.user.username if followup.assigned_to else 'Unassigned'
            
            notifications.append({
                'id': f'followup:{followup.id}',
                'type': 'followup',
                'title': 'Today Follow-up',
                'message': f"{followup.title} - {followup.lead.full_name} ({assigned_name})",
                'date': followup.followup_date.strftime('%d %b %Y'),
                'time': followup.followup_time.strftime('%I:%M %p') if followup.followup_time else '',
                'priority': 'high',
                'is_today': True,
                'url': reverse('lead_profile', args=[followup.lead.id])
            })

        # Upcoming exams (all)
        upcoming_exams = Exam.objects.filter(
            exam_date__gte=today,
            exam_date__lte=week_ahead
        ).select_related('batch', 'batch__trainer__user').order_by('exam_date')[:5]

        for exam in upcoming_exams:
            is_today = exam.exam_date == today
            is_tomorrow = exam.exam_date == tomorrow
            trainer_name = exam.batch.trainer.user.get_full_name() or exam.batch.trainer.user.username if exam.batch.trainer else 'No Trainer'

            notifications.append({
                'id': f'exam:{exam.id}',
                'type': 'exam',
                'title': f"{'Today' if is_today else 'Tomorrow' if is_tomorrow else 'Upcoming'} Exam",
                'message': f"{exam.title} - {exam.batch.batch_name} ({trainer_name})",
                'date': exam.exam_date.strftime('%d %b %Y'),
                'time': '',
                'priority': 'high' if is_today else 'medium' if is_tomorrow else 'low',
                'is_today': is_today,
                'url': f"/exams/{exam.id}/"
            })

        # Pending assignments (all)
        pending_assignments = Assignment.objects.filter(
            due_date__gte=today,
            due_date__lte=week_ahead
        ).select_related('batch', 'batch__trainer__user').order_by('due_date')[:5]

        for assignment in pending_assignments:
            is_today = assignment.due_date == today
            is_tomorrow = assignment.due_date == tomorrow
            trainer_name = assignment.batch.trainer.user.get_full_name() or assignment.batch.trainer.user.username if assignment.batch.trainer else 'No Trainer'

            notifications.append({
                'id': f'assignment:{assignment.id}',
                'type': 'assignment',
                'title': f"Assignment Due {'Today' if is_today else 'Tomorrow' if is_tomorrow else 'Soon'}",
                'message': f"{assignment.title} - {assignment.batch.batch_name} ({trainer_name})",
                'date': assignment.due_date.strftime('%d %b %Y'),
                'time': '',
                'priority': 'high' if is_today else 'medium' if is_tomorrow else 'low',
                'is_today': is_today,
                'url': f"/assignments/{assignment.id}/"
            })

    dismissed_keys = set(
        DismissedNotification.objects.filter(user=request.user).values_list(
            'notification_key',
            flat=True
        )
    )
    notifications = [
        notification
        for notification in notifications
        if notification.get('id') not in dismissed_keys
    ]

    # Sort by priority and date, then split by today's items before truncating.
    # This keeps CRE follow-ups due today visible even when many overdue items exist.
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    notifications.sort(key=lambda x: (priority_order.get(x['priority'], 3), x['date']))

    today_notifications = [n for n in notifications if n.get('is_today')]
    pending_notifications = [n for n in notifications if not n.get('is_today')]

    today_notifications = today_notifications[:10]
    pending_notifications = pending_notifications[:10]
    notifications = today_notifications + pending_notifications

    return JsonResponse({
        'today_notifications': today_notifications,
        'pending_notifications': pending_notifications,
        'notifications': notifications,
        'count': len(today_notifications),
        'total_count': len(notifications)
    })


@login_required
@require_POST
def dismiss_notification(request):
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = {}

    notification_key = (payload.get('notification_id') or '').strip()
    allowed_prefixes = ('followup:', 'exam:', 'assignment:')
    if not notification_key.startswith(allowed_prefixes):
        return JsonResponse({
            'success': False,
            'message': 'Invalid notification.'
        }, status=400)

    DismissedNotification.objects.get_or_create(
        user=request.user,
        notification_key=notification_key
    )

    return JsonResponse({'success': True})


@login_required
def global_search(request):
    """Global search across all entities"""
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'results': []})
    
    results = {
        'leads': [],
        'students': [],
        'users': [],
        'courses': [],
        'batches': [],
        'modules': [],
        'topics': []
    }
    
    # Search Leads
    lead_courses_prefetch = Prefetch(
        'lead_courses',
        queryset=LeadCourse.objects.select_related('course', 'department')
    )
    
    leads = StudentEnquiry.objects.select_related(
        'status', 'assigned__user', 'created_by'
    ).prefetch_related(lead_courses_prefetch).filter(
        Q(full_name__icontains=query) |
        Q(email__icontains=query) |
        Q(mobile__icontains=query)
    )[:10]
    
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.role == 'cre':
        leads = leads.filter(assigned=user_profile)
    
    for lead in leads:
        courses = [lc.course.course_name for lc in lead.lead_courses.all()]
        results['leads'].append({
            'id': lead.id,
            'name': lead.full_name,
            'mobile': lead.mobile,
            'email': lead.email,
            'status': lead.status.status_name if lead.status else 'N/A',
            'courses': ', '.join(courses) if courses else 'N/A',
            'url': f'/leads/{lead.id}/',
            'type': 'Lead'
        })
    
    # Search Students
    students = Student.objects.prefetch_related(
        'student_courses__course'
    ).filter(
        Q(student_name__icontains=query) |
        Q(email__icontains=query) |
        Q(mobile__icontains=query)
    )[:10]
    
    for student in students:
        courses = [sc.course.course_name for sc in student.student_courses.all()]
        results['students'].append({
            'id': student.id,
            'name': student.student_name,
            'mobile': student.mobile,
            'email': student.email,
            'status': student.status,
            'courses': ', '.join(courses) if courses else 'N/A',
            'url': f'/students/view/{student.id}/',
            'type': 'Student'
        })
    
    # Search Users
    users = User.objects.select_related('userprofile').filter(
        Q(username__icontains=query) |
        Q(email__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query)
    ).exclude(is_superuser=True)[:10]
    
    for user in users:
        role = user.userprofile.role if hasattr(user, 'userprofile') else 'N/A'
        results['users'].append({
            'id': user.id,
            'name': user.get_full_name() or user.username,
            'username': user.username,
            'email': user.email,
            'role': role,
            'url': f'/users/',
            'type': 'User'
        })
    
    # Search Courses
    courses = Course.objects.select_related('department').filter(
        Q(course_name__icontains=query) |
        Q(department__department_name__icontains=query)
    )[:10]
    
    for course in courses:
        results['courses'].append({
            'id': course.id,
            'name': course.course_name,
            'department': course.department.department_name if course.department else 'N/A',
            'fees': str(course.fees) if course.fees else 'N/A',
            'duration': f"{course.duration_months} months",
            'url': f'/course/',
            'type': 'Course'
        })
    
    # Search Batches
    batches = Batch.objects.select_related(
        'course', 'department', 'trainer__user'
    ).filter(
        Q(batch_name__icontains=query) |
        Q(course__course_name__icontains=query) |
        Q(department__department_name__icontains=query)
    )[:10]
    
    for batch in batches:
        trainer_name = batch.trainer.user.get_full_name() if batch.trainer else 'N/A'
        results['batches'].append({
            'id': batch.id,
            'name': batch.batch_name,
            'course': batch.course.course_name if batch.course else 'N/A',
            'trainer': trainer_name,
            'status': batch.status,
            'mode': batch.mode,
            'url': f'/batch/',
            'type': 'Batch'
        })
    
    # Search Modules
    modules = Module.objects.select_related('course').filter(
        Q(module_name__icontains=query) |
        Q(course__course_name__icontains=query)
    )[:10]
    
    for module in modules:
        results['modules'].append({
            'id': module.id,
            'name': module.module_name,
            'course': module.course.course_name if module.course else 'N/A',
            'url': f'/syllabus/',
            'type': 'Module'
        })
    
    # Search Topics
    topics = Topic.objects.select_related('module', 'module__course').filter(
        Q(topic_name__icontains=query) |
        Q(module__module_name__icontains=query) |
        Q(module__course__course_name__icontains=query)
    )[:10]
    
    for topic in topics:
        results['topics'].append({
            'id': topic.id,
            'name': topic.topic_name,
            'module': topic.module.module_name if topic.module else 'N/A',
            'course': topic.module.course.course_name if topic.module and topic.module.course else 'N/A',
            'url': f'/syllabus/',
            'type': 'Topic'
        })
    
    # Flatten results for display
    all_results = []
    for category, items in results.items():
        all_results.extend(items)
    
    return JsonResponse({
        'results': all_results,
        'total': len(all_results),
        'categories': {
            'leads': len(results['leads']),
            'students': len(results['students']),
            'users': len(results['users']),
            'courses': len(results['courses']),
            'batches': len(results['batches']),
            'modules': len(results['modules']),
            'topics': len(results['topics'])
        }
    })



@login_required
def question_paper(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    user_profile = request.user.userprofile

    if user_profile.role == "trainer" and exam.batch.trainer != user_profile:
        messages.error(request, "You can only manage question papers for your own batches.")
        return redirect('exam_list')

    question_paper = QuestionPaper.objects.filter(exam=exam).first()

    if request.method == 'POST':
        total_marks = request.POST.get('total_marks')
        part_names = request.POST.getlist('part_name[]')
        part_marks = request.POST.getlist('part_marks[]')
        part_questions = request.POST.getlist('part_questions[]')
        part_question_marks = request.POST.getlist('part_question_marks[]')
        part_question_data = request.POST.getlist('part_question_data[]')

        if not total_marks or not part_names:
            messages.error(request, "Total marks and at least one part are required.")
            return redirect('question_paper', exam_id=exam_id)

        if len(part_question_marks) < len(part_questions):
            part_question_marks.extend([''] * (len(part_questions) - len(part_question_marks)))
        if len(part_question_data) < len(part_questions):
            part_question_data.extend([''] * (len(part_questions) - len(part_question_data)))

        validated_parts = []
        for idx, (name, marks, questions, question_marks_raw, question_data_raw) in enumerate(zip(part_names, part_marks, part_questions, part_question_marks, part_question_data)):
            name = name.strip()
            questions = questions.strip()

            if not name or not marks or not questions:
                continue

            try:
                part_mark_value = int(marks)
            except (TypeError, ValueError):
                messages.error(request, f"Invalid marks provided for part {idx + 1}.")
                return redirect('question_paper', exam_id=exam_id)

            structured_questions = []
            if question_data_raw.strip():
                try:
                    parsed_question_data = json.loads(question_data_raw)
                except json.JSONDecodeError:
                    messages.error(request, f"Question format in part {name or idx + 1} is invalid.")
                    return redirect('question_paper', exam_id=exam_id)

                if not isinstance(parsed_question_data, list):
                    messages.error(request, f"Question format in part {name or idx + 1} is invalid.")
                    return redirect('question_paper', exam_id=exam_id)

                for question_entry in parsed_question_data:
                    question_text = str((question_entry or {}).get('text', '')).strip()
                    raw_mark = str((question_entry or {}).get('mark', '')).strip()
                    raw_subquestions = (question_entry or {}).get('subquestions', [])

                    if not question_text:
                        continue

                    try:
                        mark_value = int(raw_mark or 0)
                    except (TypeError, ValueError):
                        messages.error(request, f"Question marks in part {name or idx + 1} must be numbers.")
                        return redirect('question_paper', exam_id=exam_id)

                    if mark_value < 0:
                        messages.error(request, f"Question marks in part {name or idx + 1} cannot be negative.")
                        return redirect('question_paper', exam_id=exam_id)

                    subquestions = []
                    if raw_subquestions:
                        if not isinstance(raw_subquestions, list):
                            messages.error(request, f"Subquestions in part {name or idx + 1} are invalid.")
                            return redirect('question_paper', exam_id=exam_id)

                        for subquestion in raw_subquestions:
                            subquestion_text = str((subquestion or {}).get('text', '')).strip()
                            subquestion_mark_raw = str((subquestion or {}).get('mark', '')).strip()
                            if subquestion_text:
                                if not subquestion_mark_raw:
                                    messages.error(request, f"Each subquestion in part {name or idx + 1} must have a mark.")
                                    return redirect('question_paper', exam_id=exam_id)

                                try:
                                    subquestion_mark = int(subquestion_mark_raw)
                                except (TypeError, ValueError):
                                    messages.error(request, f"Subquestion marks in part {name or idx + 1} must be numbers.")
                                    return redirect('question_paper', exam_id=exam_id)

                                if subquestion_mark < 1:
                                    messages.error(request, f"Subquestion marks in part {name or idx + 1} must be at least 1.")
                                    return redirect('question_paper', exam_id=exam_id)

                                subquestions.append({
                                    'text': subquestion_text,
                                    'mark': subquestion_mark,
                                })

                    total_question_mark = mark_value + sum(subquestion['mark'] for subquestion in subquestions)
                    if total_question_mark < 1:
                        messages.error(request, f"Each question in part {name or idx + 1} must contribute at least 1 mark in total.")
                        return redirect('question_paper', exam_id=exam_id)

                    structured_questions.append({
                        'text': question_text,
                        'mark': mark_value,
                        'subquestions': subquestions,
                        'total_mark': total_question_mark,
                    })
            else:
                question_lines = [question.strip() for question in questions.splitlines() if question.strip()]
                question_mark_lines = [mark.strip() for mark in question_marks_raw.splitlines() if mark.strip()]

                if len(question_lines) != len(question_mark_lines):
                    messages.error(request, f"Each question in part {name or idx + 1} must have a mark.")
                    return redirect('question_paper', exam_id=exam_id)

                for question_text, mark in zip(question_lines, question_mark_lines):
                    try:
                        mark_value = int(mark)
                    except (TypeError, ValueError):
                        messages.error(request, f"Question marks in part {name or idx + 1} must be numbers.")
                        return redirect('question_paper', exam_id=exam_id)

                    if mark_value < 1:
                        messages.error(request, f"Question marks in part {name or idx + 1} must be at least 1.")
                        return redirect('question_paper', exam_id=exam_id)

                    structured_questions.append({
                        'text': question_text,
                        'mark': mark_value,
                        'subquestions': [],
                        'total_mark': mark_value,
                    })

            if not structured_questions:
                messages.error(request, f"At least one valid question is required in part {name or idx + 1}.")
                return redirect('question_paper', exam_id=exam_id)

            if sum(question.get('total_mark', question['mark']) for question in structured_questions) > part_mark_value:
                messages.error(request, f"Total question marks in part {name or idx + 1} cannot exceed part marks.")
                return redirect('question_paper', exam_id=exam_id)

            validated_parts.append({
                'name': name,
                'marks': part_mark_value,
                'questions': '\n'.join(question['text'] for question in structured_questions),
                'question_marks': '\n'.join(str(question['mark']) for question in structured_questions),
                'question_data': structured_questions,
                'order': idx,
            })

        # Create or update question paper
        if question_paper:
            question_paper.total_marks = total_marks
            question_paper.save()
            # Delete existing parts
            question_paper.parts.all().delete()
        else:
            question_paper = QuestionPaper.objects.create(
                exam=exam,
                total_marks=total_marks
            )

        # Create parts
        for part in validated_parts:
            QuestionPart.objects.create(
                question_paper=question_paper,
                part_name=part['name'],
                marks=part['marks'],
                questions=part['questions'],
                question_marks=part['question_marks'],
                question_data=part['question_data'],
                order=part['order']
            )

        messages.success(request, "Question paper saved successfully.")
        return redirect('question_paper', exam_id=exam_id)

    context = {
        'exam': exam,
        'question_paper': question_paper,
    }
    return render(request, 'question_paper.html', context)


@login_required
def download_question_paper(request, exam_id):
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    def alpha_index(index):
        value = index + 1
        output = ''
        while value > 0:
            value -= 1
            output = chr(97 + (value % 26)) + output
            value //= 26
        return output

    def alpha_label(index):
        value = index
        output = ''
        while value > 0:
            value -= 1
            output = chr(65 + (value % 26)) + output
            value //= 26
        return output
    
    exam = get_object_or_404(Exam, id=exam_id)
    question_paper = get_object_or_404(QuestionPaper, exam=exam)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="question_paper_{exam.title}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=17,
        textColor=colors.black,
        spaceAfter=8,
        alignment=TA_CENTER,
        fontName='Times-Bold'
    )

    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_CENTER,
        spaceAfter=14,
        fontName='Times-Bold'
    )

    part_style = ParagraphStyle(
        'PartStyle',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.black,
        alignment=TA_CENTER,
        fontName='Times-Bold'
    )
    
    question_style = ParagraphStyle(
        'QuestionStyle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        leftIndent=0
    )

    mark_style = ParagraphStyle(
        'MarkStyle',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_RIGHT,
        fontName='Times-Bold',
        textColor=colors.black
    )
    
    # Title
    elements.append(Paragraph(exam.title, title_style))
    elements.append(Paragraph("Question Paper", subtitle_style))
    elements.append(Spacer(1, 0.08*inch))
    
    # Header info
    header_data = [
        ['Batch:', exam.batch.batch_name, 'Date:', exam.exam_date.strftime('%d %b, %Y')],
        ['Total Marks:', str(question_paper.total_marks), 'Duration:', '3 Hours']
    ]
    
    header_table = Table(header_data, colWidths=[1.35*inch, 2.15*inch, 1.35*inch, 2.15*inch])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Times-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Times-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(header_table)
    elements.append(Paragraph("<b>Note:</b> Attempt all sections.", styles['Normal']))
    elements.append(Spacer(1, 0.18*inch))
    
    # Parts and questions
    question_number = 1
    for part_index, part in enumerate(question_paper.parts.all(), start=1):
        part_header = Table(
            [['', Paragraph(f"Section {alpha_label(part_index)}: {part.part_name}", part_style), Paragraph(f"({part.marks} Marks)", mark_style)]],
            colWidths=[1.1*inch, 4.7*inch, 1.0*inch]
        )
        part_header.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.HexColor('#94a3b8')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(part_header)
        elements.append(Spacer(1, 0.12*inch))

        for item in part.question_items:
            question_table = Table(
                [[Paragraph(f"{question_number}. {item['text']}", question_style), Paragraph(f"({item['mark']} Marks)" if item['mark'] else '', mark_style)]],
                colWidths=[5.8*inch, 1.0*inch]
            )
            question_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(question_table)

            for sub_idx, subquestion in enumerate(item.get('subquestions', [])):
                subquestion_table = Table(
                    [[Paragraph(f"{alpha_index(sub_idx)}. {subquestion['text']}", question_style), Paragraph(f"({subquestion.get('mark')})" if subquestion.get('mark') else '', mark_style)]],
                    colWidths=[5.55*inch, 1.25*inch]
                )
                subquestion_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 28),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(subquestion_table)
            question_number += 1
        
        elements.append(Spacer(1, 0.2*inch))
    
    doc.build(elements)
    return response

#===================EMAIL==============================

